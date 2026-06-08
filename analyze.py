"""
Stock Portfolio Analyzer
Reads a portfolio CSV and produces an HTML dashboard + Excel report.

Supported brokerage export formats (auto-detected):
  - Fidelity: "Symbol", "Quantity", "Cost Basis Total"
  - Schwab:   "Symbol", "Quantity", "Cost Basis"
  - Generic:  "Symbol", "Shares", "Cost Basis"

Run: python analyze.py portfolio.csv [--cash 4900] [--age 22]
"""

import sys
import os
import math
import argparse
import json
from datetime import datetime
from pathlib import Path

import pandas as pd
import yfinance as yf
from jinja2 import Template
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ---------------------------------------------------------------------------
# Investor profile defaults (overridden by CLI args)
# ---------------------------------------------------------------------------
DEFAULT_CASH      = 4900
DEFAULT_AGE       = 22
# High risk tolerance: weight toward growth/speculative picks
RISK_PROFILE      = "aggressive"

# ---------------------------------------------------------------------------
# S&P 500 sector benchmark weights (approximate, as of 2025)
# ---------------------------------------------------------------------------
SP500_SECTOR_WEIGHTS = {
    "Technology": 29.0,
    "Healthcare": 12.5,
    "Financials": 13.0,
    "Consumer Cyclical": 10.5,
    "Communication Services": 8.5,
    "Industrials": 8.5,
    "Consumer Defensive": 6.0,
    "Energy": 3.5,
    "Utilities": 2.5,
    "Real Estate": 2.5,
    "Basic Materials": 2.5,
}

# ---------------------------------------------------------------------------
# Stock universe — S&P picks + high-growth non-S&P picks per sector
# Each entry: (symbol, in_sp500, risk_tier)
#   risk_tier: "low" | "medium" | "high" | "speculative"
# ---------------------------------------------------------------------------
SECTOR_UNIVERSE = {
    "Technology": [
        ("MSFT",  True,  "low"),
        ("AAPL",  True,  "low"),
        ("NVDA",  True,  "medium"),
        ("AVGO",  True,  "medium"),
        ("AMD",   True,  "medium"),
        ("CRM",   True,  "medium"),
        ("ORCL",  True,  "low"),
        # Non-S&P high-growth
        ("PLTR",  False, "high"),
        ("RBLX",  False, "high"),
        ("SOUN",  False, "speculative"),
        ("BBAI",  False, "speculative"),
        ("IONQ",  False, "speculative"),   # quantum computing
        ("RGTI",  False, "speculative"),   # quantum computing
    ],
    "Healthcare": [
        ("UNH",  True,  "low"),
        ("LLY",  True,  "medium"),
        ("ABBV", True,  "low"),
        ("MRK",  True,  "low"),
        ("TMO",  True,  "low"),
        # Non-S&P biotech/growth
        ("RXRX", False, "speculative"),   # AI drug discovery
        ("EXAS", False, "high"),
        ("PACB", False, "speculative"),
        ("VERA", False, "speculative"),
        ("ARWR", False, "high"),
    ],
    "Financials": [
        ("JPM",  True,  "low"),
        ("V",    True,  "low"),
        ("MA",   True,  "low"),
        ("GS",   True,  "medium"),
        # Fintech / non-S&P
        ("AFRM", False, "speculative"),   # buy-now-pay-later
        ("UPST", False, "speculative"),   # AI lending
        ("SQ",   False, "high"),
        ("NU",   False, "high"),          # Latin America fintech
        ("HOOD", False, "high"),
    ],
    "Consumer Cyclical": [
        ("AMZN", True,  "low"),
        ("TSLA", True,  "high"),
        ("HD",   True,  "low"),
        ("BKNG", True,  "medium"),
        # High-growth non-S&P
        ("CELH", False, "high"),          # energy drinks
        ("ONON", False, "high"),          # On Running shoes
        ("DKNG", False, "high"),          # sports betting
        ("DASH", False, "high"),          # DoorDash
    ],
    "Communication Services": [
        ("GOOGL", True,  "low"),
        ("META",  True,  "medium"),
        ("NFLX",  True,  "medium"),
        ("DIS",   True,  "medium"),
        # Non-S&P
        ("SPOT",  False, "high"),
        ("SNAP",  False, "speculative"),
        ("RDFN",  False, "speculative"),
    ],
    "Industrials": [
        ("CAT",  True,  "low"),
        ("HON",  True,  "low"),
        ("RTX",  True,  "low"),
        ("GE",   True,  "medium"),
        # Defense tech / non-S&P
        ("RKLB", False, "speculative"),   # Rocket Lab
        ("ASTS", False, "speculative"),   # satellite broadband
        ("ACHR", False, "speculative"),   # eVTOL / air taxi
        ("JOBY", False, "speculative"),   # eVTOL
        ("KTOS", False, "high"),          # defense drones
    ],
    "Consumer Defensive": [
        ("WMT",  True,  "low"),
        ("COST", True,  "low"),
        ("PG",   True,  "low"),
        ("KO",   True,  "low"),
    ],
    "Energy": [
        ("XOM",  True,  "low"),
        ("CVX",  True,  "low"),
        ("COP",  True,  "medium"),
        # Clean energy / non-S&P
        ("FSLR", True,  "high"),
        ("ENPH", True,  "high"),
        ("ARRY", False, "speculative"),
        ("STEM", False, "speculative"),
        ("BE",   False, "high"),          # Bloom Energy
    ],
    "Utilities": [
        ("NEE",  True,  "low"),
        ("VST",  True,  "medium"),        # power for AI data centers
        ("CEG",  True,  "medium"),        # nuclear / AI power
        ("NRG",  False, "medium"),
    ],
    "Real Estate": [
        ("PLD",  True,  "low"),
        ("EQIX", True,  "low"),
        ("AMT",  True,  "low"),
        ("IIPR", False, "high"),          # cannabis REITs
    ],
    "Basic Materials": [
        ("LIN",  True,  "low"),
        ("FCX",  True,  "medium"),
        ("NEM",  True,  "medium"),
        ("MP",   False, "high"),          # rare earth mining
        ("SIGA", False, "speculative"),
        ("LAC",  False, "speculative"),   # lithium mining
        ("LIXT", False, "speculative"),
    ],
}

# Sectors beyond the standard 11 — pure high-growth themes (non-S&P heavy)
THEMATIC_PICKS = {
    "AI & Data Infrastructure": [
        ("SMCI", False, "high"),
        ("DELL", True,  "medium"),
        ("CDNS", True,  "medium"),
        ("ARM",  False, "high"),
        ("CEVA", False, "high"),
        ("AI",   False, "high"),          # C3.ai
        ("BIGB", False, "speculative"),
    ],
    "Cybersecurity": [
        ("CRWD", True,  "medium"),
        ("PANW", True,  "medium"),
        ("S",    False, "high"),          # SentinelOne
        ("ZS",   False, "high"),          # Zscaler
        ("OKTA", True,  "high"),            # Okta identity security
        ("DDOG", True,  "medium"),
    ],
    "Biotech / Gene Editing": [
        ("CRSP", False, "speculative"),
        ("BEAM", False, "speculative"),
        ("NTLA", False, "speculative"),
        ("EDIT", False, "speculative"),
        ("FATE", False, "speculative"),
    ],
    "Space & Defense Tech": [
        ("RKLB", False, "speculative"),
        ("ASTS", False, "speculative"),
        ("SPCE", False, "speculative"),
        ("KTOS", False, "high"),
        ("LUNR", False, "speculative"),
    ],
}

RISK_TIER_COLOR = {
    "low":         "#34d399",   # green
    "medium":      "#60a5fa",   # blue
    "high":        "#fbbf24",   # amber
    "speculative": "#f87171",   # red
}

RISK_TIER_LABEL = {
    "low":         "Low Risk",
    "medium":      "Medium Risk",
    "high":        "High Risk",
    "speculative": "Speculative",
}

# Cash allocation splits by risk profile (must sum to 1.0)
# aggressive: tilt toward high/speculative
CASH_ALLOCATION = {
    "aggressive": {"low": 0.15, "medium": 0.25, "high": 0.35, "speculative": 0.25},
    "moderate":   {"low": 0.30, "medium": 0.35, "high": 0.25, "speculative": 0.10},
    "conservative":{"low": 0.50,"medium": 0.35, "high": 0.15, "speculative": 0.00},
}

# ---------------------------------------------------------------------------
# Portfolio loading
# ---------------------------------------------------------------------------

def load_portfolio(filepath: str) -> pd.DataFrame:
    """Load and normalize brokerage CSV exports into (symbol, shares, cost_basis)."""
    df = pd.read_csv(filepath)
    df.columns = df.columns.str.strip()

    col_map = {}

    # Symbol column
    for c in ["Symbol", "Ticker", "Stock"]:
        if c in df.columns:
            col_map["symbol"] = c
            break

    # Shares column
    for c in ["Shares", "Quantity", "Qty", "Units"]:
        if c in df.columns:
            col_map["shares"] = c
            break

    # Cost basis column
    for c in ["Cost Basis", "Cost Basis Total", "Total Cost", "Avg Cost"]:
        if c in df.columns:
            col_map["cost"] = c
            break

    if "symbol" not in col_map or "shares" not in col_map:
        raise ValueError(
            "Could not find Symbol/Shares columns. "
            "Expected headers like 'Symbol', 'Quantity', 'Shares'."
        )

    result = pd.DataFrame()
    result["symbol"] = df[col_map["symbol"]].astype(str).str.strip().str.upper()
    result["shares"] = pd.to_numeric(df[col_map["shares"]], errors="coerce")

    if "cost" in col_map:
        result["cost_basis"] = pd.to_numeric(
            df[col_map["cost"]].astype(str).str.replace(r"[$,]", "", regex=True),
            errors="coerce",
        )
    else:
        result["cost_basis"] = None

    # Drop blanks / cash rows
    result = result[result["symbol"].str.match(r"^[A-Z]{1,5}$")]
    result = result.dropna(subset=["shares"])
    result = result[result["shares"] > 0]
    return result.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

ETF_UNIVERSE = {
    "Broad Market": [
        ("VOO",  "Vanguard S&P 500 ETF",           "Core S&P 500 exposure at rock-bottom cost. The simplest long-term hold."),
        ("VTI",  "Vanguard Total Stock Market ETF", "Every US-listed stock in one fund — ultimate domestic diversification."),
        ("SCHB", "Schwab US Broad Market ETF",      "Cheapest broad-market option; nearly identical to VTI."),
        ("IVV",  "iShares Core S&P 500 ETF",        "S&P 500 alternative to SPY with a lower expense ratio."),
        ("SPY",  "SPDR S&P 500 ETF Trust",          "Most liquid US equity ETF — tight spreads, ideal for active trading."),
    ],
    "Growth": [
        ("QQQ",  "Invesco NASDAQ-100 ETF",          "Top 100 NASDAQ names — heavy tech/AI tilt, strong long-term returns."),
        ("VUG",  "Vanguard Growth ETF",             "Large-cap US growth at low cost. Broad growth exposure."),
        ("SCHG", "Schwab US Large-Cap Growth ETF",  "One of the cheapest growth ETFs available."),
        ("MGK",  "Vanguard Mega Cap Growth ETF",    "Concentrated in the biggest growth names (Mag 7 heavy)."),
    ],
    "International": [
        ("VEA",  "Vanguard FTSE Developed Markets ETF","Europe, Japan, Australia — diversifies away from US concentration."),
        ("VWO",  "Vanguard FTSE Emerging Markets ETF", "China, India, Brazil exposure. Higher risk, higher long-run potential."),
        ("IEMG", "iShares Core MSCI Emerging Mkts ETF","Cheaper EM alternative with broader country coverage."),
        ("EFA",  "iShares MSCI EAFE ETF",               "Developed markets ex-US/Canada. Classic international core."),
    ],
    "Thematic / Sector": [
        ("SOXX", "iShares Semiconductor ETF",       "Pure semiconductor exposure — cyclical but strong secular tailwind."),
        ("SMH",  "VanEck Semiconductor ETF",         "Top 25 semis, more concentrated than SOXX. Higher conviction play."),
        ("BOTZ", "Global X Robotics & AI ETF",      "Robotics, automation, and AI hardware beneficiaries."),
        ("ICLN", "iShares Global Clean Energy ETF", "Solar, wind, and clean energy — long-term energy transition bet."),
    ],
}


def fetch_stock_data(symbols: list[str]) -> dict:
    """Fetch price, sector, key metrics, beta, and growth for a list of tickers.
    Automatically detects ETFs and fetches ETF-specific fields instead of stock metrics."""
    import time
    print(f"  Fetching data for {len(symbols)} tickers...")
    data = {}
    for i, sym in enumerate(symbols):
        if i > 0:
            time.sleep(0.4)  # pace requests to avoid rate limiting
        info = None
        for attempt in range(4):
            try:
                info = yf.Ticker(sym).info
                if not info or (info.get("regularMarketPrice") is None and
                                info.get("currentPrice") is None and
                                info.get("previousClose") is None and
                                info.get("navPrice") is None):
                    raise ValueError("empty info returned")
                break
            except Exception as e:
                if attempt < 3:
                    wait = [5, 15, 30][min(attempt, 2)]
                    time.sleep(wait)
                    continue
                info = None
                print(f"    Warning: could not fetch {sym}: {e}")
                data[sym] = {"name": sym, "sector": "Unknown", "price": 0,
                             "auto_risk": "medium", "is_etf": False}
        if info is None:
            continue
        try:
            price = (info.get("currentPrice") or info.get("regularMarketPrice")
                     or info.get("previousClose") or info.get("navPrice") or 0)
            beta  = info.get("beta")

            # ── ETF / Mutual Fund path ──
            quote_type = (info.get("quoteType") or "").upper()
            if quote_type in ("ETF", "MUTUALFUND"):
                expense_ratio    = (info.get("expenseRatio") or 0) * 100
                etf_yield        = (info.get("yield") or info.get("dividendYield") or 0) * 100
                aum              = info.get("totalAssets") or 0
                etf_category     = info.get("category") or info.get("fundFamily") or "ETF"
                ytd_return       = (info.get("ytdReturn") or 0) * 100
                three_year_ret   = (info.get("threeYearAverageReturn") or 0) * 100
                five_year_ret    = (info.get("fiveYearAverageReturn") or 0) * 100

                if beta is None:
                    auto_risk = "medium"
                elif beta > 1.5:
                    auto_risk = "high"
                elif beta > 1.1:
                    auto_risk = "medium"
                else:
                    auto_risk = "low"

                data[sym] = {
                    "name":             info.get("shortName") or info.get("longName") or sym,
                    "sector":           etf_category,
                    "price":            price,
                    "is_etf":           True,
                    "expense_ratio":    expense_ratio,
                    "etf_yield":        etf_yield,
                    "aum":              aum,
                    "etf_category":     etf_category,
                    "ytd_return":       ytd_return,
                    "three_year_ret":   three_year_ret,
                    "five_year_ret":    five_year_ret,
                    "beta":             beta,
                    "auto_risk":        auto_risk,
                    "market_cap":       aum,
                    "in_sp500":         False,
                    "dividend_yield":   etf_yield,
                    # Stock fields null for ETFs
                    "pe": None, "forward_pe": None, "pb": None,
                    "eps_growth": None, "rev_growth": None, "growth_score": None,
                    "analyst_target": None, "recommendation": None,
                    "52w_high": info.get("fiftyTwoWeekHigh"),
                    "52w_low":  info.get("fiftyTwoWeekLow"),
                }
                continue

            # ── Individual stock path ──
            market_cap = info.get("marketCap") or 0
            eps_growth = (info.get("earningsGrowth") or 0) * 100
            rev_growth = (info.get("revenueGrowth") or 0) * 100

            analyst_target = info.get("targetMeanPrice")
            upside_pct = ((analyst_target - price) / price * 100) if analyst_target and price else 0
            growth_score = round(
                eps_growth * 0.4 + rev_growth * 0.3 + min(upside_pct, 60) * 0.3, 1
            )

            if beta is None:
                auto_risk = "medium"
            elif beta > 2.0 or market_cap < 2e9:
                auto_risk = "speculative"
            elif beta > 1.4 or market_cap < 10e9:
                auto_risk = "high"
            elif beta > 0.9:
                auto_risk = "medium"
            else:
                auto_risk = "low"

            data[sym] = {
                "name":           info.get("shortName", sym),
                "sector":         info.get("sector", "Unknown"),
                "price":          price,
                "is_etf":         False,
                "pe":             info.get("trailingPE"),
                "forward_pe":     info.get("forwardPE"),
                "pb":             info.get("priceToBook"),
                "dividend_yield": (info.get("dividendYield") or 0) * 100,
                "eps_growth":     eps_growth,
                "rev_growth":     rev_growth,
                "growth_score":   growth_score,
                "analyst_target": analyst_target,
                "recommendation": info.get("recommendationKey", "").replace("_", " ").title(),
                "52w_high":       info.get("fiftyTwoWeekHigh"),
                "52w_low":        info.get("fiftyTwoWeekLow"),
                "market_cap":     market_cap,
                "beta":           beta,
                "auto_risk":      auto_risk,
                "in_sp500":       market_cap > 5e9,
            }
        except Exception as e:
            print(f"    Warning: could not fetch {sym}: {e}")
            data[sym] = {"name": sym, "sector": "Unknown", "price": 0,
                         "auto_risk": "medium", "is_etf": False}
    return data


def fetch_chart_data(symbols: list[str]) -> dict:
    """
    Fetch price history for interactive charts across multiple timeframes.
    Returns {symbol: {intraday: [{t,c},...], history: [{t,c},...], alltime: [{t,c},...]}}.
    intraday  = 1-day 5-min bars   → used for 1D tab
    history   = 2-year daily bars  → sliced in JS for 5D / 1M / 3M / YTD / 1Y tabs
    alltime   = max monthly bars   → used for ALL tab
    """
    import json as _json
    print(f"  Fetching chart data for {len(symbols)} tickers (3 timeframes each)...")

    def _to_points(hist):
        pts = []
        for ts, row in hist.iterrows():
            try:
                t_ms = int(ts.timestamp() * 1000)
                c = float(row.get("Close", 0) or 0)
                if c > 0:
                    pts.append({"t": t_ms, "c": round(c, 2)})
            except Exception:
                pass
        return pts

    chart_data = {}
    for sym in symbols:
        chart_data[sym] = {"intraday": [], "history": [], "alltime": []}
        try:
            tk = yf.Ticker(sym)
            try:
                chart_data[sym]["intraday"] = _to_points(
                    tk.history(period="1d", interval="5m", auto_adjust=True)
                )
            except Exception:
                pass
            try:
                chart_data[sym]["history"] = _to_points(
                    tk.history(period="2y", interval="1d", auto_adjust=True)
                )
            except Exception:
                pass
            try:
                chart_data[sym]["alltime"] = _to_points(
                    tk.history(period="max", interval="1mo", auto_adjust=True)
                )
            except Exception:
                pass
        except Exception as e:
            print(f"    Warning: chart data failed for {sym}: {e}")

    return chart_data


def fetch_news(symbols: list[str], max_per_symbol: int = 6) -> dict:
    """
    Fetch recent news headlines for a list of tickers via yfinance.
    Handles the nested content structure introduced in yfinance 0.2.x+.
    Returns {symbol: [article_dict, ...]} sorted newest-first.
    """
    from datetime import datetime as dt, timezone

    def _parse_age(pub_date_str: str) -> tuple[str, int]:
        """Parse ISO pubDate string → (human label, unix timestamp)."""
        try:
            pub = dt.fromisoformat(pub_date_str.replace("Z", "+00:00"))
            ts  = int(pub.timestamp())
            now = dt.now(timezone.utc)
            delta = now - pub
            total_mins = int(delta.total_seconds() // 60)
            if total_mins < 60:
                return (f"{total_mins}m ago" if total_mins > 1 else "Just now"), ts
            elif total_mins < 1440:
                return f"{total_mins // 60}h ago", ts
            elif delta.days == 1:
                return "Yesterday", ts
            else:
                return f"{delta.days}d ago", ts
        except Exception:
            return "", 0

    def _best_thumb(resolutions: list) -> str | None:
        """Pick the smallest thumbnail that's at least 120px wide."""
        candidates = [r for r in resolutions if r.get("width", 0) >= 120]
        if not candidates:
            candidates = resolutions
        if not candidates:
            return None
        return min(candidates, key=lambda r: r.get("width", 9999)).get("url")

    news_data = {}
    for sym in symbols:
        try:
            raw = yf.Ticker(sym).news or []
            articles = []
            seen_ids = set()

            for item in raw:
                if len(articles) >= max_per_symbol:
                    break

                # New yfinance structure: item = {"id": ..., "content": {...}}
                content = item.get("content") or item  # fall back to flat dict
                uid = content.get("id") or item.get("id", "")
                if uid in seen_ids:
                    continue
                seen_ids.add(uid)

                title = content.get("title") or item.get("title", "")
                if not title:
                    continue

                summary = (content.get("summary") or "").strip()

                # Publisher
                provider = content.get("provider") or {}
                publisher = (provider.get("displayName")
                             or item.get("publisher", "")
                             or "Yahoo Finance")

                # Link — prefer clickThroughUrl, then canonicalUrl, then legacy "link"
                link = "#"
                for key in ("clickThroughUrl", "canonicalUrl"):
                    obj = content.get(key)
                    if obj and obj.get("url"):
                        link = obj["url"]
                        break
                if link == "#":
                    link = item.get("link", "#")

                # Timestamp / age
                pub_date = content.get("pubDate") or content.get("displayTime") or ""
                age_str, ts = _parse_age(pub_date) if pub_date else ("", 0)
                if not ts:
                    # Legacy fallback
                    ts = item.get("providerPublishTime", 0)
                    if ts:
                        age_str, _ = _parse_age(
                            dt.fromtimestamp(ts, tz=timezone.utc).isoformat()
                        )

                # Thumbnail
                thumb = None
                try:
                    thumb_obj = content.get("thumbnail") or item.get("thumbnail") or {}
                    resolutions = thumb_obj.get("resolutions") or []
                    thumb = _best_thumb(resolutions)
                    if not thumb and thumb_obj.get("originalUrl"):
                        thumb = thumb_obj["originalUrl"]
                except Exception:
                    pass

                articles.append({
                    "title":     title,
                    "summary":   summary[:180] + ("…" if len(summary) > 180 else ""),
                    "publisher": publisher,
                    "link":      link,
                    "age":       age_str,
                    "timestamp": ts,
                    "thumbnail": thumb,
                })

            articles.sort(key=lambda a: a["timestamp"], reverse=True)
            news_data[sym] = articles

        except Exception as e:
            print(f"    Warning: news fetch failed for {sym}: {e}")
            news_data[sym] = []

    return news_data


def fetch_economy_news(max_articles: int = 40) -> list:
    """
    Fetch macro/economy-focused news by pulling headlines from a set of
    broad market / macro indicators and ETFs, then deduplicating and
    returning a flat list sorted newest-first.
    """
    MACRO_TICKERS = [
        "^GSPC",   # S&P 500 — broad market headlines
        "^DJI",    # Dow Jones
        "^TNX",    # 10-Year Treasury yield
        "^VIX",    # Volatility / fear index
        "GC=F",    # Gold futures
        "CL=F",    # Crude oil futures
        "TLT",     # Long-duration bonds (rates)
        "DX-Y.NYB",# US Dollar index
    ]
    from datetime import datetime as dt, timezone

    def _parse_age(pub_date_str: str) -> tuple[str, int]:
        try:
            pub = dt.fromisoformat(pub_date_str.replace("Z", "+00:00"))
            ts  = int(pub.timestamp())
            now = dt.now(timezone.utc)
            delta = now - pub
            total_mins = int(delta.total_seconds() // 60)
            if total_mins < 60:
                return (f"{total_mins}m ago" if total_mins > 1 else "Just now"), ts
            elif total_mins < 1440:
                return f"{total_mins // 60}h ago", ts
            elif delta.days == 1:
                return "Yesterday", ts
            else:
                return f"{delta.days}d ago", ts
        except Exception:
            return "", 0

    def _best_thumb(resolutions: list) -> str | None:
        candidates = [r for r in resolutions if r.get("width", 0) >= 120]
        if not candidates:
            candidates = resolutions
        if not candidates:
            return None
        return min(candidates, key=lambda r: r.get("width", 9999)).get("url")

    seen_ids  = set()
    all_articles = []

    for sym in MACRO_TICKERS:
        try:
            raw = yf.Ticker(sym).news or []
            for item in raw:
                content = item.get("content") or item
                uid = content.get("id") or item.get("id", "")
                # Also deduplicate by title to avoid repeats across tickers
                title = content.get("title") or item.get("title", "")
                dedup_key = uid or title.lower()[:60]
                if not title or dedup_key in seen_ids:
                    continue
                seen_ids.add(dedup_key)

                summary = (content.get("summary") or "").strip()

                provider = content.get("provider") or {}
                publisher = (provider.get("displayName")
                             or item.get("publisher", "")
                             or "Yahoo Finance")

                link = "#"
                for key in ("clickThroughUrl", "canonicalUrl"):
                    obj = content.get(key)
                    if obj and obj.get("url"):
                        link = obj["url"]
                        break
                if link == "#":
                    link = item.get("link", "#")

                pub_date = content.get("pubDate") or content.get("displayTime") or ""
                age_str, ts = _parse_age(pub_date) if pub_date else ("", 0)
                if not ts:
                    ts = item.get("providerPublishTime", 0)
                    if ts:
                        age_str, _ = _parse_age(
                            dt.fromtimestamp(ts, tz=timezone.utc).isoformat()
                        )

                thumb = None
                try:
                    thumb_obj = content.get("thumbnail") or item.get("thumbnail") or {}
                    resolutions = thumb_obj.get("resolutions") or []
                    thumb = _best_thumb(resolutions)
                    if not thumb and thumb_obj.get("originalUrl"):
                        thumb = thumb_obj["originalUrl"]
                except Exception:
                    pass

                all_articles.append({
                    "title":     title,
                    "summary":   summary[:200] + ("…" if len(summary) > 200 else ""),
                    "publisher": publisher,
                    "link":      link,
                    "age":       age_str,
                    "timestamp": ts,
                    "thumbnail": thumb,
                })
        except Exception as e:
            print(f"    Warning: economy news fetch failed for {sym}: {e}")

    all_articles.sort(key=lambda a: a["timestamp"], reverse=True)
    return all_articles[:max_articles]


def fetch_etf_recs(held_symbols: set, max_per_category: int = 2) -> list[dict]:
    """
    Score the ETF_UNIVERSE and return the top `max_per_category` picks per
    category, excluding any ETFs the user already holds.
    Scoring: 40% expense ratio, 35% 3-year return, 15% AUM, 10% yield.
    """
    import math as _math

    # Flatten universe, skip held
    candidates = []
    sym_to_meta = {}
    for category, etfs in ETF_UNIVERSE.items():
        for sym, name, description in etfs:
            if sym.upper() in {s.upper() for s in held_symbols}:
                continue
            candidates.append(sym)
            sym_to_meta[sym] = {"category": category, "name": name, "description": description}

    if not candidates:
        return []

    # Batch fetch
    tk = yf.Tickers(" ".join(candidates))
    results = []

    for sym in candidates:
        meta = sym_to_meta[sym]
        try:
            info = tk.tickers[sym].info
            price         = (info.get("navPrice") or info.get("regularMarketPrice")
                             or info.get("previousClose") or 0)
            expense_ratio = (info.get("expenseRatio") or 0) * 100
            etf_yield     = (info.get("yield") or info.get("dividendYield") or 0) * 100
            aum_raw       = info.get("totalAssets") or 0
            aum_b         = aum_raw / 1e9  # billions
            three_yr      = (info.get("threeYearAverageReturn") or 0) * 100
            five_yr       = (info.get("fiveYearAverageReturn") or 0) * 100
            ytd           = (info.get("ytdReturn") or 0) * 100
            beta          = info.get("beta")
            etf_category  = info.get("category") or ""

            # ── Score ──
            er_score  = max(0.0, 100.0 - (expense_ratio or 0.5) * 100.0)
            ret_score = min(100.0, max(0.0, (three_yr or 8.0) / 20.0 * 100.0))
            aum_score = min(100.0, max(0.0,
                _math.log10(max(0.1, aum_b)) / _math.log10(500) * 100.0))
            yld_score = min(100.0, (etf_yield or 0) / 5.0 * 100.0)
            score = round(0.40 * er_score + 0.35 * ret_score
                          + 0.15 * aum_score + 0.10 * yld_score, 1)

            def _fmt_aum(b):
                if b >= 100:  return f"${b:,.0f}B"
                if b >= 1:    return f"${b:.1f}B"
                return f"${b*1000:.0f}M"

            results.append({
                "symbol":       sym,
                "name":         meta["name"],
                "description":  meta["description"],
                "category":     meta["category"],
                "etf_category": etf_category,
                "price":        price,
                "expense_ratio":expense_ratio,
                "etf_yield":    etf_yield,
                "aum":          aum_raw,
                "aum_label":    _fmt_aum(aum_b),
                "three_yr":     three_yr,
                "five_yr":      five_yr,
                "ytd":          ytd,
                "beta":         beta,
                "score":        score,
            })
        except Exception as e:
            print(f"    Warning: ETF rec fetch failed for {sym}: {e}")

    # Sort within each category, keep top N
    from collections import defaultdict
    by_cat = defaultdict(list)
    for r in results:
        by_cat[r["category"]].append(r)

    etf_picks = []
    for category in ETF_UNIVERSE:           # preserve display order
        picks = sorted(by_cat.get(category, []), key=lambda x: x["score"], reverse=True)
        etf_picks.extend(picks[:max_per_category])

    return etf_picks


def fetch_etf_holdings(symbols: list[str]) -> dict:
    """Fetch top-10 holdings for each ETF symbol via yfinance funds_data.
    Returns {sym: [{symbol, name, pct}, ...]}"""
    import time
    holdings = {}
    for sym in symbols:
        try:
            df = yf.Ticker(sym).funds_data.top_holdings
            if df is not None and not df.empty:
                # Symbol is the index; columns are 'Name' and 'Holding Percent'
                rows = []
                for ticker, row in df.head(10).iterrows():
                    name    = str(row.get("Name", ticker)).strip()
                    raw_pct = row.get("Holding Percent", 0) or 0
                    # Values are decimal (0.07 = 7%); multiply by 100
                    pct = round(float(raw_pct) * 100, 2)
                    rows.append({"symbol": str(ticker).strip(), "name": name, "pct": pct})
                holdings[sym] = rows
            else:
                holdings[sym] = []
        except Exception:
            holdings[sym] = []
        time.sleep(0.25)
    return holdings


def fetch_sp500_sector_data() -> dict:
    """
    Build a per-sector map of every S&P 500 stock with daily % change and
    market cap, used to power the sector drill-down treemap.
    Returns: { "Information Technology": [{t, n, s, c, m}, ...], ... }
    Keys: t=ticker, n=name, s=sub-industry, c=change_pct, m=market_cap
    """
    import concurrent.futures

    # ── 1. S&P 500 constituents from Wikipedia ──
    try:
        import requests, io
        headers = {"User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        )}
        resp   = requests.get(
            "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies",
            headers=headers, timeout=15)
        resp.raise_for_status()
        tables = pd.read_html(io.StringIO(resp.text), header=0)
        sp500  = tables[0][["Symbol", "Security",
                             "GICS Sector", "GICS Sub-Industry"]].copy()
        sp500.columns = ["ticker", "name", "sector", "sub_industry"]
        sp500["ticker"] = sp500["ticker"].str.replace(".", "-", regex=False)
        tickers = sp500["ticker"].tolist()
        print(f"    S&P 500: {len(tickers)} stocks, "
              f"{sp500['sector'].nunique()} sectors")
    except Exception as e:
        print(f"    Warning: S&P 500 list unavailable: {e}")
        return {}

    # ── 2. Batch price download → daily % change ──
    changes = {t: 0.0 for t in tickers}
    try:
        raw   = yf.download(tickers, period="5d", interval="1d",
                            auto_adjust=True, progress=False, threads=True)
        try:
            close = raw["Close"]
        except KeyError:
            close = raw.xs("Close", axis=1, level=0)
        for sym in tickers:
            try:
                s = close[sym].dropna()
                if len(s) >= 2:
                    changes[sym] = round(
                        (float(s.iloc[-1]) - float(s.iloc[-2]))
                        / float(s.iloc[-2]) * 100, 2)
            except Exception:
                pass
    except Exception as e:
        print(f"    Warning: S&P 500 price batch failed: {e}")

    # ── 3. Market caps via threaded fast_info ──
    market_caps: dict[str, int] = {}

    def _mc(sym):
        try:
            fi = yf.Ticker(sym).fast_info
            return sym, int(getattr(fi, "market_cap", 0) or 0)
        except Exception:
            return sym, 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=24) as ex:
        for sym, mc in ex.map(_mc, tickers):
            market_caps[sym] = mc

    # ── 4. Group by sector, sort by market cap ──
    sector_map: dict[str, list] = {}
    for _, row in sp500.iterrows():
        sym = row["ticker"]
        sec = row["sector"]
        sector_map.setdefault(sec, []).append({
            "t": sym,
            "n": row["name"],
            "s": row["sub_industry"],
            "c": changes.get(sym, 0.0),
            "m": market_caps.get(sym, 0),
        })
    for sec in sector_map:
        sector_map[sec].sort(key=lambda x: x["m"], reverse=True)

    return sector_map


def fetch_spy_benchmark() -> dict:
    """
    Fetch SPY YTD, 1-year, and 3-year annualised returns for benchmark comparison.
    Returns dict with keys: ytd, one_yr, three_yr_ann, current_price.
    """
    try:
        hist = yf.Ticker("SPY").history(period="5y", interval="1d", auto_adjust=True)
        if hist.empty:
            return {}
        idx   = hist.index
        cur   = float(hist["Close"].iloc[-1])

        # YTD — first trading day of this calendar year
        this_year = idx[-1].year
        ytd_series = hist[idx.year == this_year]
        ytd_ret = ((cur - float(ytd_series["Close"].iloc[0])) /
                   float(ytd_series["Close"].iloc[0]) * 100) if not ytd_series.empty else None

        # 1-year
        one_yr_loc = idx.searchsorted(idx[-1] - pd.DateOffset(years=1))
        one_yr_ret = ((cur - float(hist["Close"].iloc[one_yr_loc])) /
                      float(hist["Close"].iloc[one_yr_loc]) * 100) if one_yr_loc < len(hist) else None

        # 3-year annualised
        three_yr_loc = idx.searchsorted(idx[-1] - pd.DateOffset(years=3))
        if three_yr_loc < len(hist):
            total_3 = (cur - float(hist["Close"].iloc[three_yr_loc])) / float(hist["Close"].iloc[three_yr_loc])
            three_yr_ann = ((1 + total_3) ** (1 / 3) - 1) * 100
        else:
            three_yr_ann = None

        return {
            "ytd":           round(ytd_ret,      2) if ytd_ret      is not None else None,
            "one_yr":        round(one_yr_ret,   2) if one_yr_ret   is not None else None,
            "three_yr_ann":  round(three_yr_ann, 2) if three_yr_ann is not None else None,
            "current_price": round(cur, 2),
        }
    except Exception as e:
        print(f"    Warning: SPY benchmark fetch failed: {e}")
        return {}


def fetch_market_overview() -> dict:
    """
    Fetch live data for:
      - Market strip  (S&P, Nasdaq, Dow, Russell, VIX, 10Y, Gold, Oil)
      - Sector heatmap (11 SPDR sector ETFs)
      - Yield curve   (3M, 5Y, 10Y, 30Y Treasury rates)
    Returns a single dict with keys 'strip', 'sectors', 'yield_curve'.
    """
    STRIP = [
        ("^GSPC",   "S&P 500"),
        ("^IXIC",   "NASDAQ"),
        ("^DJI",    "DOW"),
        ("^RUT",    "RUSSELL 2K"),
        ("^VIX",    "VIX"),
        ("^TNX",    "10Y YIELD"),
        ("GC=F",    "GOLD"),
        ("CL=F",    "WTI OIL"),
        ("DX-Y.NYB","USD INDEX"),
    ]
    SECTORS = [
        ("XLK",  "Technology",     "💻", "Information Technology"),
        ("XLC",  "Comm. Services", "📡", "Communication Services"),
        ("XLY",  "Cons. Disc.",    "🛍", "Consumer Discretionary"),
        ("XLP",  "Cons. Staples",  "🛒", "Consumer Staples"),
        ("XLE",  "Energy",         "⚡", "Energy"),
        ("XLF",  "Financials",     "🏦", "Financials"),
        ("XLV",  "Health Care",    "🏥", "Health Care"),
        ("XLI",  "Industrials",    "🏭", "Industrials"),
        ("XLB",  "Materials",      "⚗", "Materials"),
        ("XLRE", "Real Estate",    "🏢", "Real Estate"),
        ("XLU",  "Utilities",      "💡", "Utilities"),
    ]
    YIELD_CURVE = [
        ("^IRX", "3M"),
        ("^FVX", "5Y"),
        ("^TNX", "10Y"),
        ("^TYX", "30Y"),
    ]

    all_syms = list(dict.fromkeys(
        [s for s, _ in STRIP]
        + [s for s, _, _, _ in SECTORS]
        + [s for s, _ in YIELD_CURVE]
    ))

    # Batch download last 5 trading days
    prices = {}   # sym -> (current, prev_close)
    try:
        raw = yf.download(all_syms, period="5d", interval="1d",
                          auto_adjust=True, progress=False, threads=True)
        # Multi-ticker download has MultiIndex columns (field, ticker)
        try:
            close = raw["Close"]
        except KeyError:
            close = raw.xs("Close", axis=1, level=0)

        for sym in all_syms:
            try:
                series = close[sym].dropna()
                if len(series) >= 2:
                    prices[sym] = (float(series.iloc[-1]), float(series.iloc[-2]))
                elif len(series) == 1:
                    prices[sym] = (float(series.iloc[-1]), float(series.iloc[-1]))
            except Exception:
                pass
    except Exception as e:
        print(f"    Warning: market overview batch fetch failed: {e}")

    def _pct(sym):
        if sym not in prices:
            return None, None, 0.0
        cur, prev = prices[sym]
        pct = ((cur - prev) / prev * 100) if prev else 0.0
        return cur, prev, round(pct, 2)

    def _fmt_price(sym, cur):
        rate_syms = {"^TNX", "^IRX", "^FVX", "^TYX"}
        if sym in rate_syms:
            return f"{cur:.2f}%"
        if sym == "^VIX":
            return f"{cur:.2f}"
        if cur >= 10_000:
            return f"{cur:,.0f}"
        if cur >= 100:
            return f"{cur:,.2f}"
        return f"{cur:.2f}"

    # ── Strip ──
    strip = []
    for sym, label in STRIP:
        cur, _, pct = _pct(sym)
        if cur is None:
            continue
        strip.append({
            "label":      label,
            "price":      _fmt_price(sym, cur),
            "change_pct": pct,
            "direction":  "up" if pct >= 0 else "down",
        })

    # ── Sectors ──
    sectors = []
    for sym, label, icon, gics in SECTORS:
        _, _, pct = _pct(sym)
        if pct >= 2:    bg = "#14532d"
        elif pct >= 1:  bg = "#166534"
        elif pct >= 0:  bg = "#1a3a2a"
        elif pct >= -1: bg = "#3a1a1a"
        elif pct >= -2: bg = "#7f1d1d"
        else:           bg = "#991b1b"
        sectors.append({
            "label": label, "etf": sym, "icon": icon, "gics": gics,
            "change_pct": pct, "bg": bg,
        })

    # ── Yield curve ──
    yield_curve = []
    for sym, label in YIELD_CURVE:
        cur, _, _ = _pct(sym)
        if cur is not None:
            yield_curve.append({"label": label, "value": round(cur, 2)})

    # ── Spot prices for sidebar (Gold, Oil, DXY) ──
    spot = {}
    for sym, label in [("GC=F", "Gold"), ("CL=F", "WTI Oil"), ("DX-Y.NYB", "USD Index")]:
        cur, _, pct = _pct(sym)
        if cur is not None:
            spot[label] = {"price": _fmt_price(sym, cur), "change_pct": pct,
                           "direction": "up" if pct >= 0 else "down"}

    return {"strip": strip, "sectors": sectors, "yield_curve": yield_curve, "spot": spot}


def fetch_extended_data(symbols: list[str], stock_data: dict) -> None:
    """
    Fetch 3M/6M price momentum and earnings data for a list of symbols.
    Updates stock_data dicts in place — call after recommendations are filtered.
    """
    from datetime import datetime as dt

    print(f"  Downloading 6-month price history for {len(symbols)} tickers...")

    # ---- Batch momentum download (much faster than individual calls) ----
    try:
        if len(symbols) == 1:
            hist_raw  = yf.download(symbols[0], period="6mo", auto_adjust=True, progress=False)
            hist_close = {symbols[0]: hist_raw["Close"]}
        else:
            hist_raw  = yf.download(symbols, period="6mo", auto_adjust=True,
                                    progress=False, group_by="ticker")
            hist_close = {}
            for sym in symbols:
                try:
                    hist_close[sym] = hist_raw[sym]["Close"]
                except Exception:
                    hist_close[sym] = None
    except Exception as e:
        print(f"    Warning: batch history download failed: {e}")
        hist_close = {s: None for s in symbols}

    # ---- Per-ticker earnings data ----
    for sym in symbols:
        d = stock_data.get(sym, {})

        # -- Momentum --
        mom_3m = mom_6m = None
        try:
            series = hist_close.get(sym)
            if series is not None and len(series) > 5:
                cur   = float(series.dropna().iloc[-1])
                p3m   = float(series.dropna().iloc[max(0, len(series.dropna()) - 63)])
                p6m   = float(series.dropna().iloc[0])
                mom_3m = round((cur - p3m) / p3m * 100, 1)
                mom_6m = round((cur - p6m) / p6m * 100, 1)
        except Exception:
            pass

        # -- Earnings --
        earnings_date_str     = None
        days_to_earnings      = None
        eps_est_next          = None
        last_eps_actual       = None
        last_eps_estimate     = None
        last_eps_surprise_pct = None
        last_q_period         = None

        try:
            ticker = yf.Ticker(sym)
            info   = ticker.info

            # Next earnings date — try info timestamps first
            ts = info.get("earningsTimestampStart") or info.get("earningsTimestamp")
            if ts:
                dte = dt.fromtimestamp(int(ts))
                earnings_date_str = dte.strftime("%b %d, %Y")
                days_to_earnings  = (dte - dt.now()).days

            # Fallback: calendar dict
            if not earnings_date_str:
                try:
                    cal = ticker.calendar
                    if isinstance(cal, dict) and cal.get("Earnings Date"):
                        ed = cal["Earnings Date"]
                        if isinstance(ed, (list, tuple)) and ed:
                            dte = ed[0]
                            if hasattr(dte, "strftime"):
                                earnings_date_str = dte.strftime("%b %d, %Y")
                                days_to_earnings  = (dte.replace(tzinfo=None) - dt.now()).days
                except Exception:
                    pass

            # Next quarter EPS estimate from calendar
            try:
                cal = ticker.calendar
                if isinstance(cal, dict):
                    eps_est_next = cal.get("Earnings Average") or cal.get("EPS Average")
            except Exception:
                pass

            # Last quarter actual vs. estimate from earnings_history
            try:
                eh = ticker.earnings_history
                if eh is not None and not eh.empty:
                    row = eh.iloc[-1]
                    def _get(r, key):
                        return r.get(key) if hasattr(r, "get") else getattr(r, key, None)
                    last_eps_actual   = _get(row, "epsActual")
                    last_eps_estimate = _get(row, "epsEstimate")
                    surprise          = _get(row, "surprisePercent")
                    if surprise is not None:
                        last_eps_surprise_pct = round(float(surprise) * 100, 1)
                    elif last_eps_actual is not None and last_eps_estimate:
                        last_eps_surprise_pct = round(
                            (float(last_eps_actual) - float(last_eps_estimate))
                            / abs(float(last_eps_estimate)) * 100, 1
                        )
                    try:
                        last_q_period = str(eh.index[-1])[:7]
                    except Exception:
                        pass
            except Exception:
                pass

        except Exception as e:
            print(f"    Warning: earnings data failed for {sym}: {e}")

        d.update({
            "mom_3m":               mom_3m,
            "mom_6m":               mom_6m,
            "earnings_date":        earnings_date_str,
            "days_to_earnings":     days_to_earnings,
            "eps_est_next":         eps_est_next,
            "last_eps_actual":      last_eps_actual,
            "last_eps_estimate":    last_eps_estimate,
            "last_eps_surprise_pct": last_eps_surprise_pct,
            "last_q_period":        last_q_period,
        })
        stock_data[sym] = d


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------

def build_portfolio_df(portfolio: pd.DataFrame, stock_data: dict) -> pd.DataFrame:
    rows = []
    for _, row in portfolio.iterrows():
        sym = row["symbol"]
        d = stock_data.get(sym, {})
        price = d.get("price", 0) or 0
        shares = row["shares"]
        market_value = price * shares
        cost = row.get("cost_basis")
        gain_pct = ((market_value - cost) / cost * 100) if cost and cost > 0 else None
        upside = None
        if d.get("analyst_target") and price > 0:
            upside = (d["analyst_target"] - price) / price * 100

        rows.append({
            "Symbol":           sym,
            "Name":             d.get("name", sym),
            "Sector":           d.get("sector", "Unknown"),
            "Shares":           shares,
            "Price":            price,
            "Market Value":     market_value,
            "Cost Basis":       cost,
            "Gain/Loss %":      gain_pct,
            "Beta":             d.get("beta"),
            "Risk Tier":        d.get("auto_risk", "medium"),
            # Stock-specific
            "P/E":              d.get("pe"),
            "Fwd P/E":          d.get("forward_pe"),
            "P/B":              d.get("pb"),
            "Div Yield %":      d.get("dividend_yield"),
            "EPS Growth %":     d.get("eps_growth"),
            "Rev Growth %":     d.get("rev_growth"),
            "Growth Score":     d.get("growth_score"),
            "Analyst Target":   d.get("analyst_target"),
            "Analyst Upside %": upside,
            "Recommendation":   d.get("recommendation"),
            # ETF fields (None for stocks)
            "is_etf":           d.get("is_etf", False),
            "Expense Ratio":    d.get("expense_ratio"),
            "ETF Yield":        d.get("etf_yield"),
            "AUM":              d.get("aum"),
            "ETF Category":     d.get("etf_category"),
            "YTD Return":       d.get("ytd_return"),
            "3Y Return":        d.get("three_year_ret"),
            "5Y Return":        d.get("five_year_ret"),
            "Market Cap":       d.get("market_cap"),
        })

    return pd.DataFrame(rows)


def analyze_holdings(portfolio_df: pd.DataFrame, stock_data: dict,
                     held_symbols: set) -> list[dict]:
    """
    For each holding score its hold/watch/sell signal, explain reasons,
    and suggest same-sector replacements for anything flagged WATCH or SELL.
    Requires fetch_extended_data() to have already been called for portfolio symbols.
    """
    results = []

    for _, row in portfolio_df.iterrows():
        sym      = row["Symbol"]
        d        = stock_data.get(sym, {})

        # ETFs use different scoring — skip from stock-specific hold/sell analysis
        if d.get("is_etf") or row.get("is_etf"):
            continue

        price    = d.get("price") or row.get("Price") or 0
        target   = d.get("analyst_target")
        upside   = ((target - price) / price * 100) if target and price else None
        rec_key  = (d.get("recommendation") or "").lower()
        eps_g    = d.get("eps_growth") or 0
        rev_g    = d.get("rev_growth") or 0
        mom_3m   = d.get("mom_3m")
        pe       = d.get("pe")
        sector   = d.get("sector") or row.get("Sector", "Unknown")
        gain_pct = row.get("Gain/Loss %")

        score     = 100.0
        reasons   = []    # sell signals
        positives = []    # hold signals

        # Analyst rating
        if any(w in rec_key for w in ("strong sell", "sell")):
            score -= 45
            reasons.append(f"Analyst consensus is <strong>{rec_key.title()}</strong> — "
                           "Wall Street has turned bearish on this stock")
        elif "underperform" in rec_key:
            score -= 28
            reasons.append("Analyst rating: <strong>Underperform</strong> — "
                           "analysts expect it to lag the broader market")
        elif "hold" in rec_key:
            score -= 8
        elif any(w in rec_key for w in ("strong buy", "buy")):
            positives.append(f"Analyst consensus: <strong>{rec_key.title()}</strong>")

        # Upside to analyst target
        if upside is not None:
            if upside < 0:
                score -= 35
                reasons.append(
                    f"Analyst 12-month target (${target:.2f}) is <strong>below current price</strong> "
                    f"— Wall Street sees negative forward return"
                )
            elif upside < 8:
                score -= 18
                reasons.append(
                    f"Minimal upside: only <strong>{upside:.1f}%</strong> to analyst target "
                    f"(${target:.2f}) — risk/reward is unfavorable"
                )
            elif upside > 15:
                positives.append(
                    f"<strong>{upside:.1f}% upside</strong> remaining to analyst target (${target:.2f})"
                )

        # Price momentum
        if mom_3m is not None:
            if mom_3m < -20:
                score -= 22
                reasons.append(
                    f"Strong downtrend: <strong>{mom_3m:+.1f}%</strong> over 3 months — "
                    "sellers are in control"
                )
            elif mom_3m < -10:
                score -= 12
                reasons.append(
                    f"Weakening price momentum: <strong>{mom_3m:+.1f}%</strong> over 3 months"
                )
            elif mom_3m > 15:
                positives.append(
                    f"Strong price momentum: <strong>{mom_3m:+.1f}%</strong> over 3 months"
                )

        # Earnings growth
        if eps_g < -15:
            score -= 18
            reasons.append(
                f"Earnings declining sharply: <strong>EPS {eps_g:+.1f}%</strong> YoY "
                "— the profit engine is weakening"
            )
        elif eps_g < 0:
            score -= 10
            reasons.append(f"Negative earnings growth: <strong>EPS {eps_g:+.1f}%</strong> YoY")
        elif eps_g > 20:
            positives.append(f"Strong earnings growth: <strong>EPS +{eps_g:.1f}%</strong> YoY")

        # Expensive valuation vs slow growth
        if pe and pe > 40 and eps_g < 10:
            score -= 12
            reasons.append(
                f"Expensive at <strong>P/E {pe:.1f}x</strong> with slowing earnings growth "
                f"({eps_g:+.1f}%) — paying a premium for less growth"
            )

        # Revenue
        if rev_g > 15:
            positives.append(f"Revenue growing at <strong>{rev_g:+.1f}%</strong> YoY")

        score = max(0.0, min(100.0, score))

        if score >= 72:
            verdict = "HOLD"; verdict_color = "#34d399"; verdict_bg = "#052e16"
        elif score >= 45:
            verdict = "WATCH"; verdict_color = "#fbbf24"; verdict_bg = "#451a03"
        else:
            verdict = "SELL"; verdict_color = "#f87171"; verdict_bg = "#450a0a"

        # Replacement candidates (same sector, not already held)
        replacements = []
        if score < 72:
            universe   = SECTOR_UNIVERSE.get(sector, [])
            tier_order = ["speculative", "high", "medium", "low"]
            candidates = [(s, t) for s, _, t in universe
                          if s not in held_symbols and s != sym]
            candidates.sort(
                key=lambda x: tier_order.index(x[1]) if x[1] in tier_order else 99
            )
            for cand_sym, cand_tier in candidates:
                if len(replacements) >= 2:
                    break
                cd = stock_data.get(cand_sym, {})
                cp = cd.get("price") or 0
                ct = cd.get("analyst_target")
                if not cp or not ct or ct <= cp:
                    continue
                ok, _ = is_viable_pick(cd)
                if not ok:
                    continue
                replacements.append({
                    "symbol":       cand_sym,
                    "name":         cd.get("name", cand_sym),
                    "price":        cp,
                    "upside":       round((ct - cp) / cp * 100, 1),
                    "risk_tier":    cand_tier,
                    "recommendation": cd.get("recommendation", ""),
                    "eps_growth":   cd.get("eps_growth"),
                })

        results.append({
            "symbol":         sym,
            "name":           d.get("name", sym),
            "sector":         sector,
            "score":          round(score, 1),
            "verdict":        verdict,
            "verdict_color":  verdict_color,
            "verdict_bg":     verdict_bg,
            "reasons":        reasons,
            "positives":      positives,
            "upside":         upside,
            "mom_3m":         mom_3m,
            "eps_growth":     eps_g,
            "analyst_target": target,
            "recommendation": d.get("recommendation", ""),
            "replacements":   replacements,
            "price":          price,
            "gain_pct":       gain_pct,
        })

    # Sort: SELL first → WATCH → HOLD
    order = {"SELL": 0, "WATCH": 1, "HOLD": 2}
    results.sort(key=lambda r: order.get(r["verdict"], 3))
    return results


def compute_portfolio_beta(portfolio_df: pd.DataFrame) -> float | None:
    """Weighted-average portfolio beta using market value weights."""
    total = portfolio_df["Market Value"].sum()
    if not total:
        return None
    weighted = 0.0
    covered  = 0.0
    for _, row in portfolio_df.iterrows():
        beta = row.get("Beta")
        mv   = row.get("Market Value") or 0
        if beta is not None and not (isinstance(beta, float) and (math.isnan(beta) or math.isinf(beta))):
            weighted += float(beta) * (mv / total)
            covered  += mv / total
    return round(weighted / covered, 2) if covered > 0 else None


def sector_analysis(df: pd.DataFrame) -> pd.DataFrame:
    total = df["Market Value"].sum()
    sector_mv = df.groupby("Sector")["Market Value"].sum().reset_index()
    sector_mv["Portfolio Weight %"] = sector_mv["Market Value"] / total * 100
    sector_mv["S&P 500 Weight %"] = sector_mv["Sector"].map(SP500_SECTOR_WEIGHTS).fillna(0)
    sector_mv["Over/Under %"] = sector_mv["Portfolio Weight %"] - sector_mv["S&P 500 Weight %"]
    sector_mv = sector_mv.sort_values("Portfolio Weight %", ascending=False)
    return sector_mv


def is_viable_pick(d: dict) -> tuple[bool, str]:
    """
    Returns (ok, reason) — a stock passes only if it has enough real data
    and a positive outlook. Rejects:
      - No price or price <= 0
      - Analyst target exists but is BELOW current price (negative upside)
      - Sell/Underperform analyst rating
      - No analyst target AND no meaningful growth metrics
      - Less than 2 non-None key metrics (too sparse to evaluate)
    """
    price  = d.get("price") or 0
    target = d.get("analyst_target")
    rec    = (d.get("recommendation") or "").lower()
    eps_g  = d.get("eps_growth")
    rev_g  = d.get("rev_growth")
    fwd_pe = d.get("forward_pe")
    beta   = d.get("beta")

    if price <= 0:
        return False, "no price data"

    if target and target < price:
        return False, f"analyst target ${target:.2f} is below current price ${price:.2f}"

    if any(w in rec for w in ("sell", "underperform", "strong sell")):
        return False, f"analyst rating is '{rec}'"

    # Must have at least an analyst target OR meaningful growth data
    has_target  = target is not None
    has_growth  = (eps_g is not None and eps_g > 0) or (rev_g is not None and rev_g > 0)
    if not has_target and not has_growth:
        return False, "no analyst coverage and no positive growth data"

    # Count how many key metrics are populated
    key_metrics = [price, target, eps_g, rev_g, fwd_pe, beta,
                   d.get("pe"), d.get("growth_score")]
    populated = sum(1 for m in key_metrics if m is not None)
    if populated < 3:
        return False, "insufficient data to evaluate"

    return True, "ok"


def _build_rec_dict(sym, d, category, gap_pct, tier, in_sp):
    price = d.get("price", 0) or 0
    upside = None
    if d.get("analyst_target") and price > 0:
        upside = (d["analyst_target"] - price) / price * 100
    return {
        "category":       category,
        "gap_pct":        round(gap_pct, 1) if gap_pct else None,
        "symbol":         sym,
        "name":           d.get("name", sym),
        "price":          price,
        "beta":           d.get("beta"),
        "risk_tier":      tier,
        "in_sp500":       in_sp,
        "pe":             d.get("pe"),
        "forward_pe":     d.get("forward_pe"),
        "div_yield":      d.get("dividend_yield"),
        "eps_growth":     d.get("eps_growth"),
        "rev_growth":     d.get("rev_growth"),
        "growth_score":   d.get("growth_score"),
        "analyst_target": d.get("analyst_target"),
        "upside":         upside,
        "recommendation": d.get("recommendation"),
        # Extended fields — populated by fetch_extended_data() in main()
        "mom_3m":               d.get("mom_3m"),
        "mom_6m":               d.get("mom_6m"),
        "earnings_date":        d.get("earnings_date"),
        "days_to_earnings":     d.get("days_to_earnings"),
        "eps_est_next":         d.get("eps_est_next"),
        "last_eps_actual":      d.get("last_eps_actual"),
        "last_eps_estimate":    d.get("last_eps_estimate"),
        "last_eps_surprise_pct": d.get("last_eps_surprise_pct"),
        "last_q_period":        d.get("last_q_period"),
    }


def generate_recommendations(
    sector_df: pd.DataFrame,
    held_symbols: set,
    stock_data: dict,
    cash: float = DEFAULT_CASH,
    risk_profile: str = RISK_PROFILE,
) -> tuple[list[dict], list[dict], dict]:
    """
    Returns (sector_recs, thematic_recs, cash_plan).

    sector_recs   — picks in underweight S&P sectors
    thematic_recs — high-growth non-S&P theme picks
    cash_plan     — suggested dollar allocation per risk tier
    """
    alloc = CASH_ALLOCATION[risk_profile]

    tier_order = (
        ["speculative", "high", "medium", "low"]
        if risk_profile == "aggressive"
        else ["low", "medium", "high", "speculative"]
    )

    # ---- Sector recommendations ----
    sector_recs = []
    underweight = sector_df[sector_df["Over/Under %"] < -1.0].sort_values("Over/Under %")

    for _, row in underweight.iterrows():
        sector = row["Sector"]
        gap = abs(row["Over/Under %"])
        universe = SECTOR_UNIVERSE.get(sector, [])

        # Fetch all candidates first so we can filter on real data
        all_syms = [s for s, _, _ in universe if s not in held_symbols]
        missing  = [s for s in all_syms if s not in stock_data]
        if missing:
            stock_data.update(fetch_stock_data(missing))

        # Sort by tier preference, then filter for viability, cap at 3
        ordered = []
        for tier in tier_order:
            ordered += [(sym, in_sp, t) for sym, in_sp, t in universe
                        if t == tier and sym not in held_symbols]

        added = 0
        for sym, in_sp, tier in ordered:
            if added >= 3:
                break
            d = stock_data.get(sym, {})
            ok, reason = is_viable_pick(d)
            if not ok:
                print(f"    Skipping {sym}: {reason}")
                continue
            sector_recs.append(_build_rec_dict(sym, d, sector, gap, tier, in_sp))
            added += 1

    # ---- Thematic (non-S&P growth) recommendations ----
    thematic_recs = []
    if risk_profile in ("aggressive", "moderate"):
        for theme, picks in THEMATIC_PICKS.items():
            all_syms = [s for s, _, _ in picks if s not in held_symbols]
            missing  = [s for s in all_syms if s not in stock_data]
            if missing:
                stock_data.update(fetch_stock_data(missing))

            ordered = sorted(
                [(s, t) for s, _, t in picks if s not in held_symbols],
                key=lambda x: tier_order.index(x[1]) if x[1] in tier_order else 99
            )

            added = 0
            for sym, tier in ordered:
                if added >= 3:
                    break
                d = stock_data.get(sym, {})
                ok, reason = is_viable_pick(d)
                if not ok:
                    print(f"    Skipping {sym}: {reason}")
                    continue
                thematic_recs.append(_build_rec_dict(sym, d, theme, None, tier, False))
                added += 1

    # ---- Cash allocation plan ----
    cash_plan = {
        tier: {"pct": pct * 100, "dollars": round(cash * pct, 0)}
        for tier, pct in alloc.items()
        if pct > 0
    }

    return sector_recs, thematic_recs, cash_plan


# ---------------------------------------------------------------------------
# Buy / sell context builder
# ---------------------------------------------------------------------------

STOP_LOSS_PCT = {
    "low":         0.08,
    "medium":      0.12,
    "high":        0.18,
    "speculative": 0.25,
}

METRIC_EXPLANATIONS = {
    "Beta":         ("Volatility vs the market",
                     "A beta >1 means the stock moves more than the market. "
                     "Higher beta = bigger swings up AND down."),
    "P/E":          ("Price-to-Earnings (trailing)",
                     "How much you pay per $1 of past earnings. "
                     "Lower = cheaper, but a high P/E can still be worth it if growth is strong."),
    "Fwd P/E":      ("Price-to-Earnings (forward)",
                     "Same as P/E but uses expected future earnings. "
                     "A lower forward P/E than trailing P/E signals analysts expect earnings to grow."),
    "EPS Growth":   ("Earnings Per Share growth",
                     "How fast the company's profit per share is growing year-over-year. "
                     "Higher is better — this is the core engine of long-term stock gains."),
    "Rev Growth":   ("Revenue growth",
                     "How fast total sales are growing. "
                     "Important for pre-profit companies where earnings don't exist yet."),
    "Growth Score": ("Composite growth score (0–100+)",
                     "Weighted blend of EPS growth (40%), revenue growth (30%), "
                     "and analyst upside (30%). Higher = more growth momentum."),
    "Upside":       ("Analyst price target upside",
                     "Gap between today's price and Wall Street's average 12-month price target. "
                     "Take with a grain of salt — analysts can be late or wrong."),
    "Div Yield":    ("Dividend yield",
                     "Annual dividend as a % of price. "
                     "Less relevant for aggressive growth plays — reinvested earnings matter more."),
    "Analyst Target": ("Mean analyst 12-month price target",
                       "Average of all Wall Street analyst targets. "
                       "Useful as a rough fair-value anchor, not a guarantee."),
}


def build_rec_context(rec: dict, tier_budget: float = 0) -> dict:
    """Add buy_at, target_price, stop_loss, suggested shares, and why_text to a rec dict."""
    price      = rec.get("price") or 0
    risk_tier  = rec.get("risk_tier", "medium")
    stop_pct   = STOP_LOSS_PCT.get(risk_tier, 0.15)
    analyst_t  = rec.get("analyst_target")
    eps_g      = rec.get("eps_growth") or 0
    rev_g      = rec.get("rev_growth") or 0
    beta       = rec.get("beta")
    fwd_pe     = rec.get("forward_pe")
    growth_s   = rec.get("growth_score") or 0
    upside     = rec.get("upside") or 0
    in_sp      = rec.get("in_sp500", False)
    category   = rec.get("category", "")
    gap        = rec.get("gap_pct")

    # Buy zone: at or slightly below current price
    buy_at = round(price * 0.99, 2) if price else None

    # Profit target — analyst target is already guaranteed above price by is_viable_pick.
    # Fallback estimate from growth score only if no analyst target.
    if analyst_t and analyst_t > price:
        target = round(analyst_t, 2)
    elif price and growth_s > 0:
        target = round(price * (1 + min(growth_s, 80) / 100), 2)
    else:
        target = None

    # Safety check: if target still ended up <= buy_at, clear it
    if target and buy_at and target <= buy_at:
        target = None

    stop_loss = round(price * (1 - stop_pct), 2) if price else None

    # ---- Projected timeline to reach price target ----
    # Analyst consensus targets are 12-month forward estimates by convention.
    # We adjust ±months based on growth momentum and upside magnitude.
    if upside > 0:
        eps_g_pos  = max(eps_g, 0)
        rev_g_pos  = max(rev_g, 0)
        momentum   = eps_g_pos * 0.5 + rev_g_pos * 0.5   # blended growth pace

        # Momentum compresses or extends the 12-month baseline
        if momentum >= 60:
            mom_adj = -3
        elif momentum >= 30:
            mom_adj = -1
        elif momentum <= 5:
            mom_adj = +3
        else:
            mom_adj = 0

        # Larger upside generally takes more time
        if upside > 60:
            upside_adj = +6
        elif upside > 40:
            upside_adj = +3
        elif upside > 20:
            upside_adj = 0
        else:
            upside_adj = -2

        est_months = 12 + mom_adj + upside_adj
        est_months = max(min(est_months, 36), 3)

        if est_months <= 6:
            timeline_label   = f"~{est_months} months"
            timeline_context = "Near-term — strong growth momentum"
        elif est_months <= 12:
            timeline_label   = f"~{est_months} months"
            timeline_context = "Within analyst 12-month horizon"
        elif est_months <= 18:
            timeline_label   = f"~{est_months} months"
            timeline_context = "~1–1.5 year hold"
        elif est_months <= 24:
            timeline_label   = f"~{est_months} months"
            timeline_context = "~1.5–2 year hold"
        else:
            timeline_label   = f"~{est_months} months"
            timeline_context = f"~{est_months // 12}+ year patient hold"
    else:
        est_months       = None
        timeline_label   = "—"
        timeline_context = "No positive target available"

    # Suggested shares: how many whole shares fit within this stock's tier budget
    if buy_at and buy_at > 0 and tier_budget > 0:
        suggested_shares = int(tier_budget // buy_at)
        suggested_spend  = round(suggested_shares * buy_at, 2)
        leftover         = round(tier_budget - suggested_spend, 2)
    else:
        suggested_shares = None
        suggested_spend  = None
        leftover         = None

    # "Why" narrative — built from actual data
    bullets = []

    if gap:
        bullets.append(
            f"Your portfolio is <strong>underweight {category} by {gap}%</strong> vs the S&P 500 — "
            f"adding this stock helps rebalance your sector exposure."
        )

    if risk_tier in ("high", "speculative"):
        bullets.append(
            f"Classified as <strong>{risk_tier}</strong> — suited to your aggressive profile "
            f"and long time horizon. Expect higher volatility in exchange for higher upside."
        )

    if eps_g > 20:
        bullets.append(
            f"<strong>EPS growing at {eps_g:+.1f}%</strong> year-over-year — "
            f"strong earnings momentum is one of the most reliable drivers of share price appreciation."
        )
    elif eps_g > 0:
        bullets.append(f"Earnings growing at {eps_g:+.1f}% — positive but moderate growth.")

    if rev_g > 20:
        bullets.append(
            f"<strong>Revenue growing at {rev_g:+.1f}%</strong> — "
            f"top-line expansion indicates the business is capturing market share."
        )

    if fwd_pe and fwd_pe < 20:
        bullets.append(
            f"Forward P/E of <strong>{fwd_pe:.1f}x</strong> is reasonable — "
            f"you're not overpaying relative to expected earnings."
        )
    elif fwd_pe and fwd_pe > 50:
        bullets.append(
            f"Forward P/E of <strong>{fwd_pe:.1f}x</strong> is elevated — "
            f"the market is pricing in high growth. This is a bet on execution."
        )

    if upside > 20:
        bullets.append(
            f"Analyst consensus sees <strong>{upside:+.1f}% upside</strong> to the average "
            f"price target of ${analyst_t:,.2f}." if analyst_t else
            f"Analyst consensus sees <strong>{upside:+.1f}% upside</strong> to current price."
        )

    if beta and beta > 1.5:
        bullets.append(
            f"Beta of <strong>{beta:.2f}</strong> — this stock moves ~{beta:.1f}x the market. "
            f"Use position sizing to manage risk (consider capping at 3–5% of portfolio)."
        )

    if not in_sp:
        bullets.append(
            "This stock is <strong>outside the S&P 500</strong> — higher growth potential "
            "but less institutional coverage. Do additional research before buying."
        )

    if growth_s > 40:
        bullets.append(
            f"<strong>Growth Score: {growth_s:.0f}</strong> — one of the stronger composite "
            f"scores in this analysis, reflecting strong EPS, revenue, and analyst momentum."
        )

    if not bullets:
        bullets.append("No specific standout metrics — consider as a diversification play.")

    why_html = "<br>".join(f"• {b}" for b in bullets)

    return {
        **rec,
        "buy_at":           buy_at,
        "target":           target,
        "stop_loss":        stop_loss,
        "stop_pct":         int(stop_pct * 100),
        "suggested_shares": suggested_shares,
        "suggested_spend":  suggested_spend,
        "leftover":         leftover,
        "tier_budget":      round(tier_budget, 2) if tier_budget else None,
        "why_html":         why_html,
        "timeline_months":  est_months,
        "timeline_label":   timeline_label,
        "timeline_context": timeline_context,
    }


# ---------------------------------------------------------------------------
# Buy score ranking
# ---------------------------------------------------------------------------

REC_SCORE = {"strong buy": 1.0, "buy": 0.8, "hold": 0.4,
             "underperform": 0.1, "sell": 0.0, "strong sell": 0.0}

def compute_buy_score(rec: dict) -> float:
    """
    Composite 0-100 score weighting:
      30% — analyst upside (capped at 100%)
      20% — growth score
      20% — analyst recommendation
      10% — revenue growth (capped at 60%)
      10% — EPS growth (capped at 60%)
      10% — 3-month price momentum (-30% → 0pts, 0% → 33pts, +60% → 100pts)
    """
    upside    = min(rec.get("upside") or 0, 100)
    growth_s  = min(rec.get("growth_score") or 0, 100)
    rec_key   = (rec.get("recommendation") or "").lower()
    rec_score = REC_SCORE.get(rec_key, 0.5) * 100
    rev_g     = min(rec.get("rev_growth") or 0, 60)
    eps_g     = min(rec.get("eps_growth") or 0, 60)
    mom_3m    = rec.get("mom_3m") or 0
    # Scale momentum: -30 → 0, 0 → 33, +60 → 100
    mom_score = max(0.0, min((mom_3m + 30) / 90 * 100, 100))

    score = (
        upside    * 0.30 +
        growth_s  * 0.20 +
        rec_score * 0.20 +
        rev_g     * 0.10 +
        eps_g     * 0.10 +
        mom_score * 0.10
    )
    return round(score, 1)


def rank_recommendations(sector_recs: list, thematic_recs: list) -> list:
    ranked = []
    for rec in sector_recs + thematic_recs:
        score = compute_buy_score(rec)
        ranked.append({**rec, "buy_score": score})
    ranked.sort(key=lambda r: r["buy_score"], reverse=True)
    for i, r in enumerate(ranked, 1):
        r["rank"] = i
    return ranked


# ---------------------------------------------------------------------------
# HTML report
# ---------------------------------------------------------------------------

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Portfolio Analysis — {{ date }}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.plot.ly/plotly-basic-3.0.0.min.js"></script>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
         background: #0f172a; color: #e2e8f0; min-height: 100vh; }
  .header { background: linear-gradient(135deg, #1e3a5f, #0f2942);
            padding: 32px 40px; border-bottom: 1px solid #1e3a5f; }
  .header h1 { font-size: 28px; font-weight: 700; color: #60a5fa; }
  .header p  { color: #94a3b8; margin-top: 4px; font-size: 14px; }
  .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
                gap: 16px; padding: 24px 40px; }
  .stat-card { background: #1e293b; border: 1px solid #334155;
               border-radius: 12px; padding: 20px; }
  .stat-card .label { font-size: 12px; color: #64748b; text-transform: uppercase;
                      letter-spacing: .05em; }
  .stat-card .value { font-size: 26px; font-weight: 700; margin-top: 6px; color: #f1f5f9; }
  .stat-card .sub   { font-size: 13px; color: #64748b; margin-top: 4px; }
  .positive { color: #34d399 !important; }
  .negative { color: #f87171 !important; }
  .section { padding: 0 40px 32px; }
  .section h2 { font-size: 18px; font-weight: 600; color: #93c5fd;
                margin-bottom: 16px; padding-top: 24px;
                border-top: 1px solid #1e293b; }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { background: #1e293b; color: #64748b; font-weight: 600;
       text-transform: uppercase; letter-spacing: .05em;
       padding: 10px 12px; text-align: left; border-bottom: 1px solid #334155; }
  td { padding: 10px 12px; border-bottom: 1px solid #1e293b; color: #cbd5e1; }
  tr:hover td { background: #1e293b; }
  .badge { display: inline-block; padding: 2px 8px; border-radius: 9999px;
           font-size: 11px; font-weight: 600; }
  .badge-green  { background: #064e3b; color: #34d399; }
  .badge-yellow { background: #451a03; color: #fbbf24; }
  .badge-red    { background: #450a0a; color: #f87171; }
  .badge-blue   { background: #1e3a5f; color: #60a5fa; }
  .sector-bar { display: flex; align-items: center; gap: 8px; }
  .bar-wrap { flex: 1; background: #1e293b; border-radius: 4px; height: 8px; overflow: hidden; }
  .bar-fill-port { background: #3b82f6; height: 100%; border-radius: 4px; }
  .bar-fill-sp   { background: #64748b; height: 100%; border-radius: 4px; }
  .rec-card { background: #1e293b; border: 1px solid #334155;
              border-radius: 10px; padding: 16px 20px; margin-bottom: 12px; }
  .rec-header { display: flex; justify-content: space-between; align-items: flex-start; }
  .rec-symbol { font-size: 20px; font-weight: 700; color: #60a5fa; }
  .rec-name   { font-size: 13px; color: #64748b; margin-top: 2px; }
  .rec-sector { font-size: 11px; color: #fbbf24; font-weight: 600;
                text-transform: uppercase; letter-spacing: .05em; }
  .rec-metrics { display: flex; gap: 24px; margin-top: 12px; flex-wrap: wrap; }
  .rec-metric .k { font-size: 11px; color: #64748b; }
  .rec-metric .v { font-size: 15px; font-weight: 600; color: #e2e8f0; margin-top: 2px; }
  .gap-note { font-size: 12px; color: #f87171; margin-top: 6px; }
  footer { text-align: center; padding: 24px; font-size: 12px; color: #334155; }
  /* Dropdown detail panel */
  .dropdown-btn { display: flex; align-items: center; gap: 6px; margin-top: 14px;
                  background: none; border: 1px solid #334155; color: #94a3b8;
                  font-size: 12px; padding: 5px 12px; border-radius: 6px; cursor: pointer;
                  transition: all .15s; }
  .dropdown-btn:hover { border-color: #60a5fa; color: #60a5fa; }
  .dropdown-btn .arrow { transition: transform .2s; display: inline-block; }
  .dropdown-btn.open .arrow { transform: rotate(180deg); }
  .detail-panel { display: none; margin-top: 14px; border-top: 1px solid #334155;
                  padding-top: 14px; }
  .detail-panel.open { display: block; }
  .detail-section { margin-bottom: 16px; }
  .detail-section h4 { font-size: 12px; font-weight: 700; color: #64748b;
                       text-transform: uppercase; letter-spacing: .06em; margin-bottom: 8px; }
  .why-box { background: #0f172a; border-left: 3px solid #3b82f6; border-radius: 0 6px 6px 0;
             padding: 10px 14px; font-size: 13px; color: #cbd5e1; line-height: 1.6; }
  .buysell-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; }
  .bs-card { border-radius: 8px; padding: 12px 14px; }
  .bs-card.buy  { background: #064e3b; border: 1px solid #34d39944; }
  .bs-card.target { background: #1e3a5f; border: 1px solid #60a5fa44; }
  .bs-card.stop { background: #450a0a; border: 1px solid #f8717144; }
  .bs-card .bs-label { font-size: 11px; font-weight: 700; text-transform: uppercase;
                       letter-spacing: .05em; margin-bottom: 4px; }
  .bs-card.buy    .bs-label { color: #34d399; }
  .bs-card.target .bs-label { color: #60a5fa; }
  .bs-card.stop     .bs-label { color: #f87171; }
  .bs-card.timeline { background: #1c1a2e; border: 1px solid #a78bfa44; }
  .bs-card.timeline .bs-label { color: #a78bfa; }
  .bs-card .bs-price { font-size: 20px; font-weight: 700; color: #f1f5f9; }
  .bs-card .bs-note  { font-size: 11px; color: #64748b; margin-top: 3px; }
  .metric-table { width: 100%; border-collapse: collapse; font-size: 12px; }
  .metric-table tr { border-bottom: 1px solid #1e293b; }
  .metric-table td { padding: 6px 8px; }
  .metric-table td:first-child { color: #60a5fa; font-weight: 600; width: 120px; white-space: nowrap; }
  .metric-table td:nth-child(2) { color: #f1f5f9; width: 90px; }
  .metric-table td:last-child { color: #94a3b8; font-style: italic; }
  /* Header performance banner */
  .perf-banner { display: flex; gap: 0; margin-top: 20px; padding-top: 20px;
                 border-top: 1px solid #1e3a5f; flex-wrap: wrap; align-items: center; }
  .perf-item { padding: 0 28px 0 0; }
  .perf-item:not(:last-child) { border-right: 1px solid #1e3a5f; margin-right: 28px; }
  .perf-label { font-size: 11px; color: #64748b; text-transform: uppercase;
                letter-spacing: .06em; margin-bottom: 4px; }
  .perf-value { font-size: 32px; font-weight: 800; line-height: 1; }
  .perf-sub   { font-size: 13px; color: #64748b; margin-top: 4px; }
  /* Hold/Sell analysis */
  .holding-card { background: #1e293b; border: 1px solid #334155; border-radius: 10px;
                  padding: 18px 20px; }
  .verdict-badge { display: inline-flex; align-items: center; gap: 6px;
                   padding: 4px 12px; border-radius: 6px; font-size: 13px; font-weight: 700; }
  .signal-list { list-style: none; margin-top: 10px; }
  .signal-list li { font-size: 13px; color: #cbd5e1; padding: 4px 0;
                    padding-left: 20px; position: relative; line-height: 1.5; }
  .signal-list li::before { position: absolute; left: 0; }
  .signal-list.bad  li::before { content: "▼"; color: #f87171; font-size: 10px; top: 6px; }
  .signal-list.good li::before { content: "▲"; color: #34d399; font-size: 10px; top: 6px; }
  .replace-card { background: #0f172a; border: 1px solid #334155; border-radius: 8px;
                  padding: 10px 14px; display: flex; align-items: center;
                  justify-content: space-between; gap: 12px; }
  /* Tab navigation */
  .tab-nav { display: flex; gap: 2px; padding: 0 40px;
             background: #0a1628; border-bottom: 1px solid #1e3a5f;
             position: sticky; top: 0; z-index: 200; }
  .tab-btn { padding: 13px 22px; font-size: 13px; font-weight: 600; color: #64748b;
             border: none; background: none; cursor: pointer;
             border-bottom: 2px solid transparent; transition: all .15s; letter-spacing: .01em; }
  .tab-btn:hover { color: #93c5fd; }
  .tab-btn.active { color: #60a5fa; border-bottom-color: #3b82f6; }
  .tab-pane { display: none; }
  .tab-pane.active { display: block; }
  /* News — Yahoo Finance style */
  .news-symbol-header { font-size: 15px; font-weight: 700; padding-bottom: 10px;
                        margin-bottom: 2px; border-bottom: 2px solid #1e293b; }
  .news-list { list-style: none; }
  .news-item { display: flex; gap: 12px; align-items: flex-start;
               padding: 14px 0; border-bottom: 1px solid #1e293b; }
  .news-item:last-child { border-bottom: none; }
  .news-thumb { width: 130px; height: 86px; object-fit: cover; border-radius: 6px;
                flex-shrink: 0; background: #1e293b; }
  .news-body { flex: 1; min-width: 0; }
  .news-source-row { display: flex; align-items: center; gap: 8px; margin-bottom: 5px; }
  .news-publisher { font-size: 11px; font-weight: 700; color: #60a5fa;
                    text-transform: uppercase; letter-spacing: .04em; }
  .news-age { font-size: 11px; color: #475569; }
  .news-title { font-size: 15px; font-weight: 600; color: #e2e8f0; line-height: 1.4;
                text-decoration: none; display: -webkit-box;
                -webkit-line-clamp: 2; -webkit-box-orient: vertical; overflow: hidden; }
  .news-title:hover { color: #60a5fa; text-decoration: underline; }
  .news-summary { font-size: 13px; color: #64748b; margin-top: 5px; line-height: 1.5;
                  display: -webkit-box; -webkit-line-clamp: 2;
                  -webkit-box-orient: vertical; overflow: hidden; }
  .news-section-wrap { background: #1e293b; border: 1px solid #334155;
                       border-radius: 10px; padding: 0 20px; margin-bottom: 20px; }
  .news-section-wrap .news-item:first-child { padding-top: 16px; }
  .news-section-wrap .news-item:last-child { padding-bottom: 16px; border-bottom: none; }
  /* Chart modal */
  .chart-modal-overlay { display:none;position:fixed;inset:0;z-index:2000;
    background:rgba(0,0,0,.75);align-items:center;justify-content:center; }
  .chart-modal-box { background:#0f172a;border:1px solid #334155;border-radius:14px;
    width:92%;max-width:960px;padding:24px 28px;position:relative; }
  .tf-row { display:flex;gap:4px;flex-wrap:wrap; }
  .tf-btn { padding:5px 14px;background:none;border:1px solid #334155;color:#64748b;
    font-size:12px;font-weight:600;cursor:pointer;border-radius:6px;transition:all .15s; }
  .tf-btn:hover { color:#93c5fd;border-color:#60a5fa44; }
  .tf-btn.active { background:#1e3a5f;color:#60a5fa;border-color:#3b82f655; }
  details summary::-webkit-details-marker { display:none; }
  details summary { user-select:none; }
  .chart-btn { display:inline-flex;align-items:center;gap:5px;margin-top:14px;margin-left:6px;
    background:none;border:1px solid #334155;color:#94a3b8;font-size:12px;padding:5px 12px;
    border-radius:6px;cursor:pointer;transition:all .15s; }
  .chart-btn:hover { border-color:#a78bfa;color:#a78bfa; }
  @keyframes fillbar {
    from { width: 0% }
    to   { width: var(--bar-end, 0%) }
  }
  @keyframes pulse {
    0%, 100% { opacity: 1 }
    50%       { opacity: 0.5 }
  }
  .earnings-soon { animation: pulse 2s infinite; }
  .earnings-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }

  /* ── ETF Picks ── */
  .etf-card {
    background: #1e293b;
    border: 1px solid #334155;
    border-radius: 10px;
    padding: 18px 20px;
    display: flex;
    flex-direction: column;
    gap: 12px;
  }
  .etf-card:hover { border-color: #38bdf8; }
  .etf-header { display: flex; justify-content: space-between; align-items: flex-start; gap: 12px; }
  .etf-symbol { font-size: 22px; font-weight: 800; color: #38bdf8; }
  .etf-name   { font-size: 13px; color: #94a3b8; margin-top: 2px; }
  .etf-cat-badge {
    font-size: 10px; font-weight: 700; letter-spacing: .8px;
    padding: 3px 9px; border-radius: 20px;
    background: #0c2a3a; color: #38bdf8; border: 1px solid #1e4a6a;
    white-space: nowrap;
  }
  .etf-score-ring {
    width: 48px; height: 48px; border-radius: 50%;
    background: conic-gradient(#38bdf8 var(--pct), #1e293b 0);
    display: flex; align-items: center; justify-content: center;
    flex-shrink: 0;
    position: relative;
  }
  .etf-score-ring::after {
    content: ''; position: absolute;
    width: 36px; height: 36px; border-radius: 50%;
    background: #1e293b;
  }
  .etf-score-val {
    position: relative; z-index: 1;
    font-size: 12px; font-weight: 800; color: #38bdf8;
  }
  .etf-desc { font-size: 13px; color: #94a3b8; line-height: 1.55; }
  .etf-metrics {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 8px;
  }
  .etf-metric { background: #0f172a; border-radius: 6px; padding: 7px 10px; text-align: center; }
  .etf-metric .k { font-size: 9px; color: #475569; font-weight: 700; letter-spacing: .8px; text-transform: uppercase; }
  .etf-metric .v { font-size: 14px; font-weight: 700; color: #e2e8f0; margin-top: 2px; }
  .etf-badge { display: inline-block; font-size: 10px; font-weight: 700;
               padding: 2px 8px; border-radius: 4px; }
  .etf-badge-etf { background: #0c2a3a; color: #38bdf8; border: 1px solid #1e4a6a; }

  /* ── Market Overview Strip ── */
  .market-strip {
    background: #070e1a;
    border-bottom: 1px solid #1e3a5f;
    display: flex;
    overflow-x: auto;
    scrollbar-width: none;
    padding: 0 16px;
  }
  .market-strip::-webkit-scrollbar { display: none; }
  .strip-item {
    display: flex;
    flex-direction: column;
    align-items: center;
    padding: 7px 18px;
    border-right: 1px solid #1a2a40;
    min-width: 108px;
    flex-shrink: 0;
    gap: 1px;
  }
  .strip-label {
    font-size: 9px; font-weight: 700; color: #475569;
    letter-spacing: 1.2px; text-transform: uppercase;
  }
  .strip-price { font-size: 13px; font-weight: 600; color: #e2e8f0; font-variant-numeric: tabular-nums; }
  .strip-chg   { font-size: 11px; font-weight: 700; }
  .strip-up    { color: #34d399; }
  .strip-down  { color: #f87171; }

  /* ── Sector Heatmap ── */
  .heatmap-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(130px, 1fr));
    gap: 6px;
    margin-bottom: 4px;
  }
  .heatmap-tile {
    border-radius: 7px;
    padding: 10px 12px;
    display: flex;
    flex-direction: column;
    gap: 3px;
    transition: filter .15s;
    cursor: default;
  }
  .heatmap-tile { transition: filter .15s; }
  .heatmap-icon   { font-size: 16px; line-height: 1; }
  .heatmap-sector { font-size: 11px; font-weight: 600; color: #e2e8f0; }
  .heatmap-etf    { font-size: 9px;  color: rgba(255,255,255,.45); letter-spacing:.8px; }
  .heatmap-pct    { font-size: 15px; font-weight: 800; margin-top: 2px; }

  /* ── Sector Treemap Modal ── */
  .sector-modal-overlay {
    display: none; position: fixed; inset: 0;
    background: rgba(0,0,0,.85); z-index: 1000;
    align-items: center; justify-content: center;
    padding: 24px;
  }
  .sector-modal-overlay.open { display: flex; }
  .sector-modal {
    background: #0f172a; border: 1px solid #334155;
    border-radius: 14px; width: 100%; max-width: 1300px;
    max-height: 90vh; display: flex; flex-direction: column;
    overflow: hidden;
  }
  .sector-modal-header {
    display: flex; justify-content: space-between; align-items: center;
    padding: 16px 22px; border-bottom: 1px solid #1e293b; flex-shrink: 0;
  }
  .sector-modal-title {
    font-size: 18px; font-weight: 700; color: #e2e8f0;
  }
  .sector-modal-subtitle {
    font-size: 12px; color: #475569; margin-top: 2px;
  }
  #sector-treemap { flex: 1; min-height: 0; }

  /* ── ETF Holdings Modal ── */
  .holdings-modal-overlay {
    display: none; position: fixed; inset: 0; z-index: 3000;
    background: rgba(0,0,0,.75); align-items: center; justify-content: center;
  }
  .holdings-modal-overlay.open { display: flex; }
  .holdings-modal {
    background: #1e293b; border: 1px solid #334155; border-radius: 14px;
    width: min(480px, 95vw); max-height: 80vh; display: flex; flex-direction: column;
    overflow: hidden;
  }
  .holdings-modal-header {
    display: flex; align-items: center; justify-content: space-between;
    padding: 16px 20px; border-bottom: 1px solid #334155;
  }
  .holdings-modal-title { font-size: 15px; font-weight: 700; color: #f1f5f9; }
  .holdings-modal-sub   { font-size: 11px; color: #64748b; margin-top: 2px; }
  .holdings-table-wrap  { overflow-y: auto; padding: 12px 20px 20px; }
  .holdings-table { width: 100%; border-collapse: collapse; font-size: 12px; }
  .holdings-table th { color: #64748b; font-weight: 600; text-align: left;
    padding: 6px 8px; border-bottom: 1px solid #334155; }
  .holdings-table td { padding: 7px 8px; border-bottom: 1px solid #1e293b; color: #e2e8f0; }
  .holdings-table tr:hover td { background: #0f172a; }
  .holdings-bar { height: 6px; border-radius: 3px; background: #38bdf8;
    display: inline-block; vertical-align: middle; margin-left: 6px; }

  /* ── Earnings & Alerts tiles ── */
  .earn-list, .alert-list { display:flex; flex-direction:column; gap:5px; margin-top:6px; }
  .earn-row, .alert-row-item { display:flex; justify-content:space-between; align-items:center; gap:10px; font-size:12px; }
  .earn-sym { font-weight:700; color:#60a5fa; min-width:44px; }
  .earn-soon { color:#fbbf24; font-weight:700; }
  .earn-today { color:#f87171; font-weight:800; }
  .earn-normal { color:#94a3b8; }
  .alert-sym-lbl { font-weight:700; color:#60a5fa; min-width:44px; }
  .alert-price-lbl { color:#94a3b8; }
  .alert-hit { color:#34d399; font-weight:800; }
  .alert-del { background:none; border:none; color:#475569; cursor:pointer; font-size:13px; padding:0 2px; }
  .alert-del:hover { color:#f87171; }
  .alert-inputs { display:flex; gap:5px; margin-top:8px; align-items:center; }
  .alert-in { background:#0f172a; border:1px solid #334155; border-radius:6px;
    padding:4px 7px; color:#e2e8f0; font-size:12px; }
  .alert-in-sym { width:52px; text-transform:uppercase; }
  .alert-in-px  { width:62px; }
  .alert-add-btn { background:#1d4ed8; border:none; border-radius:6px; color:#fff;
    font-size:11px; font-weight:700; padding:4px 8px; cursor:pointer; }
  .alert-add-btn:hover { background:#2563eb; }

  /* ── Simulation Tab ── */
  .sim-summary {
    display: grid; grid-template-columns: repeat(auto-fill, minmax(160px, 1fr));
    gap: 12px; margin-bottom: 28px;
  }
  .sim-stat-card {
    background: #1e293b; border: 1px solid #334155; border-radius: 10px;
    padding: 14px 16px;
  }
  .sim-stat-card .label { font-size: 10px; color: #64748b; text-transform: uppercase;
    letter-spacing: .6px; margin-bottom: 4px; }
  .sim-stat-card .val   { font-size: 20px; font-weight: 800; color: #f1f5f9; }
  .sim-stat-card .sub   { font-size: 11px; color: #64748b; margin-top: 2px; }
  .sim-add-form {
    background: #1e293b; border: 1px solid #334155; border-radius: 12px;
    padding: 20px; margin-bottom: 24px;
    display: grid; grid-template-columns: 1fr 1fr 1fr 1fr auto; gap: 10px; align-items: end;
  }
  .sim-add-form label { font-size: 11px; color: #64748b; display: block; margin-bottom: 4px; }
  .sim-add-form input {
    width: 100%; background: #0f172a; border: 1px solid #334155; border-radius: 6px;
    padding: 8px 10px; color: #e2e8f0; font-size: 13px; box-sizing: border-box;
  }
  .sim-add-form input:focus { outline: none; border-color: #3b82f6; }
  .sim-add-btn {
    background: #1d4ed8; color: #fff; border: none; border-radius: 8px;
    padding: 9px 18px; font-size: 13px; font-weight: 600; cursor: pointer; white-space: nowrap;
  }
  .sim-add-btn:hover { background: #2563eb; }
  .sim-table { width: 100%; border-collapse: collapse; font-size: 13px; }
  .sim-table th { color: #475569; font-weight: 600; text-align: left;
    padding: 10px 12px; border-bottom: 2px solid #334155; font-size: 11px;
    text-transform: uppercase; letter-spacing: .5px; }
  .sim-table td { padding: 12px; border-bottom: 1px solid #1e293b; vertical-align: middle; }
  .sim-table tr:hover td { background: #1e293b; }
  .sim-sym { font-size: 15px; font-weight: 800; color: #60a5fa; }
  .sim-name { font-size: 10px; color: #475569; margin-top: 2px; }
  .sim-remove { background: none; border: none; color: #475569; cursor: pointer;
    font-size: 16px; padding: 2px 6px; border-radius: 4px; }
  .sim-remove:hover { color: #f87171; background: #1e293b; }
  .quick-add-btn {
    background: #1e3a5f; color: #93c5fd; border: 1px solid #1d4ed8;
    border-radius: 6px; padding: 3px 9px; font-size: 11px; font-weight: 600;
    cursor: pointer; white-space: nowrap;
  }
  .quick-add-btn:hover { background: #1d4ed8; color: #fff; }
  .quick-add-btn.added { background: #14532d; color: #4ade80; border-color: #16a34a; }
  @media (max-width: 700px) {
    .sim-add-form { grid-template-columns: 1fr 1fr; }
  }

  /* ── Ticker Hover Tooltip ── */
  .ticker-tip {
    position: fixed; z-index: 9999; pointer-events: none;
    background: #0f172a; border: 1px solid #334155; border-radius: 12px;
    width: 260px; padding: 14px 16px; box-shadow: 0 8px 32px rgba(0,0,0,.6);
    opacity: 0; transform: translateY(6px);
    transition: opacity .15s ease, transform .15s ease;
  }
  .ticker-tip.visible { opacity: 1; transform: translateY(0); }
  .ticker-tip-header { display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 8px; }
  .ticker-tip-sym { font-size: 18px; font-weight: 800; color: #60a5fa; }
  .ticker-tip-name { font-size: 11px; color: #94a3b8; margin-top: 2px; max-width: 180px;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .ticker-tip-sector { font-size: 10px; color: #475569; margin-top: 1px; }
  .ticker-tip-price { font-size: 16px; font-weight: 700; color: #f1f5f9; text-align: right; }
  .ticker-tip-chg { font-size: 11px; font-weight: 600; text-align: right; margin-top: 1px; }
  .ticker-tip-spark { margin: 8px 0; height: 40px; }
  .ticker-tip-spark canvas { width: 100% !important; height: 40px !important; }
  .ticker-tip-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 6px 8px; margin-top: 8px; }
  .ticker-tip-stat { }
  .ticker-tip-stat .k { font-size: 9px; color: #475569; text-transform: uppercase; letter-spacing: .6px; }
  .ticker-tip-stat .v { font-size: 12px; font-weight: 600; color: #e2e8f0; margin-top: 1px; }
  .ticker-tip-rec { display: inline-block; margin-top: 8px; padding: 2px 8px; border-radius: 99px;
    font-size: 10px; font-weight: 700; letter-spacing: .4px; }
  [data-sym] { cursor: pointer; }

  /* ── Yield Curve Sidebar (Economy tab) ── */
  .economy-layout {
    display: grid;
    grid-template-columns: 1fr 280px;
    gap: 24px;
    align-items: start;
  }
  @media (max-width: 900px) {
    .economy-layout { grid-template-columns: 1fr; }
  }
  .yield-sidebar {
    position: sticky;
    top: 56px;
    background: #1e293b;
    border: 1px solid #334155;
    border-radius: 12px;
    padding: 18px 16px;
    display: flex;
    flex-direction: column;
    gap: 18px;
  }
  .yield-sidebar h3 {
    font-size: 12px; font-weight: 700; color: #64748b;
    letter-spacing: 1px; text-transform: uppercase; margin-bottom: 2px;
  }
  .spot-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 5px 0;
    border-bottom: 1px solid #334155;
    font-size: 13px;
  }
  .spot-row:last-child { border-bottom: none; }
  .spot-label { color: #94a3b8; font-weight: 500; }
  .spot-val   { font-weight: 700; font-variant-numeric: tabular-nums; }

  /* ── News sub-tab navigation ── */
  .news-subnav {
    display: flex;
    gap: 10px;
    padding: 20px 28px 0;
    border-bottom: 2px solid #1e293b;
    margin-bottom: 0;
  }
  .news-subtab-btn {
    background: transparent;
    border: none;
    border-bottom: 3px solid transparent;
    color: #94a3b8;
    font-size: 15px;
    font-weight: 600;
    padding: 10px 20px 12px;
    cursor: pointer;
    transition: color .2s, border-color .2s;
    margin-bottom: -2px;
    letter-spacing: .3px;
  }
  .news-subtab-btn:hover { color: #e2e8f0; }
  .news-subtab-btn.active {
    color: #38bdf8;
    border-bottom-color: #38bdf8;
  }
  .news-subtab-pane { display: none; }
  .news-subtab-pane.active { display: block; }
</style>
<script>
function switchTab(tabId) {
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.querySelectorAll('.tab-pane').forEach(p => p.classList.remove('active'));
  document.querySelector('[data-tab="' + tabId + '"]').classList.add('active');
  document.getElementById(tabId).classList.add('active');
}

// ── S&P 500 Sector Treemap ──
const SP500_DATA = {{ sp500_sector_json | safe }};

function openSectorMap(gicsName, label, pct) {
  var stocks = SP500_DATA[gicsName];
  if (!stocks || stocks.length === 0) return;

  var modal = document.getElementById('sector-modal');
  modal.classList.add('open');
  document.getElementById('sector-modal-title').textContent = label + ' — ' + gicsName;
  document.getElementById('sector-modal-subtitle').textContent =
    stocks.length + ' S&P 500 companies  ·  ETF today: ' +
    (pct >= 0 ? '+' : '') + pct.toFixed(2) + '%';

  var labels  = stocks.map(function(s){ return s.t; });
  var values  = stocks.map(function(s){ return Math.max(s.m, 1e8); });
  var colors  = stocks.map(function(s){ return s.c; });
  var custom  = stocks.map(function(s){ return [s.n, s.c, s.s]; });

  // Color scale: dark red (-3%) → neutral → dark green (+3%)
  var colorscale = [
    [0.00, '#7f1d1d'],
    [0.30, '#b91c1c'],
    [0.45, '#3a1212'],
    [0.50, '#1e293b'],
    [0.55, '#0f2918'],
    [0.70, '#166534'],
    [1.00, '#14532d'],
  ];

  var trace = {
    type: 'treemap',
    labels: labels,
    parents: labels.map(function(){ return ''; }),
    values: values,
    customdata: custom,
    texttemplate: '<b>%{label}</b><br>%{customdata[1]:.2f}%',
    hovertemplate:
      '<b>%{label}</b> — %{customdata[0]}<br>' +
      '%{customdata[2]}<br>' +
      '<b>%{customdata[1]:.2f}%</b> today<extra></extra>',
    marker: {
      colors: colors,
      colorscale: colorscale,
      cmin: -3, cmax: 3,
      showscale: false,
      line: { width: 1.5, color: '#0f172a' },
    },
    tiling: { packing: 'squarify', squarifyratio: 1.4 },
    textfont: { size: 12, color: '#ffffff' },
    pathbar: { visible: false },
  };

  var layout = {
    paper_bgcolor: '#0f172a',
    plot_bgcolor:  '#0f172a',
    margin: { t: 4, l: 4, r: 4, b: 4 },
    font: { color: '#e2e8f0', family: '-apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif' },
  };

  Plotly.react('sector-treemap', [trace], layout,
    { responsive: true, displayModeBar: false });
}

function closeSectorMap() {
  document.getElementById('sector-modal').classList.remove('open');
}

// Close on overlay click
document.addEventListener('DOMContentLoaded', function() {
  document.getElementById('sector-modal').addEventListener('click', function(e) {
    if (e.target === this) closeSectorMap();
  });
  document.getElementById('holdings-modal').addEventListener('click', function(e) {
    if (e.target === this) closeHoldings();
  });
});

function openHoldings(sym, name) {
  var rows = ETF_HOLDINGS[sym] || [];
  var modal = document.getElementById('holdings-modal');
  document.getElementById('holdings-modal-title').textContent = sym + ' — Top Holdings';
  document.getElementById('holdings-modal-sub').textContent = name;
  var tbody = document.getElementById('holdings-tbody');
  if (rows.length === 0) {
    tbody.innerHTML = '<tr><td colspan="3" style="text-align:center;color:#64748b;padding:20px">No holdings data available</td></tr>';
  } else {
    var maxPct = Math.max.apply(null, rows.map(function(r){ return r.pct; }));
    tbody.innerHTML = rows.map(function(r, i) {
      var barW = maxPct > 0 ? Math.round(r.pct / maxPct * 80) : 0;
      return '<tr>' +
        '<td style="color:#64748b;width:28px">' + (i+1) + '</td>' +
        '<td><strong style="color:#38bdf8">' + (r.symbol || '—') + '</strong>' +
          '<span style="color:#94a3b8;font-size:11px;margin-left:6px">' + (r.name || '') + '</span></td>' +
        '<td style="text-align:right;white-space:nowrap">' +
          r.pct.toFixed(2) + '%' +
          '<span class="holdings-bar" style="width:' + barW + 'px"></span>' +
        '</td>' +
      '</tr>';
    }).join('');
  }
  modal.classList.add('open');
}
function closeHoldings() {
  document.getElementById('holdings-modal').classList.remove('open');
}

// ── Simulation Tab ────────────────────────────────────────────────────────────
var SIM_KEY = 'sim_positions_v1';

function simLoad() {
  try { return JSON.parse(localStorage.getItem(SIM_KEY) || '[]'); } catch(e) { return []; }
}
function simSave(positions) {
  localStorage.setItem(SIM_KEY, JSON.stringify(positions));
}

function simCurrentPrice(sym) {
  var info = STOCK_INFO[sym];
  if (info && info.price) return info.price;
  // fallback: last point in chart history
  var cd = CHART_DATA[sym];
  if (cd && cd.history && cd.history.length) return cd.history[cd.history.length-1][1];
  return null;
}

function simSpyReturn(entryDateStr) {
  // Returns SPY % return from entry date to now using CHART_DATA["SPY"]
  var cd = CHART_DATA['SPY'];
  if (!cd || !cd.history || !cd.history.length) return null;
  var entryTs = new Date(entryDateStr).getTime();
  var pts = cd.history;
  // find closest point on or after entry date
  var entryPt = null;
  for (var i = 0; i < pts.length; i++) {
    if (pts[i][0] >= entryTs) { entryPt = pts[i]; break; }
  }
  if (!entryPt) entryPt = pts[0];
  var curPt = pts[pts.length - 1];
  if (!entryPt[1] || !curPt[1]) return null;
  return (curPt[1] - entryPt[1]) / entryPt[1] * 100;
}

function simAddFromForm() {
  var sym    = (document.getElementById('sim-sym-input').value || '').toUpperCase().trim();
  var price  = parseFloat(document.getElementById('sim-price-input').value);
  var shares = parseFloat(document.getElementById('sim-shares-input').value);
  var date   = document.getElementById('sim-date-input').value;
  if (!sym) { alert('Please enter a ticker symbol.'); return; }
  if (!price || price <= 0) { alert('Please enter a valid entry price.'); return; }
  if (!shares || shares <= 0) { alert('Please enter a valid number of shares.'); return; }
  if (!date) { date = new Date().toISOString().slice(0,10); }
  simAdd(sym, price, shares, date);
  document.getElementById('sim-sym-input').value   = '';
  document.getElementById('sim-price-input').value = '';
  document.getElementById('sim-shares-input').value = '';
}

function simAdd(sym, entryPrice, shares, entryDate, note) {
  var positions = simLoad();
  positions.push({ sym: sym, entryPrice: entryPrice, shares: shares,
                   entryDate: entryDate || new Date().toISOString().slice(0,10),
                   note: note || '', addedAt: Date.now() });
  simSave(positions);
  renderSim();
  // Mark quick-add button as added
  document.querySelectorAll('.quick-add-btn[data-sym="' + sym + '"]').forEach(function(b) {
    b.textContent = '✓ Added'; b.classList.add('added');
  });
}

function simRemove(idx) {
  var positions = simLoad();
  positions.splice(idx, 1);
  simSave(positions);
  renderSim();
}

function simFmtPct(v) {
  if (v === null || v === undefined || isNaN(v)) return '—';
  return (v >= 0 ? '+' : '') + v.toFixed(2) + '%';
}
function simFmtDollar(v) {
  if (v === null || v === undefined || isNaN(v)) return '—';
  return (v >= 0 ? '+$' : '-$') + Math.abs(v).toLocaleString('en-US', {minimumFractionDigits:2, maximumFractionDigits:2});
}

function renderSim() {
  var positions = simLoad();

  // Mark quick-add buttons
  var inSim = {};
  positions.forEach(function(p) { inSim[p.sym] = true; });
  document.querySelectorAll('.quick-add-btn').forEach(function(b) {
    if (inSim[b.dataset.sym]) { b.textContent = '✓ Added'; b.classList.add('added'); }
    else { b.textContent = '+ Sim'; b.classList.remove('added'); }
  });

  // Summary stats
  var totalCost = 0, totalCurrent = 0, winCount = 0;
  var validPositions = positions.map(function(p, i) {
    var cur = simCurrentPrice(p.sym);
    var cost = p.entryPrice * p.shares;
    var curVal = cur ? cur * p.shares : null;
    var pnlDollar = curVal !== null ? curVal - cost : null;
    var pnlPct    = pnlDollar !== null ? pnlDollar / cost * 100 : null;
    var spyRet    = simSpyReturn(p.entryDate);
    var alpha     = (pnlPct !== null && spyRet !== null) ? pnlPct - spyRet : null;
    if (curVal !== null) { totalCost += cost; totalCurrent += curVal; }
    if (pnlPct !== null && pnlPct > 0) winCount++;
    return Object.assign({}, p, { i:i, cur:cur, cost:cost, curVal:curVal,
                                   pnlDollar:pnlDollar, pnlPct:pnlPct, spyRet:spyRet, alpha:alpha });
  });

  var totalPnl  = totalCurrent - totalCost;
  var totalPct  = totalCost > 0 ? totalPnl / totalCost * 100 : null;

  // Render summary
  var summaryEl = document.getElementById('sim-summary');
  if (positions.length === 0) {
    summaryEl.innerHTML = '';
  } else {
    var spyOverall = validPositions.length ? simSpyReturn(validPositions[0].entryDate) : null;
    summaryEl.innerHTML = [
      ['Sim Portfolio', '$' + totalCurrent.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2}), ''],
      ['Total P&L', simFmtDollar(totalPnl), totalPnl >= 0 ? 'color:#34d399' : 'color:#f87171'],
      ['Total Return', simFmtPct(totalPct), totalPct >= 0 ? 'color:#34d399' : 'color:#f87171'],
      ['Positions', positions.length + ' open', ''],
      ['Winners', winCount + ' / ' + positions.length, ''],
    ].map(function(s) {
      return '<div class="sim-stat-card"><div class="label">' + s[0] + '</div>' +
             '<div class="val" style="' + s[2] + '">' + s[1] + '</div></div>';
    }).join('');
  }

  // Render table
  var wrap = document.getElementById('sim-table-wrap');
  if (positions.length === 0) {
    wrap.innerHTML = '<div style="text-align:center;color:#475569;padding:48px 0;font-size:14px">' +
      'No positions yet — add one above or click <strong style="color:#93c5fd">+ Sim</strong> on a recommendation.</div>';
    return;
  }

  var rows = validPositions.map(function(p) {
    var info = STOCK_INFO[p.sym] || {};
    var name = info.name || p.sym;
    var curStr  = p.cur  ? '$' + p.cur.toFixed(2) : '—';
    var pnlCol  = p.pnlPct === null ? '' : (p.pnlPct >= 0 ? 'color:#34d399' : 'color:#f87171');
    var alphaCol = p.alpha === null ? '' : (p.alpha >= 0 ? 'color:#34d399' : 'color:#f87171');
    var spyStr  = p.spyRet !== null ? simFmtPct(p.spyRet) : '—';
    var alphaStr = p.alpha !== null ? simFmtPct(p.alpha) : '—';
    return '<tr>' +
      '<td><div class="sim-sym" data-sym="' + p.sym + '">' + p.sym + '</div>' +
          '<div class="sim-name">' + name.substring(0,30) + '</div></td>' +
      '<td>' + p.entryDate + '</td>' +
      '<td>$' + p.entryPrice.toFixed(2) + '</td>' +
      '<td>' + curStr + '</td>' +
      '<td>' + p.shares + '</td>' +
      '<td><strong style="' + pnlCol + '">' + simFmtDollar(p.pnlDollar) + '</strong></td>' +
      '<td><strong style="' + pnlCol + '">' + simFmtPct(p.pnlPct) + '</strong></td>' +
      '<td style="color:#94a3b8">' + spyStr + '</td>' +
      '<td style="' + alphaCol + '">' + alphaStr + '</td>' +
      '<td><button class="sim-remove" onclick="simRemove(' + p.i + ')" title="Remove">✕</button></td>' +
    '</tr>';
  }).join('');

  wrap.innerHTML = '<table class="sim-table">' +
    '<thead><tr>' +
      '<th>Ticker</th><th>Entry Date</th><th>Entry $</th><th>Current $</th>' +
      '<th>Shares</th><th>P&L $</th><th>P&L %</th>' +
      '<th>SPY</th><th>vs SPY</th><th></th>' +
    '</tr></thead><tbody>' + rows + '</tbody></table>';
}

// Set today's date as default in the form
document.addEventListener('DOMContentLoaded', function() {
  var d = document.getElementById('sim-date-input');
  if (d) d.value = new Date().toISOString().slice(0,10);
});

// ── Earnings Calendar ────────────────────────────────────────────────────────
function renderEarnings() {
  var el = document.getElementById('earnings-tile');
  if (!el) return;
  if (!EARNINGS_UPCOMING || !EARNINGS_UPCOMING.length) {
    el.innerHTML = '<div style="font-size:12px;color:#475569;margin-top:6px">None in next 30 days</div>';
    return;
  }
  var html = '';
  var shown = EARNINGS_UPCOMING.slice(0, 6);
  for (var i = 0; i < shown.length; i++) {
    var e = shown[i];
    var cls = e.days === 0 ? 'earn-today' : (e.days <= 3 ? 'earn-soon' : 'earn-normal');
    var label = e.days === 0 ? 'TODAY' : (e.days === 1 ? 'Tomorrow' : (e.days + 'd - ' + e.date));
    html += '<div class="earn-row">';
    html += '<span class="earn-sym" data-sym="' + e.sym + '">' + e.sym + '</span>';
    html += '<span class="' + cls + '">' + label + '</span>';
    html += '</div>';
  }
  el.innerHTML = html;
}

// ── Price Alerts ─────────────────────────────────────────────────────────────
var ALERTS_KEY = 'price_alerts_v1';

function alertLoad() {
  try { return JSON.parse(localStorage.getItem(ALERTS_KEY) || '[]'); } catch(e) { return []; }
}
function alertSave(arr) { localStorage.setItem(ALERTS_KEY, JSON.stringify(arr)); }

function alertAdd() {
  var sym = (document.getElementById('al-sym').value || '').toUpperCase().trim();
  var px  = parseFloat(document.getElementById('al-px').value);
  if (!sym || isNaN(px) || px <= 0) return;
  var arr = alertLoad().filter(function(a) { return a.sym !== sym; });
  arr.push({ sym: sym, target: px });
  alertSave(arr);
  document.getElementById('al-sym').value = '';
  document.getElementById('al-px').value  = '';
  renderAlerts();
}

function alertRemove(sym) {
  alertSave(alertLoad().filter(function(a) { return a.sym !== sym; }));
  renderAlerts();
}

function renderAlerts() {
  var el = document.getElementById('alerts-tile');
  if (!el) return;
  var arr = alertLoad();
  if (!arr.length) {
    el.innerHTML = '<div style="font-size:12px;color:#475569;margin-top:4px">No alerts set</div>';
    return;
  }
  var html = '';
  for (var i = 0; i < arr.length; i++) {
    var a    = arr[i];
    var info = STOCK_INFO[a.sym];
    var cur  = info ? info.price : null;
    var near = cur !== null && Math.abs(cur - a.target) / a.target <= 0.02;
    var priceTxt = cur !== null ? ('$' + cur.toFixed(2)) : '--';
    var lbl = near ? (priceTxt + ' near target') : ('target $' + a.target.toFixed(2));
    html += '<div class="alert-row-item">';
    html += '<span class="alert-sym-lbl" data-sym="' + a.sym + '">' + a.sym + '</span>';
    html += '<span class="' + (near ? 'alert-hit' : 'alert-price-lbl') + '">' + lbl + '</span>';
    html += '<button class="alert-del" data-sym="' + a.sym + '" onclick="alertRemove(this.dataset.sym)">x</button>';
    html += '</div>';
  }
  el.innerHTML = html;
}

// ── Live Price Refresh ───────────────────────────────────────────────────────
function refreshPrices(manual) {
  var statusEl = document.getElementById('refresh-status');
  var btnEl    = document.getElementById('refresh-btn');
  if (statusEl) { statusEl.textContent = 'Fetching…'; statusEl.style.color = '#94a3b8'; }
  if (btnEl)    { btnEl.disabled = true; }

  // Collect all symbols from STOCK_INFO
  var syms = Object.keys(STOCK_INFO);
  if (!syms.length) {
    if (statusEl) statusEl.textContent = 'No symbols';
    if (btnEl) btnEl.disabled = false;
    return;
  }

  var url = 'https://query1.finance.yahoo.com/v7/finance/quote?symbols=' + syms.join(',') +
            '&fields=regularMarketPrice&formatted=false&lang=en-US&region=US';

  fetch(url, { mode: 'cors' })
    .then(function(res) {
      if (!res.ok) throw new Error('HTTP ' + res.status);
      return res.json();
    })
    .then(function(data) {
      var results = data && data.quoteResponse && data.quoteResponse.result;
      if (!results || !results.length) throw new Error('empty response');

      var updated = 0;
      for (var i = 0; i < results.length; i++) {
        var q   = results[i];
        var sym = q.symbol;
        var px  = q.regularMarketPrice;
        if (!sym || px == null) continue;
        if (STOCK_INFO[sym]) { STOCK_INFO[sym].price = px; updated++; }
      }

      // Update holdings table cells
      var rows = document.querySelectorAll('tr[data-holding-sym]');
      for (var r = 0; r < rows.length; r++) {
        var row    = rows[r];
        var sym    = row.getAttribute('data-holding-sym');
        var shares = parseFloat(row.getAttribute('data-holding-shares')) || 0;
        var cost   = parseFloat(row.getAttribute('data-holding-cost'))   || 0;
        var info   = STOCK_INFO[sym];
        if (!info || info.price == null) continue;
        var px = info.price;

        var priceCell = document.querySelector('td[data-price-cell="' + sym + '"]');
        if (priceCell) priceCell.textContent = '$' + px.toFixed(2);

        var mv     = px * shares;
        var mvCell = document.querySelector('td[data-mv-cell="' + sym + '"]');
        if (mvCell) mvCell.textContent = '$' + Math.round(mv).toLocaleString();

        var glCell = document.querySelector('td[data-gl-cell="' + sym + '"]');
        if (glCell && cost > 0) {
          var glPct = (mv - cost) / cost * 100;
          var sign  = glPct >= 0 ? '+' : '';
          glCell.innerHTML = '<span class="' + (glPct >= 0 ? 'positive' : 'negative') + '">' +
                             sign + glPct.toFixed(1) + '%</span>';
        }
      }

      // Re-render sim and alerts with updated prices
      if (typeof renderSim   === 'function') renderSim();
      if (typeof renderAlerts === 'function') renderAlerts();

      var now = new Date();
      var ts  = now.getHours() + ':' + String(now.getMinutes()).padStart(2, '0') + ':' +
                String(now.getSeconds()).padStart(2, '0');
      if (statusEl) { statusEl.textContent = 'Updated ' + ts; statusEl.style.color = '#4ade80'; }
      if (btnEl)    { btnEl.disabled = false; }
    })
    .catch(function(err) {
      if (statusEl) { statusEl.textContent = 'Offline / CORS'; statusEl.style.color = '#f87171'; }
      if (btnEl)    { btnEl.disabled = false; }
      if (manual)   { console.warn('[refreshPrices] Error:', err); }
    });
}

document.addEventListener('DOMContentLoaded', function() {
  renderEarnings();
  renderAlerts();
  refreshPrices(false);
  var inputs = ['al-sym', 'al-px'];
  for (var i = 0; i < inputs.length; i++) {
    (function(id) {
      var inp = document.getElementById(id);
      if (inp) { inp.addEventListener('keydown', function(e) { if (e.key === 'Enter') alertAdd(); }); }
    })(inputs[i]);
  }
});

// ── Ticker Hover Tooltip ─────────────────────────────────────────────────────
(function() {
  var tip = null, sparkChart = null, hoverTimer = null;

  function fmtMktCap(v) {
    if (!v) return '—';
    if (v >= 1e12) return '$' + (v/1e12).toFixed(2) + 'T';
    if (v >= 1e9)  return '$' + (v/1e9).toFixed(1) + 'B';
    if (v >= 1e6)  return '$' + (v/1e6).toFixed(0) + 'M';
    return '$' + v.toLocaleString();
  }

  function recColor(rec) {
    if (!rec) return null;
    var r = rec.toLowerCase();
    if (r.includes('strong buy') || r.includes('buy')) return {bg:'#14532d', c:'#4ade80'};
    if (r.includes('hold')) return {bg:'#1c3a5f', c:'#60a5fa'};
    if (r.includes('sell')) return {bg:'#450a0a', c:'#f87171'};
    return null;
  }

  function drawSparkline(sym) {
    var canvas = document.getElementById('tt-spark');
    var ctx = canvas.getContext('2d');
    canvas.width = 228; canvas.height = 40;
    ctx.clearRect(0, 0, 228, 40);
    var cd = CHART_DATA[sym];
    if (!cd) return;
    var pts = (cd.history && cd.history.length > 1) ? cd.history : (cd.intraday || []);
    if (pts.length < 2) return;
    var vals = pts.map(function(p){ return p[1]; });
    var mn = Math.min.apply(null, vals), mx = Math.max.apply(null, vals);
    if (mx === mn) mx = mn + 1;
    var w = 228, h = 40, pad = 3;
    var xStep = (w - pad*2) / (vals.length - 1);
    var positive = vals[vals.length-1] >= vals[0];
    var grad = ctx.createLinearGradient(0, 0, 0, h);
    grad.addColorStop(0, positive ? 'rgba(52,211,153,.3)' : 'rgba(248,113,113,.3)');
    grad.addColorStop(1, 'rgba(0,0,0,0)');
    ctx.beginPath();
    vals.forEach(function(v, i) {
      var x = pad + i * xStep;
      var y = h - pad - (v - mn) / (mx - mn) * (h - pad*2);
      i === 0 ? ctx.moveTo(x, y) : ctx.lineTo(x, y);
    });
    var lastX = pad + (vals.length-1)*xStep;
    ctx.lineTo(lastX, h); ctx.lineTo(pad, h); ctx.closePath();
    ctx.fillStyle = grad; ctx.fill();
    ctx.beginPath();
    vals.forEach(function(v, i) {
      var x = pad + i * xStep;
      var y = h - pad - (v - mn) / (mx - mn) * (h - pad*2);
      i === 0 ? ctx.moveTo(x, y) : ctx.lineTo(x, y);
    });
    ctx.strokeStyle = positive ? '#34d399' : '#f87171';
    ctx.lineWidth = 1.5; ctx.stroke();
  }

  function showTip(sym, anchorEl) {
    if (!tip) tip = document.getElementById('ticker-tip');
    var info = STOCK_INFO[sym];
    if (!info) return;

    document.getElementById('tt-sym').textContent = sym;
    document.getElementById('tt-name').textContent = info.name || '';
    document.getElementById('tt-sector').textContent = info.sector || '';

    // Price + daily change from chart data
    var cd = CHART_DATA[sym];
    var price = info.price;
    var chgPct = null;
    if (cd) {
      var pts = (cd.intraday && cd.intraday.length > 1) ? cd.intraday
              : (cd.history && cd.history.length > 1) ? cd.history.slice(-2) : null;
      if (pts && pts.length >= 2) {
        var first = pts[0][1], last = pts[pts.length-1][1];
        if (first) { chgPct = (last - first) / first * 100; price = price || last; }
      }
    }
    document.getElementById('tt-price').textContent = price ? '$' + price.toFixed(2) : '—';
    var chgEl = document.getElementById('tt-chg');
    if (chgPct !== null) {
      chgEl.textContent = (chgPct >= 0 ? '+' : '') + chgPct.toFixed(2) + '%';
      chgEl.style.color = chgPct >= 0 ? '#34d399' : '#f87171';
    } else { chgEl.textContent = ''; }

    // Sparkline
    drawSparkline(sym);

    // Stats grid
    var stats = [];
    if (info.isEtf) {
      if (info.expenseRatio != null) stats.push(['Exp. Ratio', info.expenseRatio.toFixed(2) + '%']);
      if (info.threeYr != null)      stats.push(['3Y Return',  (info.threeYr >= 0 ? '+' : '') + info.threeYr.toFixed(1) + '%']);
      if (info.marketCap)            stats.push(['AUM',        fmtMktCap(info.marketCap)]);
      if (info.beta != null)         stats.push(['Beta',       info.beta.toFixed(2)]);
    } else {
      if (info.marketCap)            stats.push(['Mkt Cap',    fmtMktCap(info.marketCap)]);
      if (info.pe != null)           stats.push(['P/E',        info.pe.toFixed(1)]);
      if (info.beta != null)         stats.push(['Beta',       info.beta.toFixed(2)]);
      if (info.analystTarget != null) {
        var upStr = info.upside != null ? ' (' + (info.upside > 0 ? '+' : '') + info.upside.toFixed(1) + '%)' : '';
        stats.push(['Target', '$' + info.analystTarget.toFixed(2) + upStr]);
      }
    }
    var grid = document.getElementById('tt-grid');
    grid.innerHTML = stats.map(function(s) {
      return '<div class="ticker-tip-stat"><div class="k">' + s[0] + '</div><div class="v">' + s[1] + '</div></div>';
    }).join('');

    // Recommendation badge
    var recEl = document.getElementById('tt-rec');
    var rc = recColor(info.rec);
    if (rc && info.rec) {
      recEl.innerHTML = '<span class="ticker-tip-rec" style="background:' + rc.bg + ';color:' + rc.c + '">' + info.rec + '</span>';
    } else { recEl.innerHTML = ''; }

    // Position tooltip
    var rect = anchorEl.getBoundingClientRect();
    var tx = rect.left, ty = rect.bottom + 8;
    if (tx + 260 > window.innerWidth - 12) tx = window.innerWidth - 272;
    if (ty + 220 > window.innerHeight - 12) ty = rect.top - 228;
    tip.style.left = tx + 'px'; tip.style.top = ty + 'px';
    tip.classList.add('visible');
  }

  function hideTip() {
    if (tip) tip.classList.remove('visible');
  }

  document.addEventListener('mouseover', function(e) {
    var el = e.target.closest('[data-sym]');
    if (!el) return;
    clearTimeout(hoverTimer);
    hoverTimer = setTimeout(function() { showTip(el.dataset.sym, el); }, 120);
  });
  document.addEventListener('mouseout', function(e) {
    var el = e.target.closest('[data-sym]');
    if (!el) return;
    clearTimeout(hoverTimer);
    hideTip();
  });
})();

function switchNewsTab(subtabId) {
  document.querySelectorAll('.news-subtab-btn').forEach(b => b.classList.remove('active'));
  document.querySelectorAll('.news-subtab-pane').forEach(p => p.classList.remove('active'));
  document.querySelector('[data-subtab="' + subtabId + '"]').classList.add('active');
  document.getElementById(subtabId).classList.add('active');
  // Render yield curve chart when Economy tab becomes visible
  if (subtabId === 'subtab-economy') renderYieldCurve();
}

// ── Yield Curve Chart ──
const YIELD_DATA = {{ market_overview.yield_curve | tojson }};
var _yieldChartInst = null;
function renderYieldCurve() {
  var canvas = document.getElementById('yield-chart');
  if (!canvas || !YIELD_DATA || YIELD_DATA.length === 0) return;
  if (_yieldChartInst) { _yieldChartInst.destroy(); _yieldChartInst = null; }
  var labels = YIELD_DATA.map(function(p){ return p.label; });
  var vals   = YIELD_DATA.map(function(p){ return p.value; });
  // Color the line: inverted curve (short > long) → red tint; normal → teal
  var inverted = vals.length >= 2 && vals[0] > vals[vals.length-1];
  var lineColor = inverted ? '#f87171' : '#38bdf8';
  _yieldChartInst = new Chart(canvas.getContext('2d'), {
    type: 'line',
    data: {
      labels: labels,
      datasets: [{
        data: vals,
        borderColor: lineColor,
        backgroundColor: lineColor + '22',
        borderWidth: 2.5,
        pointRadius: 5,
        pointBackgroundColor: lineColor,
        fill: true,
        tension: 0.35,
      }]
    },
    options: {
      responsive: true,
      animation: { duration: 600 },
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: { label: function(ctx){ return ctx.parsed.y.toFixed(2) + '%'; } }
        }
      },
      scales: {
        x: { grid: { color: '#1e293b' }, ticks: { color: '#64748b', font: { size: 10 } } },
        y: {
          grid: { color: '#1e293b' },
          ticks: { color: '#64748b', font: { size: 10 },
                   callback: function(v){ return v.toFixed(2) + '%'; } },
          suggestedMin: Math.max(0, Math.min(...vals) - 0.3),
          suggestedMax: Math.max(...vals) + 0.3,
        }
      }
    }
  });
}
// Auto-render if Economy tab is the default visible pane
document.addEventListener('DOMContentLoaded', function() {
  if (document.getElementById('subtab-economy') &&
      document.getElementById('subtab-economy').classList.contains('active')) {
    renderYieldCurve();
  }
});

function toggleDetail(btn) {
  btn.classList.toggle('open');
  // The button is now inside a flex wrapper div — step up to find the detail panel
  const panel = btn.closest('div').nextElementSibling;
  panel.classList.toggle('open');
  // Animate the timeline progress bar on first open
  if (panel.classList.contains('open')) {
    panel.querySelectorAll('.timeline-bar').forEach(function(bar) {
      const pct = bar.dataset.pct || '0';
      bar.style.setProperty('--bar-end', pct + '%');
      bar.style.animation = 'none';
      bar.offsetHeight; // reflow
      bar.style.animation = 'fillbar 1.2s ease forwards';
    });
  }
}

// ── Embedded chart data ──────────────────────────────────────────
const CHART_DATA   = {{ chart_data_json | safe }};
const ETF_HOLDINGS = {{ etf_holdings_json | safe }};
const STOCK_INFO        = {{ stock_info_json | safe }};
const EARNINGS_UPCOMING = {{ earnings_json | safe }};

// ── Chart modal ──────────────────────────────────────────────────
let _chart = null, _chartSym = null, _chartTf = '1d';

const TF_LABELS = {
  '1d':'1D','5d':'5D','1m':'1M','3m':'3M','ytd':'YTD','1y':'1Y','all':'ALL'
};

function openChart(sym) {
  _chartSym = sym;
  const modal = document.getElementById('chart-modal');
  modal.style.display = 'flex';
  document.getElementById('chart-modal-sym').textContent = sym;
  // Build timeframe buttons
  const row = document.getElementById('chart-tf-row');
  row.innerHTML = '';
  ['1d','5d','1m','3m','ytd','1y','all'].forEach(tf => {
    const b = document.createElement('button');
    b.className = 'tf-btn' + (tf === _chartTf ? ' active' : '');
    b.textContent = TF_LABELS[tf];
    b.onclick = () => switchTf(tf);
    row.appendChild(b);
  });
  renderChart(sym, _chartTf);
}

function closeChart() {
  document.getElementById('chart-modal').style.display = 'none';
  if (_chart) { _chart.destroy(); _chart = null; }
}

function switchTf(tf) {
  _chartTf = tf;
  document.querySelectorAll('.tf-btn').forEach(b => {
    b.classList.toggle('active', b.textContent === TF_LABELS[tf]);
  });
  renderChart(_chartSym, tf);
}

function renderChart(sym, tf) {
  const raw = CHART_DATA[sym];
  if (!raw) { return; }

  // Select and slice dataset
  const now = Date.now();
  const D = 86400000;
  let pts = [];
  if (tf === '1d') {
    pts = raw.intraday || [];
  } else if (tf === '5d') {
    pts = (raw.history || []).slice(-5);
  } else if (tf === '1m') {
    pts = (raw.history || []).filter(p => p.t >= now - 30*D);
  } else if (tf === '3m') {
    pts = (raw.history || []).filter(p => p.t >= now - 91*D);
  } else if (tf === 'ytd') {
    const jan1 = new Date(new Date().getFullYear(),0,1).getTime();
    pts = (raw.history || []).filter(p => p.t >= jan1);
  } else if (tf === '1y') {
    pts = (raw.history || []).filter(p => p.t >= now - 365*D);
  } else {
    pts = raw.alltime || raw.history || [];
  }

  if (!pts.length) {
    document.getElementById('chart-modal-price').textContent = 'No data';
    document.getElementById('chart-modal-change').textContent = '';
    return;
  }

  const closes = pts.map(p => p.c);
  const first = closes[0], last = closes[closes.length-1];
  const chgAbs = last - first, chgPct = chgAbs / first * 100;
  const up = last >= first;
  const lineColor = up ? '#34d399' : '#f87171';

  document.getElementById('chart-modal-price').textContent = '$' + last.toFixed(2);
  const chEl = document.getElementById('chart-modal-change');
  chEl.textContent = (chgAbs >= 0 ? '+' : '') + '$' + Math.abs(chgAbs).toFixed(2) +
                     ' (' + (chgPct >= 0 ? '+' : '') + chgPct.toFixed(2) + '%)';
  chEl.style.color = lineColor;

  // Format x-axis labels
  const labels = pts.map(p => {
    const d = new Date(p.t);
    if (tf === '1d') {
      return d.toLocaleTimeString('en-US',{hour:'2-digit',minute:'2-digit',hour12:true});
    } else if (tf === '5d') {
      return d.toLocaleDateString('en-US',{weekday:'short',month:'short',day:'numeric'});
    } else if (['1m','3m'].includes(tf)) {
      return d.toLocaleDateString('en-US',{month:'short',day:'numeric'});
    } else if (tf === 'ytd' || tf === '1y') {
      return d.toLocaleDateString('en-US',{month:'short',day:'numeric'});
    } else {
      return d.toLocaleDateString('en-US',{month:'short',year:'2-digit'});
    }
  });

  const ctx = document.getElementById('chart-canvas').getContext('2d');
  if (_chart) _chart.destroy();

  const grad = ctx.createLinearGradient(0,0,0,320);
  grad.addColorStop(0, up ? 'rgba(52,211,153,0.25)' : 'rgba(248,113,113,0.25)');
  grad.addColorStop(1, 'rgba(15,23,42,0)');

  _chart = new Chart(ctx, {
    type: 'line',
    data: {
      labels,
      datasets: [{
        data: closes,
        borderColor: lineColor,
        backgroundColor: grad,
        borderWidth: 2,
        pointRadius: 0,
        pointHoverRadius: 4,
        pointHoverBackgroundColor: lineColor,
        fill: true,
        tension: 0.1
      }]
    },
    options: {
      responsive: true,
      maintainAspectRatio: false,
      interaction: { mode:'index', intersect:false },
      plugins: {
        legend: { display:false },
        tooltip: {
          backgroundColor:'#1e293b',
          borderColor:'#334155',
          borderWidth:1,
          titleColor:'#94a3b8',
          bodyColor:'#f1f5f9',
          padding:10,
          callbacks: {
            label: ctx => ' $' + ctx.parsed.y.toFixed(2)
          }
        }
      },
      scales: {
        x: {
          grid:{ color:'#1e293b', drawBorder:false },
          ticks:{ color:'#475569', maxTicksLimit:8, maxRotation:0, font:{size:11} },
          border:{ display:false }
        },
        y: {
          position:'right',
          grid:{ color:'#1e293b', drawBorder:false },
          ticks:{ color:'#475569', callback: v => '$'+v.toFixed(v<10?2:0), font:{size:11} },
          border:{ display:false }
        }
      }
    }
  });
}

// ── Password gate — pure JS, works on file:// and https://
(function() {
  const STORED = "{{ pass_b64 }}";
  function encode(str) {
    // UTF-8 safe base64
    return btoa(encodeURIComponent(str).replace(/%([0-9A-F]{2})/g,
      function(m, p) { return String.fromCharCode('0x' + p); }));
  }
  function unlock() {
    document.getElementById('gate').style.display = 'none';
    document.getElementById('app').style.display = 'block';
    sessionStorage.setItem('pa_unlocked', STORED);
  }
  window.checkPassword = function() {
    const val = document.getElementById('gate-input').value;
    const err = document.getElementById('gate-error');
    if (encode(val) === STORED) {
      unlock();
    } else {
      err.textContent = 'Incorrect password. Try again.';
      document.getElementById('gate-input').value = '';
      document.getElementById('gate-input').focus();
    }
  };
  window.addEventListener('DOMContentLoaded', function() {
    if (STORED === 'none') { unlock(); return; }
    if (sessionStorage.getItem('pa_unlocked') === STORED) { unlock(); return; }
    document.getElementById('gate').style.display = 'flex';
    document.getElementById('app').style.display = 'none';
    document.getElementById('gate-input').addEventListener('keydown', function(e) {
      if (e.key === 'Enter') window.checkPassword();
    });
  });
})();
</script>
</head>
<body>

<!-- Password gate overlay -->
<div id="gate" style="display:none;position:fixed;inset:0;background:#0f172a;
     z-index:9999;flex-direction:column;align-items:center;justify-content:center;gap:16px">
  <div style="font-size:32px">📈</div>
  <div style="font-size:22px;font-weight:700;color:#60a5fa">Portfolio Analysis Dashboard</div>
  <div style="font-size:14px;color:#64748b;margin-bottom:8px">This report is password protected</div>
  <input id="gate-input" type="password" placeholder="Enter password"
    style="background:#1e293b;border:1px solid #334155;color:#f1f5f9;
           padding:12px 20px;border-radius:8px;font-size:16px;width:280px;outline:none;
           text-align:center;letter-spacing:2px" autofocus>
  <button onclick="checkPassword()"
    style="background:#3b82f6;color:#fff;border:none;padding:11px 32px;
           border-radius:8px;font-size:15px;font-weight:600;cursor:pointer;width:280px">
    Unlock
  </button>
  <div id="gate-error" style="color:#f87171;font-size:13px;min-height:18px"></div>
  <div style="font-size:11px;color:#334155;margin-top:16px">
    Built with Python · Yahoo Finance · Not financial advice
  </div>
</div>

<!-- Chart modal -->
<div id="chart-modal" class="chart-modal-overlay" onclick="if(event.target===this)closeChart()">
  <div class="chart-modal-box">
    <button onclick="closeChart()"
            style="position:absolute;top:14px;right:18px;background:none;border:none;
                   color:#475569;font-size:22px;cursor:pointer;line-height:1">✕</button>

    <!-- Header -->
    <div style="display:flex;justify-content:space-between;align-items:flex-start;
                flex-wrap:wrap;gap:12px;margin-bottom:18px">
      <div>
        <div id="chart-modal-sym"
             style="font-size:22px;font-weight:800;color:#60a5fa;letter-spacing:-.01em"></div>
        <div style="display:flex;gap:12px;align-items:baseline;margin-top:5px">
          <span id="chart-modal-price"
                style="font-size:26px;font-weight:700;color:#f1f5f9"></span>
          <span id="chart-modal-change"
                style="font-size:14px;font-weight:600"></span>
        </div>
      </div>
      <div id="chart-tf-row" class="tf-row"></div>
    </div>

    <!-- Canvas -->
    <div style="height:340px;position:relative">
      <canvas id="chart-canvas"></canvas>
    </div>

    <p style="font-size:11px;color:#334155;margin-top:12px;text-align:center">
      Data via Yahoo Finance · Prices adjusted for splits & dividends · Not financial advice
    </p>
  </div>
</div>

<!-- Sector Treemap Modal -->
<div id="sector-modal" class="sector-modal-overlay">
  <div class="sector-modal">
    <div class="sector-modal-header">
      <div>
        <div class="sector-modal-title" id="sector-modal-title">Sector</div>
        <div class="sector-modal-subtitle" id="sector-modal-subtitle"></div>
      </div>
      <div style="display:flex;align-items:center;gap:16px">
        <div style="display:flex;gap:10px;align-items:center;font-size:11px;color:#475569">
          <span style="display:inline-flex;align-items:center;gap:4px">
            <span style="width:10px;height:10px;border-radius:2px;background:#14532d;display:inline-block"></span>
            +3%+
          </span>
          <span style="display:inline-flex;align-items:center;gap:4px">
            <span style="width:10px;height:10px;border-radius:2px;background:#166534;display:inline-block"></span>
            +1–3%
          </span>
          <span style="display:inline-flex;align-items:center;gap:4px">
            <span style="width:10px;height:10px;border-radius:2px;background:#1e293b;display:inline-block"></span>
            ~0%
          </span>
          <span style="display:inline-flex;align-items:center;gap:4px">
            <span style="width:10px;height:10px;border-radius:2px;background:#b91c1c;display:inline-block"></span>
            -1–3%
          </span>
          <span style="display:inline-flex;align-items:center;gap:4px">
            <span style="width:10px;height:10px;border-radius:2px;background:#7f1d1d;display:inline-block"></span>
            -3%+
          </span>
        </div>
        <button onclick="closeSectorMap()"
                style="background:none;border:1px solid #334155;border-radius:6px;
                       color:#94a3b8;padding:6px 14px;cursor:pointer;font-size:13px">
          ✕ Close
        </button>
      </div>
    </div>
    <div id="sector-treemap" style="height:70vh"></div>
  </div>
</div>

<!-- ETF Holdings Modal -->
<div id="holdings-modal" class="holdings-modal-overlay">
  <div class="holdings-modal">
    <div class="holdings-modal-header">
      <div>
        <div class="holdings-modal-title" id="holdings-modal-title">Holdings</div>
        <div class="holdings-modal-sub" id="holdings-modal-sub"></div>
      </div>
      <button onclick="closeHoldings()" style="background:none;border:none;color:#64748b;font-size:20px;cursor:pointer;padding:4px">✕</button>
    </div>
    <div class="holdings-table-wrap">
      <table class="holdings-table">
        <thead>
          <tr>
            <th>#</th>
            <th>Holding</th>
            <th style="text-align:right">Weight</th>
          </tr>
        </thead>
        <tbody id="holdings-tbody"></tbody>
      </table>
    </div>
  </div>
</div>

<!-- Ticker hover tooltip -->
<div id="ticker-tip" class="ticker-tip">
  <div class="ticker-tip-header">
    <div>
      <div class="ticker-tip-sym" id="tt-sym"></div>
      <div class="ticker-tip-name" id="tt-name"></div>
      <div class="ticker-tip-sector" id="tt-sector"></div>
    </div>
    <div>
      <div class="ticker-tip-price" id="tt-price"></div>
      <div class="ticker-tip-chg" id="tt-chg"></div>
    </div>
  </div>
  <div class="ticker-tip-spark"><canvas id="tt-spark" height="40"></canvas></div>
  <div class="ticker-tip-grid" id="tt-grid"></div>
  <div id="tt-rec"></div>
</div>

<!-- Main app (hidden until unlocked) -->
<div id="app" style="display:none">

<div class="header">
  <h1>Portfolio Analysis Dashboard</h1>
  <p>Generated {{ date }} &nbsp;·&nbsp; {{ n_stocks }} positions &nbsp;·&nbsp;
     Data via Yahoo Finance &nbsp;·&nbsp; Not financial advice</p>
  <div style="display:inline-flex;align-items:center;gap:8px;margin-top:12px;
              background:#0f2942;border:1px solid #3b82f6;border-radius:8px;
              padding:8px 14px;font-size:13px;color:#93c5fd;">
    <span style="font-size:16px">📈</span>
    <span><strong>Equities only</strong> — all recommendations are for stock purchases (long positions). No options, futures contracts, or derivatives.</span>
  </div>

  {% if total_cost %}
  <div class="perf-banner">
    <div class="perf-item">
      <div class="perf-label">Current Value</div>
      <div class="perf-value" style="color:#f1f5f9">${{ "{:,.0f}".format(total_value) }}</div>
      <div class="perf-sub">{{ n_stocks }} positions</div>
    </div>
    <div class="perf-item">
      <div class="perf-label">Total Return</div>
      <div class="perf-value {{ 'positive' if total_gain >= 0 else 'negative' }}">
        {{ "+" if total_dollar_gain >= 0 else "" }}${{ "{:,.0f}".format(total_dollar_gain|abs) }}
      </div>
      <div class="perf-sub">
        <span class="{{ 'positive' if total_gain >= 0 else 'negative' }}" style="font-weight:700">
          {{ "+" if total_gain >= 0 else "" }}{{ "{:.2f}".format(total_gain) }}%
        </span>
        &nbsp;unrealised gain/loss
      </div>
    </div>
    <div class="perf-item">
      <div class="perf-label">Total Cost Basis</div>
      <div class="perf-value" style="color:#94a3b8;font-size:24px">${{ "{:,.0f}".format(total_cost) }}</div>
      <div class="perf-sub">amount invested</div>
    </div>
    {% if portfolio_beta is not none %}
    <div class="perf-item">
      <div class="perf-label">Portfolio Beta</div>
      <div class="perf-value" style="color:{{ beta_color }};font-size:24px">
        {{ "{:.2f}".format(portfolio_beta) }}
      </div>
      <div class="perf-sub">
        {% if portfolio_beta < 1.0 %}less volatile than market
        {% elif portfolio_beta < 1.3 %}moves ~{{ "{:.1f}".format(portfolio_beta) }}x market
        {% else %}{{ "{:.1f}".format(portfolio_beta) }}x market volatility{% endif %}
      </div>
    </div>
    {% endif %}

    {# ── SPY Benchmark tile ── #}
    {% if spy_benchmark and total_gain is not none %}
    {% set spy1y = spy_benchmark.get("one_yr") %}
    {% set spy_ytd = spy_benchmark.get("ytd") %}
    {% set beating = spy1y is not none and total_gain > spy1y %}
    <div class="perf-item" style="border-left:2px solid #334155;padding-left:20px">
      <div class="perf-label">vs. S&amp;P 500 (SPY)</div>
      <div class="perf-value" style="font-size:16px;margin-top:8px;display:flex;flex-direction:column;gap:4px">
        {% if spy_ytd is not none %}
        <div style="display:flex;justify-content:space-between;gap:20px">
          <span style="color:#64748b;font-size:12px;font-weight:400">YTD</span>
          <span class="{{ 'positive' if spy_ytd >= 0 else 'negative' }}" style="font-size:14px">
            {{ "{:+.1f}".format(spy_ytd) }}%
          </span>
        </div>
        {% endif %}
        {% if spy1y is not none %}
        <div style="display:flex;justify-content:space-between;gap:20px">
          <span style="color:#64748b;font-size:12px;font-weight:400">1Y</span>
          <span class="{{ 'positive' if spy1y >= 0 else 'negative' }}" style="font-size:14px">
            {{ "{:+.1f}".format(spy1y) }}%
          </span>
        </div>
        {% endif %}
        {% if spy_benchmark.get("three_yr_ann") is not none %}
        <div style="display:flex;justify-content:space-between;gap:20px">
          <span style="color:#64748b;font-size:12px;font-weight:400">3Y ann.</span>
          <span class="{{ 'positive' if spy_benchmark.three_yr_ann >= 0 else 'negative' }}" style="font-size:14px">
            {{ "{:+.1f}".format(spy_benchmark.three_yr_ann) }}%
          </span>
        </div>
        {% endif %}
      </div>
      <div class="perf-sub" style="margin-top:8px">
        {% if spy1y is not none %}
          {% set diff = total_gain - spy1y %}
          <span style="font-weight:700;color:{{ '#34d399' if beating else '#f87171' }}">
            {{ '▲ Beating' if beating else '▼ Lagging' }} by {{ "{:.1f}".format(diff|abs) }}%
          </span>
          <span style="color:#475569"> vs 1Y SPY</span>
        {% endif %}
      </div>
    </div>
    {% endif %}

    {# ── Earnings Calendar tile ── #}
    <div class="perf-item" style="border-left:2px solid #334155;padding-left:20px;min-width:175px">
      <div class="perf-label">Upcoming Earnings</div>
      <div id="earnings-tile" class="earn-list">
        <div style="font-size:12px;color:#475569;margin-top:6px">Loading...</div>
      </div>
    </div>

    {# ── Price Alerts tile ── #}
    <div class="perf-item" style="border-left:2px solid #334155;padding-left:20px;min-width:210px">
      <div class="perf-label">Price Alerts</div>
      <div id="alerts-tile" class="alert-list"></div>
      <div class="alert-inputs">
        <input id="al-sym" class="alert-in alert-in-sym" placeholder="AAPL" maxlength="6"
               oninput="this.value=this.value.toUpperCase()">
        <input id="al-px"  class="alert-in alert-in-px"  placeholder="Price" type="number" step="0.01" min="0">
        <button class="alert-add-btn" onclick="alertAdd()">+ Alert</button>
      </div>
    </div>

    {# ── Live Refresh tile ── #}
    <div class="perf-item" style="border-left:2px solid #334155;padding-left:20px;min-width:140px">
      <div class="perf-label">Live Prices</div>
      <div id="refresh-status" style="font-size:12px;color:#94a3b8;margin-top:6px">Auto-refreshing...</div>
      <button id="refresh-btn" class="alert-add-btn" style="margin-top:8px;width:100%" onclick="refreshPrices(true)">&#8635; Refresh Now</button>
    </div>

  </div>
  {% endif %}
</div>

<!-- ── Market Overview Strip ── -->
{% if market_overview.strip %}
<div class="market-strip">
  {% for item in market_overview.strip %}
  <div class="strip-item">
    <span class="strip-label">{{ item.label }}</span>
    <span class="strip-price">{{ item.price }}</span>
    <span class="strip-chg {{ 'strip-up' if item.direction == 'up' else 'strip-down' }}">
      {{ '+' if item.direction == 'up' else '' }}{{ item.change_pct }}%
    </span>
  </div>
  {% endfor %}
</div>
{% endif %}

<!-- Tab navigation -->
<div class="tab-nav">
  <button class="tab-btn active" data-tab="tab-dashboard"
          onclick="switchTab('tab-dashboard')">📊 Dashboard</button>
  <button class="tab-btn" data-tab="tab-sim"
          onclick="switchTab('tab-sim');renderSim()">🧪 Simulation</button>
  <button class="tab-btn" data-tab="tab-news"
          onclick="switchTab('tab-news')">📰 News</button>
</div>

<div id="tab-dashboard" class="tab-pane active">

<!-- ── Sector Heatmap ── -->
{% if market_overview.sectors %}
<div class="section" style="padding-bottom:16px">
  <div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:14px">
    <h2 style="margin:0">🗂 Sector Performance</h2>
    <span style="font-size:12px;color:#475569">Today vs. prior close &nbsp;·&nbsp; SPDR ETFs</span>
  </div>
  <div class="heatmap-grid">
  {% for s in market_overview.sectors %}
    <div class="heatmap-tile" style="background:{{ s.bg }}"
         title="Click to see all {{ s.gics }} stocks in S&amp;P 500"
        >
      <span class="heatmap-icon">{{ s.icon }}</span>
      <span class="heatmap-sector">{{ s.label }}</span>
      <span class="heatmap-etf">{{ s.etf }}</span>
      <span class="heatmap-pct {{ 'strip-up' if s.change_pct >= 0 else 'strip-down' }}">
        {{ '+' if s.change_pct >= 0 else '' }}{{ s.change_pct }}%
      </span>
    </div>
  {% endfor %}
  </div>
</div>
{% endif %}

<!-- Summary stats -->
<div class="stats-grid">
  <div class="stat-card">
    <div class="label">Total Market Value</div>
    <div class="value">${{ "{:,.0f}".format(total_value) }}</div>
    <div class="sub">across {{ n_stocks }} positions</div>
  </div>
  <div class="stat-card">
    <div class="label">Total Cost Basis</div>
    <div class="value">
      {% if total_cost %}${{ "{:,.0f}".format(total_cost) }}{% else %}N/A{% endif %}
    </div>
    <div class="sub">&nbsp;</div>
  </div>
  <div class="stat-card">
    <div class="label">Total Gain / Loss</div>
    {% if total_gain is not none %}
    <div class="value {{ 'positive' if total_gain >= 0 else 'negative' }}">
      {{ "+" if total_gain >= 0 else "" }}{{ "{:.1f}".format(total_gain) }}%
    </div>
    {% else %}
    <div class="value">N/A</div>
    {% endif %}
    <div class="sub">unrealised</div>
  </div>
  <div class="stat-card">
    <div class="label">Sectors Covered</div>
    <div class="value">{{ n_sectors }}</div>
    <div class="sub">of 11 S&amp;P 500 sectors</div>
  </div>
</div>

<!-- Holdings table -->
<div class="section">
  <h2>Current Holdings</h2>
  <table>
    <thead>
      <tr>
        <th>Symbol</th><th>Name</th><th>Sector / Category</th><th>Shares</th>
        <th>Price</th><th>Market Value</th><th>Gain/Loss</th>
        <th>P/E · Exp Ratio</th><th>Fwd P/E · Yield</th>
        <th>Div Yield · 3Y Ret</th><th>Upside · YTD</th><th>Rating · AUM</th>
        <th>Ideal Sell</th><th></th>
      </tr>
    </thead>
    <tbody>
      {% for r in holdings %}
      {% set is_etf = r.get("is_etf") %}
      <tr data-holding-sym="{{ r.Symbol }}" data-holding-shares="{{ r.Shares }}" data-holding-cost="{{ r['Cost Basis'] if r['Cost Basis'] is not none else 0 }}">
        <td>
          <strong data-sym="{{ r.Symbol }}" style="color:{{ '#38bdf8' if is_etf else '#60a5fa' }}">{{ r.Symbol }}</strong>
          {% if is_etf %}<span class="etf-badge etf-badge-etf" style="margin-left:5px">ETF</span>{% endif %}
        </td>
        <td>{{ r.Name }}</td>
        <td>{{ r["ETF Category"] if is_etf else r.Sector }}</td>
        <td>{{ "{:,.0f}".format(r.Shares) }}</td>
        <td data-price-cell="{{ r.Symbol }}">${{ "{:,.2f}".format(r.Price) if r.Price else "—" }}</td>
        <td data-mv-cell="{{ r.Symbol }}">${{ "{:,.0f}".format(r["Market Value"]) }}</td>
        <td data-gl-cell="{{ r.Symbol }}">
          {% if r["Gain/Loss %"] is not none %}
          <span class="{{ 'positive' if r['Gain/Loss %'] >= 0 else 'negative' }}">
            {{ "+" if r["Gain/Loss %"] >= 0 else "" }}{{ "{:.1f}".format(r["Gain/Loss %"]) }}%
          </span>
          {% else %}—{% endif %}
        </td>
        {# P/E (stocks) or Expense Ratio (ETFs) #}
        <td>
          {% if is_etf %}
            {{ "{:.2f}".format(r["Expense Ratio"]) ~ "%" if r["Expense Ratio"] is not none else "—" }}
          {% else %}
            {{ "{:.1f}".format(r["P/E"]) if r["P/E"] else "—" }}
          {% endif %}
        </td>
        {# Fwd P/E (stocks) or ETF Yield #}
        <td>
          {% if is_etf %}
            {{ "{:.2f}".format(r["ETF Yield"]) ~ "%" if r["ETF Yield"] else "—" }}
          {% else %}
            {{ "{:.1f}".format(r["Fwd P/E"]) if r["Fwd P/E"] else "—" }}
          {% endif %}
        </td>
        {# Div Yield (stocks) or 3Y Return (ETFs) #}
        <td>
          {% if is_etf %}
            {% if r["3Y Return"] is not none %}
            <span class="{{ 'positive' if r['3Y Return'] >= 0 else 'negative' }}">
              {{ "+" if r["3Y Return"] >= 0 else "" }}{{ "{:.1f}".format(r["3Y Return"]) }}%
            </span>
            {% else %}—{% endif %}
          {% else %}
            {{ "{:.2f}".format(r["Div Yield %"]) ~ "%" if r["Div Yield %"] else "—" }}
          {% endif %}
        </td>
        {# Analyst Upside (stocks) or YTD return (ETFs) #}
        <td>
          {% if is_etf %}
            {% if r["YTD Return"] is not none %}
            <span class="{{ 'positive' if r['YTD Return'] >= 0 else 'negative' }}">
              {{ "+" if r["YTD Return"] >= 0 else "" }}{{ "{:.1f}".format(r["YTD Return"]) }}% YTD
            </span>
            {% else %}—{% endif %}
          {% else %}
            {% if r["Analyst Upside %"] is not none %}
            <span class="{{ 'positive' if r['Analyst Upside %'] >= 0 else 'negative' }}">
              {{ "+" if r["Analyst Upside %"] >= 0 else "" }}{{ "{:.1f}".format(r["Analyst Upside %"]) }}%
            </span>
            {% else %}—{% endif %}
          {% endif %}
        </td>
        {# Analyst Rating (stocks) or AUM (ETFs) #}
        <td>
          {% if is_etf %}
            {% if r["AUM"] %}
              {% set aum_b = r["AUM"] / 1e9 %}
              <span style="font-size:12px;color:#38bdf8;font-weight:600">
                {{ "${:,.0f}B".format(aum_b) if aum_b >= 1 else "${:,.0f}M".format(r["AUM"]/1e6) }}
              </span>
            {% else %}—{% endif %}
          {% else %}
            {% set rec = r.Recommendation or "" %}
            {% if "buy" in rec.lower() or "strong" in rec.lower() %}
              <span class="badge badge-green">{{ rec }}</span>
            {% elif "hold" in rec.lower() %}
              <span class="badge badge-yellow">{{ rec }}</span>
            {% elif "sell" in rec.lower() or "underperform" in rec.lower() %}
              <span class="badge badge-red">{{ rec }}</span>
            {% else %}
              <span class="badge badge-blue">{{ rec or "—" }}</span>
            {% endif %}
          {% endif %}
        </td>
        {% set is_etf     = r.get("is_etf") %}
        {% set sell_price = r["Analyst Target"] %}
        {% set upside     = r["Analyst Upside %"] %}
        {% set h          = holding_map.get(r.Symbol) %}
        {% set needs_dd   = sell_price and not is_etf and (
              (upside is not none and upside < 5)
              or (h and h.verdict in ("WATCH","SELL"))
           ) %}
        <td style="min-width:110px">
          {% if is_etf and r.Price %}
            {# ETF: 8% trailing stop — standard passive holding guideline #}
            {% set stop = r.Price * 0.92 %}
            <div style="font-size:13px;font-weight:700;color:#38bdf8">${{ "{:,.2f}".format(stop) }}</div>
            <div style="font-size:10px;color:#475569;margin-top:1px">8% trailing stop</div>
          {% elif sell_price %}
            {% if upside is not none and upside < 0 %}
              {% set sell_color = "#f87171" %}
            {% elif upside is not none and upside < 5 %}
              {% set sell_color = "#fbbf24" %}
            {% else %}
              {% set sell_color = "#34d399" %}
            {% endif %}
            <div style="font-size:14px;font-weight:700;color:{{ sell_color }}">
              ${{ "{:,.2f}".format(sell_price) }}</div>
            {% if upside is not none %}
            <div style="font-size:11px;color:#64748b;margin-top:1px">
              {{ "{:+.1f}".format(upside) }}% to target</div>
            {% endif %}
            {% if needs_dd %}
            <details style="margin-top:5px">
              <summary style="font-size:11px;color:#60a5fa;cursor:pointer;
                              list-style:none;display:inline-flex;align-items:center;gap:3px">
                ▾ Why sell here
              </summary>
              <div style="font-size:12px;color:#94a3b8;margin-top:7px;line-height:1.5;
                          max-width:220px;padding:8px 10px;background:#0f172a;
                          border-radius:6px;border-left:2px solid {{ sell_color }}">
                {% if upside is not none and upside < 0 %}
                  Stock is <strong style="color:#f87171">trading above</strong>
                  its analyst consensus target — Wall Street sees no more upside.
                  Consider a limit-sell order at or below
                  <strong>${{ "{:,.2f}".format(sell_price) }}</strong>.
                {% elif upside is not none and upside < 5 %}
                  Within <strong style="color:#fbbf24">{{ "{:.1f}".format(upside) }}%</strong>
                  of analyst target. Consider taking partial profits —
                  lock in gains while leaving some exposure if the thesis continues.
                {% endif %}
                {% if h and h.reasons %}
                  <div style="margin-top:6px;padding-top:6px;border-top:1px solid #1e293b">
                    {% for reason in h.reasons[:2] %}
                    <div style="margin-bottom:4px">• {{ reason | safe }}</div>
                    {% endfor %}
                  </div>
                {% endif %}
              </div>
            </details>
            {% endif %}
          {% else %}
            <span style="color:#475569">—</span>
          {% endif %}
        </td>
        <td>
          <button class="chart-btn" onclick="openChart('{{ r.Symbol }}')">📈 Chart</button>
        </td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>

<!-- Hold or Sell Analysis -->
{% if holding_analysis %}
<div class="section">
  <h2>Hold or Sell? — Portfolio Position Review</h2>
  <p style="font-size:13px;color:#64748b;margin-bottom:20px">
    Each position is scored 0–100 on analyst sentiment, remaining upside, price momentum, and
    earnings growth. <strong style="color:#f87171">SELL</strong> =&nbsp;multiple red flags detected.
    <strong style="color:#fbbf24">WATCH</strong> =&nbsp;monitor closely.
    <strong style="color:#34d399">HOLD</strong> =&nbsp;thesis intact.
    Suggested replacements come from the same sector.
  </p>
  <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(460px,1fr));gap:16px">
  {% for h in holding_analysis %}
  <div class="holding-card" style="border-left:4px solid {{ h.verdict_color }}">

    <!-- Card header -->
    <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:8px">
      <div>
        <span style="font-size:20px;font-weight:800;color:#60a5fa">{{ h.symbol }}</span>
        <span style="font-size:14px;color:#64748b;margin-left:8px">{{ h.name }}</span>
        <div style="font-size:11px;color:#475569;margin-top:3px;text-transform:uppercase;
                    letter-spacing:.05em">{{ h.sector }}</div>
      </div>
      <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px">
        <div class="verdict-badge"
             style="background:{{ h.verdict_bg }};color:{{ h.verdict_color }};
                    border:1px solid {{ h.verdict_color }}44">
          {{ h.verdict }}
        </div>
        <div style="font-size:12px;color:#475569">Score: {{ h.score }}/100</div>
      </div>
    </div>

    <!-- Quick stats row -->
    <div style="display:flex;gap:20px;margin-top:12px;flex-wrap:wrap">
      {% if h.gain_pct is not none %}
      <div style="font-size:12px">
        <span style="color:#64748b">Portfolio gain: </span>
        <span class="{{ 'positive' if h.gain_pct >= 0 else 'negative' }}" style="font-weight:600">
          {{ "+" if h.gain_pct >= 0 else "" }}{{ "{:.1f}".format(h.gain_pct) }}%
        </span>
      </div>
      {% endif %}
      {% if h.upside is not none %}
      <div style="font-size:12px">
        <span style="color:#64748b">Remaining upside: </span>
        <span class="{{ 'positive' if h.upside >= 0 else 'negative' }}" style="font-weight:600">
          {{ "{:+.1f}".format(h.upside) }}%
        </span>
      </div>
      {% endif %}
      {% if h.mom_3m is not none %}
      <div style="font-size:12px">
        <span style="color:#64748b">3M return: </span>
        <span class="{{ 'positive' if h.mom_3m >= 0 else 'negative' }}" style="font-weight:600">
          {{ "{:+.1f}".format(h.mom_3m) }}%
        </span>
      </div>
      {% endif %}
    </div>

    <!-- Sell signals -->
    {% if h.reasons %}
    <div style="margin-top:12px">
      <div style="font-size:11px;font-weight:700;color:#f87171;text-transform:uppercase;
                  letter-spacing:.05em;margin-bottom:4px">⚠ Sell Signals</div>
      <ul class="signal-list bad">
        {% for r in h.reasons %}<li>{{ r | safe }}</li>{% endfor %}
      </ul>
    </div>
    {% endif %}

    <!-- Positives -->
    {% if h.positives %}
    <div style="margin-top:10px">
      <div style="font-size:11px;font-weight:700;color:#34d399;text-transform:uppercase;
                  letter-spacing:.05em;margin-bottom:4px">✓ Hold Case</div>
      <ul class="signal-list good">
        {% for p in h.positives %}<li>{{ p | safe }}</li>{% endfor %}
      </ul>
    </div>
    {% endif %}

    <!-- Replacement suggestions -->
    {% if h.replacements %}
    <div style="margin-top:14px;padding-top:14px;border-top:1px solid #334155">
      <div style="font-size:11px;font-weight:700;color:#a78bfa;text-transform:uppercase;
                  letter-spacing:.05em;margin-bottom:8px">
        🔄 Consider Replacing With (same sector — {{ h.sector }})
      </div>
      <div style="display:flex;flex-direction:column;gap:8px">
      {% for rep in h.replacements %}
      <div class="replace-card">
        <div>
          <span style="font-size:16px;font-weight:700;color:#a78bfa">{{ rep.symbol }}</span>
          <span style="font-size:12px;color:#64748b;margin-left:8px">{{ rep.name }}</span>
          <div style="margin-top:3px;display:flex;gap:8px;align-items:center">
            <span style="font-size:11px;font-weight:600;
                         color:{{ risk_tier_color.get(rep.risk_tier,'#94a3b8') }}">
              {{ risk_tier_label.get(rep.risk_tier, rep.risk_tier) }}
            </span>
            {% if rep.recommendation %}
            <span style="font-size:11px;color:#64748b">· {{ rep.recommendation }}</span>
            {% endif %}
          </div>
        </div>
        <div style="text-align:right;flex-shrink:0">
          <div style="font-size:14px;font-weight:700;color:#f1f5f9">
            ${{ "{:,.2f}".format(rep.price) }}</div>
          <div style="font-size:12px;color:#34d399;font-weight:600">
            +{{ "{:.1f}".format(rep.upside) }}% upside</div>
        </div>
      </div>
      {% endfor %}
      </div>
    </div>
    {% endif %}

  </div>
  {% endfor %}
  </div>
</div>
{% endif %}

<!-- Sector allocation -->
<div class="section">
  <h2>Sector Allocation vs S&P 500</h2>
  <table>
    <thead>
      <tr>
        <th>Sector</th><th>Market Value</th>
        <th>Portfolio %</th><th>S&P 500 %</th><th>Over/Under</th>
        <th>Visual</th>
      </tr>
    </thead>
    <tbody>
      {% for r in sectors %}
      <tr>
        <td>{{ r.Sector }}</td>
        <td>${{ "{:,.0f}".format(r["Market Value"]) if r["Market Value"] is not none else "—" }}</td>
        <td>{{ "{:.1f}".format(r["Portfolio Weight %"]) ~ "%" if r["Portfolio Weight %"] is not none else "—" }}</td>
        <td>{{ "{:.1f}".format(r["S&P 500 Weight %"]) ~ "%" if r["S&P 500 Weight %"] is not none else "—" }}</td>
        <td>
          {% set gap = r["Over/Under %"] %}
          {% if gap is not none %}
          <span class="{{ 'positive' if gap >= 0 else 'negative' }}">
            {{ "+" if gap >= 0 else "" }}{{ "{:.1f}".format(gap) }}%
          </span>
          {% else %}—{% endif %}
        </td>
        <td style="min-width:180px">
          <div style="font-size:10px;color:#64748b;margin-bottom:3px">
            <span style="color:#3b82f6">■</span> Portfolio &nbsp;
            <span style="color:#64748b">■</span> S&amp;P 500
          </div>
          <div class="bar-wrap" style="margin-bottom:3px">
            <div class="bar-fill-port"
                 style="width:{{ [r['Portfolio Weight %']/35*100,100]|min if r['Portfolio Weight %'] is not none else 0 }}%"></div>
          </div>
          <div class="bar-wrap">
            <div class="bar-fill-sp"
                 style="width:{{ [r['S&P 500 Weight %']/35*100,100]|min if r['S&P 500 Weight %'] is not none else 0 }}%"></div>
          </div>
        </td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>

<!-- Cash Allocation Plan -->
<div class="section">
  <h2>Cash Deployment Plan — ${{ "{:,.0f}".format(cash) }} Available</h2>
  <p style="font-size:13px;color:#64748b;margin-bottom:16px">
    Suggested split based on your aggressive risk profile (age {{ age }}).
    Higher allocation to high-risk/speculative plays given your long time horizon.
  </p>
  <div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:24px">
    {% for tier, info in cash_plan.items() %}
    <div style="background:#1e293b;border:1px solid #334155;border-radius:10px;
                padding:16px 20px;min-width:160px;border-top:3px solid {{ risk_tier_color[tier] }}">
      <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.05em">
        {{ risk_tier_label[tier] }}
      </div>
      <div style="font-size:26px;font-weight:700;color:{{ risk_tier_color[tier] }};margin-top:6px">
        ${{ "{:,.0f}".format(info.dollars) }}
      </div>
      <div style="font-size:13px;color:#64748b;margin-top:4px">{{ "{:.0f}".format(info.pct) }}% of cash</div>
    </div>
    {% endfor %}
  </div>
</div>

<!-- ── Reusable rec-card macro ── -->
{% macro rec_card(rec, accent_color) %}
<div class="rec-card" style="border-left:3px solid {{ accent_color }}">
  <div class="rec-header">
    <div>
      <div class="rec-symbol" data-sym="{{ rec.symbol }}">{{ rec.symbol }}</div>
      <div class="rec-name">{{ rec.name }}</div>
      <div style="margin-top:4px;display:flex;gap:6px;align-items:center;flex-wrap:wrap">
        <span class="badge" style="background:{{ risk_tier_color.get(rec.risk_tier,'#334155') }}22;
              color:{{ risk_tier_color.get(rec.risk_tier,'#94a3b8') }};
              border:1px solid {{ risk_tier_color.get(rec.risk_tier,'#334155') }}">
          {{ risk_tier_label.get(rec.risk_tier, rec.risk_tier) }}
        </span>
        {% if rec.in_sp500 %}<span class="badge badge-blue">S&amp;P 500</span>
        {% else %}<span class="badge badge-yellow">Non-S&amp;P</span>{% endif %}
        {% if rec.days_to_earnings is not none and rec.days_to_earnings >= 0 and rec.days_to_earnings <= 14 %}
        <span class="badge earnings-soon"
              style="background:#431407;color:#fb923c;border:1px solid #fb923c55">
          ⚠️ Earnings in {{ rec.days_to_earnings }}d
        </span>
        {% endif %}
      </div>
    </div>
    <div style="text-align:right">
      {% if rec.recommendation %}{% set r = rec.recommendation %}
      {% if "buy" in r.lower() or "strong" in r.lower() %}<span class="badge badge-green">{{ r }}</span>
      {% elif "hold" in r.lower() %}<span class="badge badge-yellow">{{ r }}</span>
      {% else %}<span class="badge badge-blue">{{ r }}</span>{% endif %}{% endif %}
      <div style="font-size:20px;font-weight:700;color:#f1f5f9;margin-top:4px">
        ${{ "{:,.2f}".format(rec.price) if rec.price else "—" }}
      </div>
    </div>
  </div>

  <!-- Key metrics row -->
  <div class="rec-metrics">
    <div class="rec-metric"><div class="k">Beta</div>
      <div class="v">{{ "{:.2f}".format(rec.beta) if rec.beta else "—" }}</div></div>
    <div class="rec-metric"><div class="k">Fwd P/E</div>
      <div class="v">{{ "{:.1f}".format(rec.forward_pe) if rec.forward_pe else "—" }}</div></div>
    <div class="rec-metric"><div class="k">EPS Growth</div>
      <div class="v">{{ "{:+.1f}".format(rec.eps_growth) ~ "%" if rec.eps_growth else "—" }}</div></div>
    <div class="rec-metric"><div class="k">Rev Growth</div>
      <div class="v">{{ "{:+.1f}".format(rec.rev_growth) ~ "%" if rec.rev_growth else "—" }}</div></div>
    <div class="rec-metric"><div class="k">Growth Score</div>
      <div class="v" style="color:{{ accent_color }}">
        {{ "{:.0f}".format(rec.growth_score) if rec.growth_score else "—" }}</div></div>
    <div class="rec-metric"><div class="k">Analyst Target</div>
      <div class="v">${{ "{:,.2f}".format(rec.analyst_target) if rec.analyst_target else "—" }}</div></div>
    <div class="rec-metric"><div class="k">Upside</div>
      <div class="v {{ 'positive' if (rec.upside or 0) >= 0 else 'negative' }}">
        {{ "{:+.1f}".format(rec.upside) ~ "%" if rec.upside is not none else "—" }}</div></div>
    <div class="rec-metric"><div class="k">3M Return</div>
      <div class="v {{ 'positive' if (rec.mom_3m or 0) >= 0 else 'negative' }}">
        {{ "{:+.1f}".format(rec.mom_3m) ~ "%" if rec.mom_3m is not none else "—" }}</div></div>
    <div class="rec-metric"><div class="k">6M Return</div>
      <div class="v {{ 'positive' if (rec.mom_6m or 0) >= 0 else 'negative' }}">
        {{ "{:+.1f}".format(rec.mom_6m) ~ "%" if rec.mom_6m is not none else "—" }}</div></div>
  </div>

  <!-- Action buttons row -->
  <div style="display:flex;gap:8px;flex-wrap:wrap">
    <button class="dropdown-btn" onclick="toggleDetail(this)">
      <span>Details — Why, What to Buy/Sell At, Metric Guide</span>
      <span class="arrow">▾</span>
    </button>
    <button class="chart-btn" onclick="openChart('{{ rec.symbol }}')">📈 Chart</button>
    <button class="quick-add-btn" data-sym="{{ rec.symbol }}"
            onclick="simAdd('{{ rec.symbol }}', {{ rec.price or 0 }}, {{ rec.suggested_shares or 1 }}, new Date().toISOString().slice(0,10))">+ Sim</button>
  </div>

  <div class="detail-panel">

    <!-- Why this stock -->
    <div class="detail-section">
      <h4>Why this stock is recommended</h4>
      <div class="why-box">{{ rec.why_html | safe }}</div>
    </div>

    <!-- Buy / Sell prices -->
    <div class="detail-section">
      <h4>Entry &amp; Exit Levels</h4>
      <div class="buysell-grid">
        <div class="bs-card buy">
          <div class="bs-label">Buy At / Entry</div>
          <div class="bs-price">${{ "{:,.2f}".format(rec.buy_at) if rec.buy_at else "—" }}</div>
          {% if rec.suggested_shares and rec.suggested_shares > 0 %}
          <div style="margin-top:8px;padding-top:8px;border-top:1px solid #34d39922">
            <div style="font-size:22px;font-weight:700;color:#34d399">
              {{ rec.suggested_shares }} share{{ "s" if rec.suggested_shares != 1 else "" }}
            </div>
            <div style="font-size:11px;color:#6ee7b7;margin-top:2px">
              ≈ ${{ "{:,.2f}".format(rec.suggested_spend) }} of your
              ${{ "{:,.0f}".format(rec.tier_budget) }} {{ rec.risk_tier }} budget
            </div>
            {% if rec.leftover and rec.leftover > 0 %}
            <div style="font-size:11px;color:#64748b;margin-top:2px">
              ${{ "{:,.2f}".format(rec.leftover) }} left over
            </div>
            {% endif %}
          </div>
          {% else %}
          <div class="bs-note" style="margin-top:6px">
            {% if rec.buy_at and rec.tier_budget and rec.buy_at > rec.tier_budget %}
            Price (${{ "{:,.0f}".format(rec.buy_at) }}) exceeds per-stock budget
            (${{ "{:,.0f}".format(rec.tier_budget) }}) — consider 1 fractional share
            {% else %}
            ~1% limit order buffer below current price
            {% endif %}
          </div>
          {% endif %}
        </div>
        <div class="bs-card target">
          <div class="bs-label">Price Target / Sell</div>
          <div class="bs-price">${{ "{:,.2f}".format(rec.target) if rec.target else "—" }}</div>
          <div class="bs-note">
            {% if rec.analyst_target %}Analyst mean 12-month target
            {% else %}Estimated from growth score{% endif %}
          </div>
        </div>
        <div class="bs-card stop">
          <div class="bs-label">Stop Loss</div>
          <div class="bs-price">${{ "{:,.2f}".format(rec.stop_loss) if rec.stop_loss else "—" }}</div>
          <div class="bs-note">{{ rec.stop_pct }}% below entry — based on {{ rec.risk_tier }} risk tier</div>
        </div>
        <div class="bs-card timeline">
          <div class="bs-label">⏱ Timeline to Target</div>
          <div class="bs-price" style="color:#a78bfa;font-size:18px">{{ rec.timeline_label }}</div>
          <div class="bs-note">{{ rec.timeline_context }}</div>
          {% if rec.timeline_months and rec.price and rec.target %}
          <div style="margin-top:10px">
            <div style="display:flex;justify-content:space-between;font-size:10px;color:#64748b;margin-bottom:3px">
              <span>Now ${{ "{:,.0f}".format(rec.price) }}</span>
              <span>Target ${{ "{:,.0f}".format(rec.target) }}</span>
            </div>
            <div style="background:#0f172a;border-radius:4px;height:6px;overflow:hidden">
              {% set bar_pct = [(rec.upside / 60 * 100)|int, 100]|min %}
              <div class="timeline-bar"
                   data-pct="{{ bar_pct }}"
                   style="width:0%;background:linear-gradient(90deg,#a78bfa,#60a5fa);height:100%;border-radius:4px"></div>
            </div>
            <div style="font-size:10px;color:#64748b;margin-top:3px;text-align:center">
              {{ "{:+.1f}".format(rec.upside) }}% to go over ~{{ rec.timeline_months }} months
            </div>
          </div>
          {% endif %}
        </div>
      </div>
      <p style="font-size:11px;color:#475569;margin-top:8px">
        ⚠ These are suggested stock (equity) entry and exit levels — no options or derivatives. Adjust based on your own conviction and position size.
      </p>
    </div>

    <!-- Earnings -->
    <div class="detail-section">
      <h4>📅 Earnings</h4>
      <div class="earnings-grid">

        <!-- Next earnings -->
        <div style="background:#0f172a;border-radius:8px;padding:14px">
          <div style="font-size:11px;color:#64748b;text-transform:uppercase;
                      letter-spacing:.05em;margin-bottom:8px">Next Earnings Date</div>
          {% if rec.earnings_date %}
          <div style="font-size:18px;font-weight:700;color:#f1f5f9">{{ rec.earnings_date }}</div>
          {% if rec.days_to_earnings is not none %}
            {% if rec.days_to_earnings < 0 %}
            <div style="font-size:12px;color:#64748b;margin-top:4px">
              {{ (rec.days_to_earnings * -1)|int }} days ago</div>
            {% elif rec.days_to_earnings <= 14 %}
            <div style="font-size:12px;color:#fb923c;font-weight:600;margin-top:4px">
              ⚠️ In {{ rec.days_to_earnings }} days — consider waiting until after earnings
              for reduced binary risk.
            </div>
            {% else %}
            <div style="font-size:12px;color:#64748b;margin-top:4px">
              In {{ rec.days_to_earnings }} days</div>
            {% endif %}
          {% endif %}
          {% if rec.eps_est_next is not none %}
          <div style="margin-top:10px;padding-top:10px;border-top:1px solid #1e293b">
            <div style="font-size:11px;color:#64748b">Next Quarter EPS Estimate</div>
            <div style="font-size:22px;font-weight:700;color:#60a5fa;margin-top:2px">
              ${{ "{:.2f}".format(rec.eps_est_next) }}</div>
            <div style="font-size:11px;color:#475569;margin-top:2px">
              Wall Street consensus estimate</div>
          </div>
          {% endif %}
          {% else %}
          <div style="color:#475569;font-size:13px">Date not available</div>
          {% endif %}
        </div>

        <!-- Last quarter actual vs estimate -->
        <div style="background:#0f172a;border-radius:8px;padding:14px">
          <div style="font-size:11px;color:#64748b;text-transform:uppercase;
                      letter-spacing:.05em;margin-bottom:8px">
            Last Quarter{% if rec.last_q_period %} · {{ rec.last_q_period }}{% endif %}
          </div>
          {% if rec.last_eps_actual is not none %}
          <div style="display:flex;gap:20px;align-items:flex-end;flex-wrap:wrap">
            <div>
              <div style="font-size:11px;color:#64748b">Actual EPS</div>
              <div style="font-size:24px;font-weight:700;color:#f1f5f9">
                ${{ "{:.2f}".format(rec.last_eps_actual) }}</div>
            </div>
            <div>
              <div style="font-size:11px;color:#64748b">Estimated</div>
              <div style="font-size:24px;font-weight:700;color:#64748b">
                {% if rec.last_eps_estimate is not none %}
                ${{ "{:.2f}".format(rec.last_eps_estimate) }}
                {% else %}—{% endif %}
              </div>
            </div>
          </div>
          {% if rec.last_eps_surprise_pct is not none %}
          <div style="margin-top:10px;display:inline-block;padding:6px 12px;border-radius:6px;
                      background:{{ '#052e16' if rec.last_eps_surprise_pct >= 0 else '#450a0a' }};
                      border:1px solid {{ '#34d39944' if rec.last_eps_surprise_pct >= 0 else '#f8717144' }}">
            <span style="font-weight:700;font-size:14px;
                         color:{{ '#34d399' if rec.last_eps_surprise_pct >= 0 else '#f87171' }}">
              {{ "+" if rec.last_eps_surprise_pct >= 0 else "" }}{{ "{:.1f}".format(rec.last_eps_surprise_pct) }}%
              {{ "Beat ✅" if rec.last_eps_surprise_pct >= 0 else "Miss ❌" }}
            </span>
          </div>
          <div style="font-size:11px;color:#475569;margin-top:6px">vs. analyst consensus</div>
          {% endif %}
          {% else %}
          <div style="color:#475569;font-size:13px">Earnings history not available</div>
          {% endif %}
        </div>

      </div>
    </div>

    <!-- Metric explanations -->
    <div class="detail-section">
      <h4>What each metric means</h4>
      <table class="metric-table">
        {% for key, (short, long) in metric_explanations.items() %}
        <tr>
          <td>{{ key }}</td>
          <td>{{ short }}</td>
          <td>{{ long }}</td>
        </tr>
        {% endfor %}
      </table>
    </div>

  </div>
</div>
{% endmacro %}

<!-- Sector Recommendations -->
{% if sector_recs %}
<div class="section">
  <h2>Sector Picks — Fill Your Underweight Gaps</h2>
  <p style="font-size:13px;color:#64748b;margin-bottom:20px">
    Stocks in sectors where your portfolio lags the S&P 500.
    Click <strong>Details</strong> on any card to see why it was picked, what to buy/sell at, and what each metric means.
  </p>
  {% set ns = namespace(last_cat="") %}
  {% for rec in sector_recs %}
    {% if rec.category != ns.last_cat %}
    <p style="font-size:14px;font-weight:600;color:#fbbf24;margin:20px 0 10px">
      {{ rec.category }}
      <span style="font-size:12px;color:#f87171;font-weight:400">
        — underweight by {{ rec.gap_pct }}% vs S&P 500
      </span>
    </p>
    {% set ns.last_cat = rec.category %}
    {% endif %}
    {{ rec_card(rec, risk_tier_color.get(rec.risk_tier,'#64748b')) }}
  {% endfor %}
</div>
{% endif %}

<!-- Thematic Picks -->
{% if thematic_recs %}
<div class="section">
  <h2>High-Growth Theme Picks — Beyond the S&amp;P 500</h2>
  <p style="font-size:13px;color:#64748b;margin-bottom:20px">
    Emerging high-conviction growth themes for a long time horizon.
    Click <strong>Details</strong> on any card for the full breakdown.
  </p>
  {% set ns2 = namespace(last_theme="") %}
  {% for rec in thematic_recs %}
    {% if rec.category != ns2.last_theme %}
    <p style="font-size:14px;font-weight:600;color:#a78bfa;margin:20px 0 10px">
      {{ rec.category }}
    </p>
    {% set ns2.last_theme = rec.category %}
    {% endif %}
    {{ rec_card(rec, risk_tier_color.get(rec.risk_tier,'#64748b')) }}
  {% endfor %}
</div>
{% endif %}

<!-- Ranked Buy List -->
{% if ranked_recs %}
<div class="section">
  <h2>All Picks Ranked — Best Buy to Worst</h2>
  <p style="font-size:13px;color:#64748b;margin-bottom:16px">
    Every recommended stock scored and ranked by a composite formula:
    <strong style="color:#94a3b8">35% analyst upside &nbsp;·&nbsp;
    25% growth score &nbsp;·&nbsp; 20% analyst rating &nbsp;·&nbsp;
    10% revenue growth &nbsp;·&nbsp; 10% EPS growth</strong>
  </p>
  <table>
    <thead>
      <tr>
        <th style="width:40px">Rank</th>
        <th>Symbol</th>
        <th>Name</th>
        <th>Category</th>
        <th>Risk</th>
        <th>Score</th>
        <th>Upside</th>
        <th>Growth Score</th>
        <th>EPS Growth</th>
        <th>Rev Growth</th>
        <th>Rating</th>
        <th>Buy At</th>
        <th>Shares</th>
        <th>Target</th>
        <th>Timeline</th>
        <th>3M Return</th>
        <th>Earnings</th>
        <th>Sim</th>
      </tr>
    </thead>
    <tbody>
      {% for rec in ranked_recs %}
      {% set score = rec.buy_score %}
      {% set bar_w = [score, 100] | min %}
      {% set score_color = "#34d399" if score >= 60 else ("#fbbf24" if score >= 35 else "#f87171") %}
      <tr>
        <td style="text-align:center">
          {% if rec.rank == 1 %}
            <span style="font-size:18px">🥇</span>
          {% elif rec.rank == 2 %}
            <span style="font-size:18px">🥈</span>
          {% elif rec.rank == 3 %}
            <span style="font-size:18px">🥉</span>
          {% else %}
            <span style="color:#475569;font-weight:600">#{{ rec.rank }}</span>
          {% endif %}
        </td>
        <td><strong data-sym="{{ rec.symbol }}" style="color:#60a5fa">{{ rec.symbol }}</strong></td>
        <td style="color:#94a3b8">{{ rec.name }}</td>
        <td style="font-size:12px;color:#64748b">{{ rec.category }}</td>
        <td>
          <span style="font-size:11px;font-weight:600;
                color:{{ risk_tier_color.get(rec.risk_tier,'#94a3b8') }}">
            {{ risk_tier_label.get(rec.risk_tier, rec.risk_tier) }}
          </span>
        </td>
        <td>
          <div style="display:flex;align-items:center;gap:8px;min-width:100px">
            <div style="flex:1;background:#0f172a;border-radius:4px;height:6px;overflow:hidden">
              <div style="width:{{ bar_w }}%;background:{{ score_color }};height:100%;border-radius:4px"></div>
            </div>
            <span style="font-size:13px;font-weight:700;color:{{ score_color }};min-width:32px">
              {{ "{:.0f}".format(score) }}
            </span>
          </div>
        </td>
        <td>
          {% if rec.upside is not none %}
          <span class="{{ 'positive' if rec.upside >= 0 else 'negative' }}">
            {{ "{:+.1f}".format(rec.upside) }}%
          </span>
          {% else %}—{% endif %}
        </td>
        <td>{{ "{:.0f}".format(rec.growth_score) if rec.growth_score else "—" }}</td>
        <td>
          {% if rec.eps_growth %}
          <span class="{{ 'positive' if rec.eps_growth >= 0 else 'negative' }}">
            {{ "{:+.1f}".format(rec.eps_growth) }}%
          </span>
          {% else %}—{% endif %}
        </td>
        <td>
          {% if rec.rev_growth %}
          <span class="{{ 'positive' if rec.rev_growth >= 0 else 'negative' }}">
            {{ "{:+.1f}".format(rec.rev_growth) }}%
          </span>
          {% else %}—{% endif %}
        </td>
        <td>
          {% set r = rec.recommendation or "" %}
          {% if "buy" in r.lower() or "strong" in r.lower() %}
            <span class="badge badge-green">{{ r }}</span>
          {% elif "hold" in r.lower() %}
            <span class="badge badge-yellow">{{ r }}</span>
          {% elif r %}
            <span class="badge badge-blue">{{ r }}</span>
          {% else %}—{% endif %}
        </td>
        <td style="color:#34d399;font-weight:600">
          ${{ "{:,.2f}".format(rec.buy_at) if rec.buy_at else "—" }}
        </td>
        <td style="color:#34d399;font-weight:700">
          {% if rec.suggested_shares and rec.suggested_shares > 0 %}
            {{ rec.suggested_shares }}
          {% else %}—{% endif %}
        </td>
        <td style="color:#60a5fa">
          ${{ "{:,.2f}".format(rec.target) if rec.target else "—" }}
        </td>
        <td style="color:#a78bfa;font-weight:600;white-space:nowrap">
          {% if rec.timeline_label and rec.timeline_label != "—" %}
            {{ rec.timeline_label }}
            <div style="font-size:10px;color:#64748b;font-weight:400">{{ rec.timeline_context }}</div>
          {% else %}—{% endif %}
        </td>
        <td>
          {% if rec.mom_3m is not none %}
          <span class="{{ 'positive' if rec.mom_3m >= 0 else 'negative' }}" style="font-weight:600">
            {{ "{:+.1f}".format(rec.mom_3m) }}%
          </span>
          {% else %}—{% endif %}
        </td>
        <td style="white-space:nowrap;font-size:12px">
          {% if rec.earnings_date %}
            {% if rec.days_to_earnings is not none and rec.days_to_earnings >= 0 and rec.days_to_earnings <= 14 %}
            <span class="earnings-soon" style="color:#fb923c;font-weight:600">
              ⚠️ {{ rec.earnings_date }}
            </span>
            {% else %}
            <span style="color:#64748b">{{ rec.earnings_date }}</span>
            {% endif %}
          {% else %}—{% endif %}
        </td>
        <td>
          <button class="quick-add-btn" data-sym="{{ rec.symbol }}"
                  onclick="simAdd('{{ rec.symbol }}', {{ rec.price or rec.buy_at or 0 }}, {{ rec.suggested_shares or 1 }}, new Date().toISOString().slice(0,10))">+ Sim</button>
        </td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>
{% endif %}

<!-- ══ ETF Picks ══ -->
{% if etf_recs %}
<div class="section">
  <h2>📦 ETF Picks — Diversify &amp; Reduce Costs</h2>
  <p style="font-size:13px;color:#64748b;margin-bottom:20px">
    Curated funds across six categories, scored on expense ratio (40%), 3-year return (35%),
    AUM/liquidity (15%), and yield (10%). ETFs you already hold are excluded.
  </p>

  {% set ns3 = namespace(last_cat="") %}
  {% for etf in etf_recs %}
    {% if etf.category != ns3.last_cat %}
    {% set ns3.last_cat = etf.category %}
    <p style="font-size:14px;font-weight:600;color:#38bdf8;margin:20px 0 10px">
      {{ etf.category }}
    </p>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:14px;margin-bottom:6px">
    {% endif %}

    <div class="etf-card" style="border-left:3px solid #38bdf8">
      <div class="etf-header">
        <div>
          <div class="etf-symbol" data-sym="{{ etf.symbol }}">{{ etf.symbol }}</div>
          <div class="etf-name">{{ etf.name }}</div>
          {% if etf.etf_category %}
          <div style="font-size:10px;color:#475569;margin-top:3px">{{ etf.etf_category }}</div>
          {% endif %}
        </div>
        <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px">
          <div class="etf-score-ring" style="--pct:{{ etf.score }}%">
            <span class="etf-score-val">{{ etf.score|int }}</span>
          </div>
        </div>
      </div>

      <div class="etf-desc">{{ etf.description }}</div>

      <div class="etf-metrics">
        <div class="etf-metric">
          <div class="k">Exp. Ratio</div>
          <div class="v" style="color:#34d399">
            {{ ("{:.2f}".format(etf.expense_ratio) ~ "%") if etf.expense_ratio is not none else "—" }}
          </div>
        </div>
        <div class="etf-metric">
          <div class="k">3Y Ann. Return</div>
          <div class="v {{ 'positive' if (etf.three_yr or 0) >= 0 else 'negative' }}">
            {{ ("{:+.1f}".format(etf.three_yr) ~ "%") if etf.three_yr is not none else "—" }}
          </div>
        </div>
        <div class="etf-metric">
          <div class="k">5Y Ann. Return</div>
          <div class="v {{ 'positive' if (etf.five_yr or 0) >= 0 else 'negative' }}">
            {{ ("{:+.1f}".format(etf.five_yr) ~ "%") if etf.five_yr is not none else "—" }}
          </div>
        </div>
        <div class="etf-metric">
          <div class="k">Yield</div>
          <div class="v">
            {{ ("{:.2f}".format(etf.etf_yield) ~ "%") if etf.etf_yield is not none else "—" }}
          </div>
        </div>
        <div class="etf-metric">
          <div class="k">AUM</div>
          <div class="v" style="color:#38bdf8">{{ etf.aum_label or "—" }}</div>
        </div>
        <div class="etf-metric">
          <div class="k">YTD Return</div>
          <div class="v {{ 'positive' if (etf.ytd or 0) >= 0 else 'negative' }}">
            {{ ("{:+.1f}".format(etf.ytd) ~ "%") if etf.ytd is not none else "—" }}
          </div>
        </div>
      </div>

      <div style="display:flex;gap:8px;flex-wrap:wrap;margin-top:2px">
        <button class="chart-btn" onclick="openChart('{{ etf.symbol }}')">📈 Chart</button>
        <button class="chart-btn" onclick="openHoldings('{{ etf.symbol }}', '{{ etf.name | replace("'", "\\'") }}')"
                style="background:#0f4c81">🗂 Holdings</button>
        {% if etf.price %}
        <span style="font-size:12px;color:#475569;align-self:center">
          ${{ "{:,.2f}".format(etf.price) }}
        </span>
        {% endif %}
      </div>
    </div>

    {# Close the grid div after the last card in a category #}
    {% set ns_next = namespace(next_cat="") %}
    {% if not loop.last %}{% set ns_next.next_cat = etf_recs[loop.index].category %}{% endif %}
    {% if loop.last or ns_next.next_cat != etf.category %}
    </div>
    {% endif %}
  {% endfor %}
</div>
{% endif %}

<footer>
  Portfolio Analyzer &nbsp;·&nbsp; Data from Yahoo Finance &nbsp;·&nbsp;
  For informational purposes only. Not financial advice. Always do your own research.
</footer>
</div><!-- /tab-dashboard -->

<!-- ═══════════════════════ SIM TAB ═══════════════════════ -->
<div id="tab-sim" class="tab-pane">

  <div class="section">
    <h2>🧪 Trial Simulation</h2>
    <p style="font-size:13px;color:#64748b;margin-bottom:20px">
      Shadow-trade picks before committing real money. Add positions at your chosen entry price
      and track live P&amp;L against the S&amp;P 500 over the same period.
    </p>

    <!-- Summary banner -->
    <div class="sim-summary" id="sim-summary"></div>

    <!-- Add position form -->
    <div class="sim-add-form">
      <div>
        <label>Ticker</label>
        <input id="sim-sym-input" placeholder="e.g. NVDA" style="text-transform:uppercase">
      </div>
      <div>
        <label>Entry Price ($)</label>
        <input id="sim-price-input" type="number" step="0.01" placeholder="195.00">
      </div>
      <div>
        <label>Shares</label>
        <input id="sim-shares-input" type="number" step="1" placeholder="10">
      </div>
      <div>
        <label>Entry Date</label>
        <input id="sim-date-input" type="date">
      </div>
      <button class="sim-add-btn" onclick="simAddFromForm()">＋ Add Position</button>
    </div>

    <!-- Positions table -->
    <div id="sim-table-wrap"></div>

    <!-- Quick-add from recs hint -->
    <p style="font-size:12px;color:#475569;margin-top:16px">
      💡 You can also click <strong style="color:#93c5fd">+ Add to Sim</strong> on any recommendation card to add it at its current price.
    </p>
  </div>
</div><!-- /tab-sim -->

<!-- ═══════════════════════ NEWS TAB ═══════════════════════ -->
<div id="tab-news" class="tab-pane">

  <!-- ── News Sub-tab Nav ── -->
  <div class="news-subnav">
    <button class="news-subtab-btn active" data-subtab="subtab-economy"
            onclick="switchNewsTab('subtab-economy')">🌐 Economy</button>
    <button class="news-subtab-btn" data-subtab="subtab-equity"
            onclick="switchNewsTab('subtab-equity')">📈 Equity News</button>
  </div>

  <!-- ══ SUB-TAB: ECONOMY ══ -->
  <div id="subtab-economy" class="news-subtab-pane active">
    <div class="section">
      <h2>🌐 Economy — Global Market &amp; Macro News</h2>
      <p style="font-size:13px;color:#64748b;margin-bottom:24px">
        World events, central bank decisions, inflation, commodities, and currency moves that drive the broader market.
      </p>

      <div class="economy-layout">

        <!-- Left: news feed -->
        <div>
          {% if economy_news %}
          <div style="display:flex;flex-direction:column;gap:10px">
            {% for art in economy_news %}
            <div class="news-item" style="background:#1e293b;border-radius:10px;padding:14px 16px;display:flex;gap:12px;align-items:flex-start">
              {% if art.thumbnail %}
              <img class="news-thumb" src="{{ art.thumbnail }}" alt=""
                   onerror="this.style.display='none'">
              {% endif %}
              <div class="news-body">
                <div class="news-source-row">
                  <span class="news-publisher" style="color:#34d399">{{ art.publisher }}</span>
                  {% if art.age %}<span class="news-age">· {{ art.age }}</span>{% endif %}
                </div>
                <a class="news-title" href="{{ art.link }}" target="_blank" rel="noopener noreferrer">
                  {{ art.title }}
                </a>
                {% if art.summary %}
                <div class="news-summary">{{ art.summary }}</div>
                {% endif %}
              </div>
            </div>
            {% endfor %}
          </div>
          {% else %}
          <div class="news-section-wrap" style="padding:20px">
            <p style="color:#475569;font-size:13px">No economy news available at this time.</p>
          </div>
          {% endif %}
        </div>

        <!-- Right: Yield Curve + Spot Prices sidebar -->
        <div class="yield-sidebar">

          <!-- Yield Curve chart -->
          <div>
            <h3>📈 US Treasury Yield Curve</h3>
            <canvas id="yield-chart" height="140"></canvas>
            <div style="display:flex;justify-content:space-between;margin-top:8px">
              {% for pt in market_overview.yield_curve %}
              <div style="text-align:center">
                <div style="font-size:9px;color:#475569;font-weight:700;letter-spacing:.8px">{{ pt.label }}</div>
                <div style="font-size:13px;font-weight:700;color:#38bdf8">{{ pt.value }}%</div>
              </div>
              {% endfor %}
            </div>
          </div>

          <!-- Spot prices -->
          {% if market_overview.spot %}
          <div>
            <h3>💹 Commodities &amp; FX</h3>
            {% for label, data in market_overview.spot.items() %}
            <div class="spot-row">
              <span class="spot-label">{{ label }}</span>
              <div style="text-align:right">
                <div class="spot-val">{{ data.price }}</div>
                <div class="strip-chg {{ 'strip-up' if data.direction == 'up' else 'strip-down' }}"
                     style="font-size:11px">
                  {{ '+' if data.direction == 'up' else '' }}{{ data.change_pct }}%
                </div>
              </div>
            </div>
            {% endfor %}
          </div>
          {% endif %}

          <p style="font-size:10px;color:#334155;margin-top:4px">
            Rates sourced from CBOE/Treasury via Yahoo Finance. Delayed ~15 min.
          </p>
        </div>

      </div><!-- /economy-layout -->
    </div>
  </div><!-- /subtab-economy -->

  <!-- ══ SUB-TAB: EQUITY NEWS ══ -->
  <div id="subtab-equity" class="news-subtab-pane">

    <!-- Portfolio Holdings News -->
    <div class="section">
      <h2>📰 Portfolio Holdings — Latest News</h2>
      <p style="font-size:13px;color:#64748b;margin-bottom:24px">
        Recent headlines for the {{ portfolio_news|length }} stocks you currently hold.
      </p>
      <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(520px,1fr));gap:28px;align-items:start">
      {% for sym, articles in portfolio_news.items() %}
      <div>
        <div class="news-symbol-header" style="color:#60a5fa">
          {{ sym }}
          {% if holdings_map.get(sym) %}
          <span style="font-size:13px;color:#64748b;font-weight:400"> · {{ holdings_map[sym] }}</span>
          {% endif %}
        </div>
        {% if articles %}
        <div class="news-section-wrap">
          <ul class="news-list">
          {% for art in articles %}
          <li class="news-item">
            {% if art.thumbnail %}
            <img class="news-thumb" src="{{ art.thumbnail }}" alt=""
                 onerror="this.style.display='none'">
            {% endif %}
            <div class="news-body">
              <div class="news-source-row">
                <span class="news-publisher">{{ art.publisher }}</span>
                {% if art.age %}<span class="news-age">· {{ art.age }}</span>{% endif %}
              </div>
              <a class="news-title" href="{{ art.link }}" target="_blank" rel="noopener noreferrer">
                {{ art.title }}
              </a>
              {% if art.summary %}
              <div class="news-summary">{{ art.summary }}</div>
              {% endif %}
            </div>
          </li>
          {% endfor %}
          </ul>
        </div>
        {% else %}
        <div class="news-section-wrap" style="padding:16px 20px">
          <p style="color:#475569;font-size:13px">No recent news available.</p>
        </div>
        {% endif %}
      </div>
      {% endfor %}
      </div>
    </div>

    <!-- Recommended Stocks News -->
    <div class="section">
      <h2>📰 Recommended Stocks — Latest News</h2>
      <p style="font-size:13px;color:#64748b;margin-bottom:24px">
        Recent headlines for stocks on your buy list — spot catalysts and risks early.
      </p>
      <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(520px,1fr));gap:28px;align-items:start">
      {% for sym, articles in rec_news.items() %}
      <div>
        <div class="news-symbol-header" style="color:#a78bfa">
          {{ sym }}
          {% if rec_names_map.get(sym) %}
          <span style="font-size:13px;color:#64748b;font-weight:400"> · {{ rec_names_map[sym] }}</span>
          {% endif %}
        </div>
        {% if articles %}
        <div class="news-section-wrap">
          <ul class="news-list">
          {% for art in articles %}
          <li class="news-item">
            {% if art.thumbnail %}
            <img class="news-thumb" src="{{ art.thumbnail }}" alt=""
                 onerror="this.style.display='none'">
            {% endif %}
            <div class="news-body">
              <div class="news-source-row">
                <span class="news-publisher" style="color:#a78bfa">{{ art.publisher }}</span>
                {% if art.age %}<span class="news-age">· {{ art.age }}</span>{% endif %}
              </div>
              <a class="news-title" href="{{ art.link }}" target="_blank" rel="noopener noreferrer">
                {{ art.title }}
              </a>
              {% if art.summary %}
              <div class="news-summary">{{ art.summary }}</div>
              {% endif %}
            </div>
          </li>
          {% endfor %}
          </ul>
        </div>
        {% else %}
        <div class="news-section-wrap" style="padding:16px 20px">
          <p style="color:#475569;font-size:13px">No recent news available.</p>
        </div>
        {% endif %}
      </div>
      {% endfor %}
      </div>
    </div>

  </div><!-- /subtab-equity -->

  <footer>
    Portfolio Analyzer &nbsp;·&nbsp; News via Yahoo Finance &nbsp;·&nbsp;
    For informational purposes only. Not financial advice.
  </footer>
</div><!-- /tab-news -->

</div><!-- /app -->
</body>
</html>"""


def _encode_password(s: str) -> str:
    import base64, urllib.parse
    return base64.b64encode(urllib.parse.quote(s, safe='').encode()).decode()


def _validate_js(html: str) -> None:
    """Extract inline <script> blocks from html and validate each with esprima.
    Prints a clear error and the offending snippet if any block fails to parse."""
    try:
        import esprima, re
    except ImportError:
        return  # esprima not installed — skip silently

    blocks = re.findall(r'<script[^>]*>(.*?)</script>', html, re.DOTALL)
    ok = True
    for i, src in enumerate(blocks):
        src = src.strip()
        if not src:
            continue
        try:
            esprima.parseScript(src, tolerant=False)
        except esprima.Error as e:
            ok = False
            # Show the line that errored with a few lines of context
            lines = src.splitlines()
            lineno = getattr(e, 'lineNumber', None)
            print(f"\n  !! JS SYNTAX ERROR in script block {i+1}: {e}")
            if lineno and lineno <= len(lines):
                start = max(0, lineno - 3)
                end   = min(len(lines), lineno + 2)
                for ln, text in enumerate(lines[start:end], start + 1):
                    marker = " >>>" if ln == lineno else "    "
                    print(f"  {marker} {ln:4d}: {text}")
            print()
    if ok:
        print("  JS validation passed")


def render_html(portfolio_df, sector_df, sector_recs, thematic_recs, cash_plan, output_path,
                cash=DEFAULT_CASH, age=DEFAULT_AGE, password: str = "",
                portfolio_beta: float | None = None,
                portfolio_news: dict | None = None, rec_news: dict | None = None,
                economy_news: list | None = None,
                market_overview: dict | None = None,
                sp500_sector_data: dict | None = None,
                spy_benchmark: dict | None = None,
                etf_recs: list | None = None,
                etf_holdings: dict | None = None,
                holding_analysis: list | None = None,
                chart_data: dict | None = None):
    total_value = portfolio_df["Market Value"].sum()
    total_cost  = portfolio_df["Cost Basis"].sum() if portfolio_df["Cost Basis"].notna().any() else None
    total_gain  = ((total_value - total_cost) / total_cost * 100) if total_cost and total_cost > 0 else None

    def clean(r):
        row = r.to_dict() if hasattr(r, "to_dict") else dict(r)
        return {k: (None if isinstance(v, float) and (math.isnan(v) or math.isinf(v)) else v)
                for k, v in row.items()}

    holdings_records = [clean(r) for _, r in portfolio_df.iterrows()]
    sector_records   = [clean(r) for _, r in sector_df.iterrows()]

    # Per-stock budget = tier's total cash / number of recs in that tier
    all_recs = sector_recs + thematic_recs
    from collections import Counter
    tier_counts = Counter(r["risk_tier"] for r in all_recs)
    def tier_budget_for(rec):
        tier = rec.get("risk_tier", "medium")
        total = cash_plan.get(tier, {}).get("dollars", 0)
        count = tier_counts.get(tier, 1)
        return total / count if count else 0

    sector_recs   = [build_rec_context(r, tier_budget_for(r)) for r in sector_recs]
    thematic_recs = [build_rec_context(r, tier_budget_for(r)) for r in thematic_recs]
    ranked_recs   = rank_recommendations(sector_recs, thematic_recs)

    # Portfolio beta color
    if portfolio_beta is None:
        beta_color = "#64748b"
    elif portfolio_beta < 0.9:
        beta_color = "#34d399"
    elif portfolio_beta < 1.3:
        beta_color = "#60a5fa"
    elif portfolio_beta < 1.6:
        beta_color = "#fbbf24"
    else:
        beta_color = "#f87171"

    # ── Build compact stock-info map for hover tooltips ──────────────────────
    import json as _json
    def _clean_float(v):
        if v is None: return None
        try:
            f = float(v)
            return None if (math.isnan(f) or math.isinf(f)) else f
        except: return None

    stock_info = {}
    # Portfolio holdings
    for _, row in portfolio_df.iterrows():
        sym = row["Symbol"]
        price = _clean_float(row.get("Price"))
        mc    = _clean_float(row.get("Market Cap"))
        at    = _clean_float(row.get("Analyst Target"))
        upside = round((at - price) / price * 100, 1) if at and price else None
        stock_info[sym] = {
            "name": row.get("Name", sym),
            "sector": row.get("Sector") or row.get("ETF Category") or "",
            "price": price, "marketCap": mc,
            "pe": _clean_float(row.get("P/E")),
            "beta": _clean_float(row.get("Beta")),
            "analystTarget": at, "upside": upside,
            "rec": str(row.get("Recommendation") or "").replace("_", " ").title(),
            "isEtf": bool(row.get("is_etf")),
        }
    # Sector + thematic recs
    for r in (sector_recs or []) + (thematic_recs or []):
        sym = r.get("symbol") or r.get("Symbol")
        if not sym or sym in stock_info: continue
        p  = _clean_float(r.get("price"))
        at = _clean_float(r.get("analyst_target"))
        up = round((at - p) / p * 100, 1) if at and p else None
        stock_info[sym] = {
            "name": r.get("name", sym),
            "sector": r.get("sector", ""),
            "price": p, "marketCap": _clean_float(r.get("market_cap")),
            "pe": _clean_float(r.get("pe")),
            "beta": _clean_float(r.get("beta")),
            "analystTarget": at, "upside": up,
            "rec": str(r.get("recommendation") or "").replace("_", " ").title(),
            "isEtf": False,
        }
    # ETF recs
    for e in (etf_recs or []):
        sym = e.get("symbol")
        if not sym or sym in stock_info: continue
        stock_info[sym] = {
            "name": e.get("name", sym),
            "sector": e.get("etf_category") or "ETF",
            "price": _clean_float(e.get("price")),
            "marketCap": _clean_float(e.get("aum")),
            "pe": None, "beta": _clean_float(e.get("beta")),
            "analystTarget": None, "upside": None, "rec": "",
            "isEtf": True,
            "expenseRatio": _clean_float(e.get("expense_ratio")),
            "threeYr": _clean_float(e.get("three_yr")),
        }

    # Build upcoming earnings list (next 30 days) from holdings + recs
    _earn_seen = set()
    _earn_upcoming = []
    for _h in (holding_analysis or []):
        _dte = _h.get("days_to_earnings")
        if _dte is not None and 0 <= _dte <= 30 and _h["symbol"] not in _earn_seen:
            _earn_seen.add(_h["symbol"])
            _earn_upcoming.append({"sym": _h["symbol"], "days": int(_dte), "date": _h.get("earnings_date") or ""})
    for _r in sector_recs + thematic_recs:
        _dte = _r.get("days_to_earnings")
        if _dte is not None and 0 <= _dte <= 30 and _r["symbol"] not in _earn_seen:
            _earn_seen.add(_r["symbol"])
            _earn_upcoming.append({"sym": _r["symbol"], "days": int(_dte), "date": _r.get("earnings_date") or ""})
    _earn_upcoming.sort(key=lambda x: x["days"])

    tmpl = Template(HTML_TEMPLATE)
    html = tmpl.render(
        date=datetime.now().strftime("%B %d, %Y"),
        n_stocks=len(portfolio_df),
        n_sectors=len(sector_df),
        total_value=total_value,
        total_cost=total_cost,
        total_gain=total_gain,
        holdings=holdings_records,
        sectors=sector_records,
        sector_recs=sector_recs,
        thematic_recs=thematic_recs,
        cash_plan=cash_plan,
        cash=cash,
        age=age,
        risk_tier_color=RISK_TIER_COLOR,
        risk_tier_label=RISK_TIER_LABEL,
        metric_explanations=METRIC_EXPLANATIONS,
        pass_b64=_encode_password(password) if password else "none",
        ranked_recs=ranked_recs,
        portfolio_beta=portfolio_beta,
        beta_color=beta_color,
        portfolio_news=portfolio_news or {},
        rec_news=rec_news or {},
        economy_news=economy_news or [],
        market_overview=market_overview or {"strip": [], "sectors": [], "yield_curve": [], "spot": {}},
        sp500_sector_json=__import__("json").dumps(sp500_sector_data or {}),
        spy_benchmark=spy_benchmark or {},
        etf_recs=etf_recs or [],
        etf_holdings_json=__import__("json").dumps(etf_holdings or {}),
        holdings_map={r["Symbol"]: r["Name"] for _, r in portfolio_df.iterrows()},
        holding_map={h["symbol"]: h for h in (holding_analysis or [])},
        rec_names_map={r["symbol"]: r["name"] for r in sector_recs + thematic_recs},
        holding_analysis=holding_analysis or [],
        total_dollar_gain=(total_value - total_cost) if total_cost else None,
        chart_data_json=__import__("json").dumps(chart_data or {}),
        stock_info_json=_json.dumps(stock_info),
        earnings_json=_json.dumps(_earn_upcoming),
    )
    Path(output_path).write_text(html, encoding="utf-8")
    print(f"  HTML report: {output_path}")
    _validate_js(html)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def render_excel(portfolio_df, sector_df, sector_recs, thematic_recs, cash_plan, output_path):
    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    pos_font    = Font(color="34D399", bold=True)
    neg_font    = Font(color="F87171", bold=True)
    thin_border = Border(bottom=Side(style="thin", color="334155"))

    RISK_FILLS = {
        "low":         PatternFill("solid", fgColor="064E3B"),
        "medium":      PatternFill("solid", fgColor="1E3A5F"),
        "high":        PatternFill("solid", fgColor="451A03"),
        "speculative": PatternFill("solid", fgColor="450A0A"),
    }

    # ---- Holdings sheet ----
    ws = wb.active
    ws.title = "Holdings"
    cols = ["Symbol","Name","Sector","Shares","Price","Market Value","Cost Basis",
            "Gain/Loss %","Beta","Risk Tier","P/E","Fwd P/E","Div Yield %",
            "EPS Growth %","Rev Growth %","Growth Score","Analyst Target","Analyst Upside %","Recommendation"]
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(1, ci, c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, (_, row) in enumerate(portfolio_df.iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            val = row.get(col)
            if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                val = None
            cell = ws.cell(ri, ci, val)
            cell.border = thin_border
            if col in ("Gain/Loss %","Analyst Upside %","EPS Growth %","Rev Growth %") and val is not None:
                cell.font = pos_font if val >= 0 else neg_font
            if col in ("Price","Market Value","Cost Basis","Analyst Target"):
                cell.number_format = '$#,##0.00'
            if col == "Shares": cell.number_format = '#,##0'
            if col == "Risk Tier" and val:
                cell.fill = RISK_FILLS.get(val, header_fill)
                cell.font = Font(color="FFFFFF", bold=True)
    for ci in range(1, len(cols)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # ---- Sector sheet ----
    ws2 = wb.create_sheet("Sector Allocation")
    scols = ["Sector","Market Value","Portfolio Weight %","S&P 500 Weight %","Over/Under %"]
    for ci, c in enumerate(scols, 1):
        cell = ws2.cell(1, ci, c); cell.fill = header_fill; cell.font = header_font
    for ri, (_, row) in enumerate(sector_df.iterrows(), 2):
        for ci, col in enumerate(scols, 1):
            val = row.get(col)
            cell = ws2.cell(ri, ci, val)
            if col == "Market Value": cell.number_format = '$#,##0'
            if "%" in col and val is not None: cell.number_format = "0.0"
            if col == "Over/Under %" and val is not None:
                cell.font = pos_font if val >= 0 else neg_font
    for ci in range(1, len(scols)+1):
        ws2.column_dimensions[get_column_letter(ci)].width = 20
    chart = BarChart()
    chart.type = "col"; chart.title = "Portfolio vs S&P 500 Sector Weights"
    chart.y_axis.title = "Weight (%)"; chart.style = 10; chart.width = 24; chart.height = 14
    n = len(sector_df) + 1
    chart.add_data(Reference(ws2, min_col=3, max_col=4, min_row=1, max_row=n), titles_from_data=True)
    chart.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=n))
    ws2.add_chart(chart, "G2")

    # ---- Recommendations sheet ----
    ws3 = wb.create_sheet("Recommendations")
    rcols = ["Category","Type","Gap vs S&P %","Symbol","Name","In S&P 500","Risk Tier",
             "Price","Beta","P/E","Fwd P/E","EPS Growth %","Rev Growth %","Growth Score",
             "Analyst Target","Upside %","Rating"]
    for ci, c in enumerate(rcols, 1):
        cell = ws3.cell(1, ci, c); cell.fill = header_fill; cell.font = header_font
    all_recs = [(r, "Sector") for r in sector_recs] + [(r, "Thematic") for r in thematic_recs]
    for ri, (rec, rtype) in enumerate(all_recs, 2):
        vals = [
            rec.get("category"), rtype, rec.get("gap_pct"),
            rec.get("symbol"), rec.get("name"),
            "Yes" if rec.get("in_sp500") else "No",
            rec.get("risk_tier"), rec.get("price"), rec.get("beta"),
            rec.get("pe"), rec.get("forward_pe"),
            rec.get("eps_growth"), rec.get("rev_growth"), rec.get("growth_score"),
            rec.get("analyst_target"), rec.get("upside"), rec.get("recommendation"),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws3.cell(ri, ci, val)
            if ci == 8 and val: cell.number_format = '$#,##0.00'
            if ci == 15 and val: cell.number_format = '$#,##0.00'
            if ci == 16 and val is not None:
                cell.font = pos_font if val >= 0 else neg_font
            if ci == 7 and val:
                cell.fill = RISK_FILLS.get(val, header_fill)
                cell.font = Font(color="FFFFFF", bold=True)
    for ci in range(1, len(rcols)+1):
        ws3.column_dimensions[get_column_letter(ci)].width = 17

    # ---- Cash Allocation Plan sheet ----
    ws4 = wb.create_sheet("Cash Allocation Plan")
    descs = {
        "low":         "Blue-chip S&P stocks — stable anchor",
        "medium":      "Growth S&P stocks — solid upside",
        "high":        "High-beta & small-cap growth plays",
        "speculative": "Moonshot bets — high risk, high reward",
    }
    for ci, h in enumerate(["Risk Tier","Allocation %","Suggested $","Description"], 1):
        cell = ws4.cell(1, ci, h); cell.fill = header_fill; cell.font = header_font
    for ri, (tier, info) in enumerate(cash_plan.items(), 2):
        ws4.cell(ri, 1, RISK_TIER_LABEL.get(tier, tier))
        ws4.cell(ri, 1).fill = RISK_FILLS.get(tier, header_fill)
        ws4.cell(ri, 1).font = Font(color="FFFFFF", bold=True)
        ws4.cell(ri, 2, info["pct"]).number_format = '0.0"%"'
        ws4.cell(ri, 3, info["dollars"]).number_format = '$#,##0'
        ws4.cell(ri, 4, descs.get(tier, ""))
    for ci in range(1, 5):
        ws4.column_dimensions[get_column_letter(ci)].width = 26

    wb.save(output_path)
    print(f"  Excel report: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Stock Portfolio Analyzer")
    parser.add_argument("portfolio", help="Path to portfolio CSV")
    parser.add_argument("--cash",   type=float, default=DEFAULT_CASH,
                        help=f"Cash available to invest (default ${DEFAULT_CASH:,.0f})")
    parser.add_argument("--age",    type=int,   default=DEFAULT_AGE,
                        help=f"Investor age (default {DEFAULT_AGE})")
    parser.add_argument("--risk",     choices=["aggressive","moderate","conservative"],
                        default=RISK_PROFILE, help="Risk profile (default aggressive)")
    parser.add_argument("--password", type=str, default="",
                        help="Password-protect the HTML report (optional)")
    args = parser.parse_args()

    csv_path = args.portfolio
    if not os.path.exists(csv_path):
        print(f"Error: file not found: {csv_path}")
        sys.exit(1)

    stem      = Path(csv_path).stem
    out_html  = str(Path(csv_path).parent / f"{stem}_report.html")
    out_excel = str(Path(csv_path).parent / f"{stem}_report.xlsx")

    print(f"\nPortfolio Analyzer")
    print(f"{'='*40}")
    print(f"Input:        {csv_path}")
    print(f"Cash to deploy: ${args.cash:,.0f}")
    print(f"Age:          {args.age}  |  Risk profile: {args.risk}")

    print("\n[1/5] Loading portfolio...")
    portfolio = load_portfolio(csv_path)
    print(f"  Found {len(portfolio)} positions: {', '.join(portfolio['symbol'])}")

    print("\n[2/5] Fetching market data...")
    stock_data = fetch_stock_data(portfolio["symbol"].tolist())

    print("\n[3/5] Building portfolio analysis...")
    portfolio_df = build_portfolio_df(portfolio, stock_data)
    sector_df    = sector_analysis(portfolio_df)

    print("\n[4/5] Generating recommendations...")
    held = set(portfolio["symbol"])
    sector_recs, thematic_recs, cash_plan = generate_recommendations(
        sector_df, held, stock_data, cash=args.cash, risk_profile=args.risk
    )
    print(f"  Sector recs:   {len(sector_recs)} picks across "
          f"{len(set(r['category'] for r in sector_recs))} underweight sectors")
    print(f"  Thematic recs: {len(thematic_recs)} picks across "
          f"{len(set(r['category'] for r in thematic_recs))} themes")

    portfolio_syms = portfolio["symbol"].tolist()
    rec_syms       = list({r["symbol"] for r in sector_recs + thematic_recs})
    all_extended   = list(set(portfolio_syms + rec_syms))

    print(f"\n[4b/5] Fetching momentum & earnings data ({len(all_extended)} tickers)...")
    fetch_extended_data(all_extended, stock_data)
    # Patch extended fields back into rec dicts
    for rec in sector_recs + thematic_recs:
        d = stock_data.get(rec["symbol"], {})
        rec.update({k: d.get(k) for k in (
            "mom_3m", "mom_6m", "earnings_date", "days_to_earnings",
            "eps_est_next", "last_eps_actual", "last_eps_estimate",
            "last_eps_surprise_pct", "last_q_period",
        )})

    portfolio_beta   = compute_portfolio_beta(portfolio_df)
    holding_analysis = analyze_holdings(portfolio_df, stock_data, held)
    sells  = sum(1 for h in holding_analysis if h["verdict"] == "SELL")
    watches = sum(1 for h in holding_analysis if h["verdict"] == "WATCH")
    print(f"  Portfolio beta: {portfolio_beta if portfolio_beta else 'N/A'}")
    print(f"  Holding review: {sells} SELL · {watches} WATCH · "
          f"{len(holding_analysis)-sells-watches} HOLD")

    print(f"\n[4c/5] Fetching chart data for all {len(all_extended)} stocks...")
    print("  Scoring ETF recommendations...")
    etf_recs        = fetch_etf_recs(held_symbols=set(portfolio_syms))
    # Watchlist
    etf_syms        = [e["symbol"] for e in etf_recs]
    chart_symbols   = list(set(all_extended + etf_syms + ["SPY"]))
    chart_data      = fetch_chart_data(chart_symbols)
    print("  Fetching ETF holdings...")
    etf_holdings    = fetch_etf_holdings(etf_syms)

    print(f"\n[4d/5] Fetching news + market overview...")
    portfolio_news  = fetch_news(portfolio_syms)
    rec_news        = fetch_news(rec_syms)
    economy_news    = fetch_economy_news(max_articles=40)
    print("  Fetching market strip / sector heatmap / yield curve...")
    market_overview = fetch_market_overview()
    print("  Fetching S&P 500 sector drill-down data...")
    sp500_sector_data = fetch_sp500_sector_data()
    print("  Fetching SPY benchmark returns...")
    spy_benchmark   = fetch_spy_benchmark()

    print("\n[5/5] Writing reports...")
    render_html(portfolio_df, sector_df, sector_recs, thematic_recs, cash_plan,
                out_html, cash=args.cash, age=args.age, password=args.password,
                portfolio_beta=portfolio_beta,
                portfolio_news=portfolio_news, rec_news=rec_news,
                economy_news=economy_news, market_overview=market_overview,
                sp500_sector_data=sp500_sector_data,
                spy_benchmark=spy_benchmark, etf_recs=etf_recs, etf_holdings=etf_holdings,
                holding_analysis=holding_analysis, chart_data=chart_data)
    render_excel(portfolio_df, sector_df, sector_recs, thematic_recs, cash_plan, out_excel)

    total = portfolio_df["Market Value"].sum()
    print(f"\n{'='*40}")
    print(f"Done! Portfolio value: ${total:,.2f}")
    print(f"Cash to deploy:       ${args.cash:,.2f}")
    print(f"Open {out_html} in your browser for the full dashboard.")


if __name__ == "__main__":
    main()
